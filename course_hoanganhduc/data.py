# -*- coding: utf-8 -*-
# Canvas API: https://canvas.instructure.com/doc/api/
# Course Management Script

import pandas as pd
import os
import pickle
try:
    import readline
except ImportError:  # Windows without pyreadline installed
    readline = None
import sys
import argparse
import re
import glob
from datetime import datetime
import openpyxl
import pytesseract
from pdf2image import convert_from_path
from PIL import Image, ImageOps, ImageFilter
import requests
import tempfile
import base64
import PyPDF2
import io
import shlex
from collections import defaultdict
import difflib
from openpyxl.styles import Alignment
import unicodedata
import numpy as np
from paddleocr import PaddleOCR
import json
import shutil
from canvasapi import Canvas
from datetime import datetime, timedelta, timezone
from itertools import combinations
import signal
import platform
from pathlib import Path
import time
import random
import traceback
import cv2
import subprocess
import csv
import copy
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import hashlib
import math
import statistics
from collections import Counter, OrderedDict
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from types import SimpleNamespace

from tqdm import tqdm  # <-- Add tqdm for progress bars

from .version import __version__

from .settings import *
from .config import *
from .models import *
from .utils import *

AI_LAST_MODEL_USED = {"provider": None, "model": None}
_AI_MODEL_CACHE = {}

def clean_excel_data(df, verbose=False):
    """
    Remove duplicate columns and rows for Student ID, Email, Name.
    Also remove empty rows and columns (columns with only header or all NaN/empty).
    Returns a cleaned DataFrame.
    If verbose is True, print more details; otherwise, print only important notice.
    """
    # Normalize columns for easier processing
    df = normalize_columns(df, verbose=verbose)
    if verbose:
        print("[CleanExcel] Normalized columns:", list(df.columns))

    # Remove duplicate rows based on Student ID, Email, Name (keep first occurrence)
    subset_cols = [col for col in ["Student ID", "Email", "Name"] if col in df.columns]
    if subset_cols:
        before = len(df)
        df = df.drop_duplicates(subset=subset_cols, keep='first')
        after = len(df)
        if verbose:
            print(f"[CleanExcel] Removed {before - after} duplicate rows based on {subset_cols}")
        elif before - after > 0:
            print(f"Notice: Removed {before - after} duplicate rows.")

    # Remove empty rows (all columns empty or NaN)
    before = len(df)
    df = df.dropna(how='all')
    df = df[~df.apply(lambda row: all(str(x).strip() == "" or pd.isna(x) for x in row), axis=1)]
    after = len(df)
    if verbose:
        print(f"[CleanExcel] Removed {before - after} empty rows.")
    elif before - after > 0:
        print(f"Notice: Removed {before - after} empty rows.")

    # Remove empty columns (all values empty or NaN except header)
    def is_empty_col(col):
        return all(str(x).strip() == "" or pd.isna(x) for x in df[col])

    empty_cols = [col for col in df.columns if is_empty_col(col)]
    if empty_cols:
        df = df.drop(columns=empty_cols)
        if verbose:
            print(f"[CleanExcel] Removed empty columns: {empty_cols}")
        else:
            print(f"Notice: Removed {len(empty_cols)} empty columns.")

    return df

def normalize_columns(df, verbose=False):
    """
    Normalize column names in the DataFrame to standard English names.
    Keep only one column for: Name, Email, Student ID.
    Prefer columns with non-empty, non-anonymous values.
    For Student ID, only keep columns where values are 8-digit numbers.
    Also, reformat names of the form "<first name>, <family name>" to "<family name> <first name>".
    Grades columns (e.g., Attendance, Midterm, Final, Participation, Assignment, Quiz) are also mapped and kept.
    Handles Canvas-style grade columns like "Bài tập Final Score", "Kiểm tra giữa kỳ Final Score", etc.
    If verbose is True, print more details; otherwise, print only important notice.
    """
    col_map = {}
    # Mapping from Vietnamese to English (capitalize right-hand side)
    mapping = {
        # GitHub username mappings (specific)
        "Github username": "GitHub Username",
        "Github account": "GitHub Username",
        "Github id": "GitHub Username",
        "Github handle": "GitHub Username",
        "GitHub username": "GitHub Username",
        "GitHub account": "GitHub Username",
        "GitHub id": "GitHub Username",
        "GitHub handle": "GitHub Username",
        "Github": "GitHub Username",
        "GitHub": "GitHub Username",
        
        # Registration/Section ID mappings - all mapped to Canvas Section
        "Registered Class ID": "Canvas Section",
        "Registration ID": "Canvas Section",
        "Registered Class": "Canvas Section",
        "Canvas Section": "Canvas Section",
        "Section ID": "Canvas Section",
        "Lớp học phần": "Canvas Section",
        "Mã Section": "Canvas Section",
        "Mã lớp đăng ký": "Canvas Section",
        
        # Email mappings (including specific Google Classroom variants)
        "Email": "Email",
        "Email sử dụng để đăng ký lớp Google Classroom": "Email",
        "Email google classroom": "Email",
        "Email đăng ký google classroom": "Email",
        "Email google": "Email",
        
        # Google Classroom Display Name mapping
        "Tên Hiển Thị Trên Google Classroom": "Google Classroom Display Name",
        
        # Regular mappings
        "Họ và Tên": "Name",
        "Họ tên": "Name",
        "Tên": "Name",
        "Mã sinh viên": "Student ID",
        "Mã SV": "Student ID",
        "Mã số sinh viên": "Student ID",
        "MSSV": "Student ID",
        "Lớp": "Class",
        "Số điện thoại": "Phone",
        "Ngày sinh": "Dob",
        "Giới tính": "Gender",
        "Student": "Name",
        "Name": "Name",
        "Full Name": "Name",
        "Class": "Class",
        "Phone": "Phone",
        "Gender": "Gender",
        "Date of Birth": "Dob",
        # Grades (full names)
        "Chuyên cần": "Attendance",
        "CC": "Attendance",
        "Điểm chuyên cần": "Attendance",
        "Attendance": "Attendance",
        "Giữa kỳ": "Midterm",
        "GK": "Midterm",
        "Điểm giữa kỳ": "Midterm",
        "Midterm": "Midterm",
        "Cuối kỳ": "Final",
        "CK": "Final",
        "Điểm cuối kỳ": "Final",
        "Final": "Final",
        # Additional grades
        "Điểm danh": "Participation",
        "DD": "Participation",
        "Bài Tập": "Assignment",
        "BT": "Assignment",
        "Quiz": "Quiz",
    }
    # Grades columns to keep (main categories, full names)
    grade_cols = {"Attendance", "Midterm", "Final", "Participation", "Assignment", "Quiz"}

    # Canvas-style grade column patterns
    canvas_patterns = [
        (r"bài[\s_]*tập.*final score", "Assignment"),
        (r"kiểm[\s_]*tra[\s_]*giữa[\s_]*kỳ.*final score", "Midterm"),
        (r"giữa[\s_]*kỳ.*final score", "Midterm"),
        (r"cuối[\s_]*kỳ.*final score", "Final"),
        (r"midterm.*score", "Midterm"),
        (r"attendance.*final score", "Attendance"),
        (r"chuyên[\s_]*cần.*final score", "Attendance"),
        (r"quiz.*final score", "Quiz"),
        (r"điểm[\s_]*danh.*final score", "Participation"),
    ]

    for col in df.columns:
        mapped = False
        # Try direct mapping
        for vn, en in mapping.items():
            if vn.lower() == col.lower():  # Exact match for GitHub and Registration fields
                col_map[col] = en
                mapped = True
                break
            elif not any(term in vn.lower() for term in ["github", "registered", "registration", "section", "canvas"]) and vn.lower() in col.lower():
                # For other fields, use partial match but exclude GitHub and Registration/Section terms to avoid wrong mappings
                col_map[col] = en
                mapped = True
                break
        if mapped:
            continue
        
        # Try GitHub patterns specifically (ensure they don't get mapped to Name)
        if re.search(r"github|git\s+username", col.lower()) and not re.search(r"name.*github", col.lower()):
            col_map[col] = "GitHub Username"
            mapped = True
            continue
            
        # Try Registration/Section patterns specifically - all map to Canvas Section
        if re.search(r"registered\s*class\s*id|registration\s*id|class\s*registration|section\s*id|canvas\s*section|mã\s*lớp", col.lower()):
            col_map[col] = "Canvas Section"
            mapped = True
            continue
            
        # Try Email patterns specifically (including Google Classroom variants)
        if re.search(r"email.*google.*classroom|email.*đăng.*ký.*google|email.*google", col.lower()):
            col_map[col] = "Email"
            mapped = True
            continue
            
        # Try Canvas-style patterns
        for pat, en in canvas_patterns:
            if re.search(pat, col.lower()):
                col_map[col] = en
                mapped = True
                break
        if mapped:
            continue

    if col_map:
        if verbose:
            print(f"[NormalizeColumns] Column mapping: {col_map}")
        else:
            print(f"Notice: {len(col_map)} columns mapped to standard names.")
        df = df.rename(columns=col_map)

    # Helper to check for anonymous-like values
    def is_anonymous(val):
        if not isinstance(val, str):
            return False
        v = val.strip().lower()
        return v in {"", "anonymous", "n/a", "na", "none", "unk", "unknown", "no name", "không tên", "chưa rõ"}

    # Helper to check for valid 8-digit student id (must be integer, not float)
    def is_valid_student_id(val):
        if isinstance(val, float):
            # Reject floats (e.g., 12345678.0)
            return False
        if not isinstance(val, str):
            val = str(val)
        val = val.strip()
        # Reject if contains a decimal point
        if '.' in val:
            return False
        return bool(re.fullmatch(r"\d{8}", val))

    # For each of Name, Email, Student ID: keep the column with most non-anonymous, non-empty values
    std_cols = ["Name", "Email", "Student ID"]
    cols_to_keep = []
    for std_col in std_cols:
        candidates = [col for col in df.columns if col.lower() == std_col.lower()]
        if not candidates:
            continue
        if std_col == "Student ID":
            # Only consider columns where values are 8-digit numbers
            best_col = None
            max_valid = -1
            for col in candidates:
                valid_count = df[col].apply(lambda x: is_valid_student_id(x)).sum()
                if valid_count > max_valid:
                    max_valid = valid_count
                    best_col = col
            if best_col and max_valid > 0:
                cols_to_keep.append(best_col)
                if verbose:
                    print(f"[NormalizeColumns] Keeping Student ID column: {best_col} (valid IDs: {max_valid})")
            elif verbose:
                print(f"[NormalizeColumns] No valid Student ID column found among: {candidates}")
        else:
            if len(candidates) == 1:
                cols_to_keep.append(candidates[0])
                if verbose:
                    print(f"[NormalizeColumns] Keeping {std_col} column: {candidates[0]}")
            else:
                best_col = None
                max_valid = -1
                for col in candidates:
                    valid_count = df[col].apply(lambda x: not is_anonymous(str(x))).sum()
                    if valid_count > max_valid:
                        max_valid = valid_count
                        best_col = col
                if best_col:
                    cols_to_keep.append(best_col)
                    if verbose:
                        print(f"[NormalizeColumns] Keeping {std_col} column: {best_col} (non-anonymous: {max_valid})")
    # Add grades columns (Attendance, Midterm, Final, Participation, Assignment, Quiz) if present
    for grade_col in grade_cols:
        candidates = [col for col in df.columns if col == grade_col]
        if candidates:
            cols_to_keep.extend([c for c in candidates if c not in cols_to_keep])
            if verbose and candidates:
                print(f"[NormalizeColumns] Keeping grade column(s): {candidates}")
    # Add other columns that are not std_cols or grades
    for col in df.columns:
        if col not in cols_to_keep and col.lower() not in [c.lower() for c in std_cols] and col not in grade_cols:
            cols_to_keep.append(col)
            if verbose:
                print(f"[NormalizeColumns] Keeping extra column: {col}")
    df = df.loc[:, cols_to_keep]

    # Reformat names of the form "<first name>, <family name>" to "<family name> <first name>"
    if "Name" in df.columns:
        def reformat_name(name):
            if not isinstance(name, str):
                return name
            parts = [p.strip() for p in name.split(",")]
            if len(parts) == 2 and all(parts):
                # "<first name>, <family name>" -> "<family name> <first name>"
                name = f"{parts[1]} {parts[0]}"
            # Apply titlecase to ensure proper capitalization (e.g., NGUYỄN VĂN AN -> Nguyễn Văn An)
            return name.title()
        df["Name"] = df["Name"].apply(reformat_name)
        if verbose:
            print("[NormalizeColumns] Reformatted names with comma to standard format and applied titlecase.")

    return df

def _get_retry_after_seconds(response):
    retry_after = response.headers.get("Retry-After")
    if retry_after:
        try:
            return int(float(retry_after))
        except Exception:
            return None
    return None


def _sleep_with_backoff(attempt, base_seconds=5, retry_after=None):
    sleep_seconds = retry_after if retry_after is not None else base_seconds * attempt
    time.sleep(sleep_seconds)


def _is_rate_limited(response=None, error_text=""):
    if response is not None and response.status_code in (429, 503):
        return True
    if error_text and "rate" in error_text.lower():
        return True
    return False


def _set_last_ai_model(provider, model):
    AI_LAST_MODEL_USED["provider"] = provider
    AI_LAST_MODEL_USED["model"] = model


def _reset_last_ai_model():
    AI_LAST_MODEL_USED["provider"] = None
    AI_LAST_MODEL_USED["model"] = None


def get_last_ai_model_used():
    provider = AI_LAST_MODEL_USED.get("provider")
    model = AI_LAST_MODEL_USED.get("model")
    if not provider:
        return ""
    if model:
        return f"{provider} ({model})"
    return provider


def _normalize_gemini_model_name(name):
    if not name:
        return ""
    value = str(name).strip()
    if value.startswith("models/"):
        return value.split("/", 1)[1]
    return value


def _get_cached_models(provider, target_capability=None):
    cached = _AI_MODEL_CACHE.get(provider)
    if not cached:
        return []
    models = cached.get("models", []) or []
    if target_capability:
        model_caps = cached.get("capabilities", {}) or {}
        models = [m for m in models if target_capability in (model_caps.get(m) or [])]
    return models


def _update_model_cache(provider, models, capabilities=None):
    _AI_MODEL_CACHE[provider] = {
        "models": models,
        "capabilities": capabilities or {},
        "updated_at": time.time(),
    }


def _pick_fallback_model(provider, current_model=None, target_capability=None):
    models = _get_cached_models(provider, target_capability=target_capability)
    if not models:
        listed = list_ai_models(provider, verbose=False)
        models = (listed.get(provider, {}) or {}).get("models", []) if isinstance(listed, dict) else []
    if not models:
        return ""
    candidates = [m for m in models if m and m != current_model]
    if not candidates:
        return ""
    return random.choice(candidates)


def refine_text_with_gemini(text, api_key=None, user_prompt=None, verbose=False):
    """
    Use Gemini API to refine OCR-extracted text while preserving layout.
    Corrects OCR errors, improves readability, but maintains the original structure.

    Args:
        text: The OCR-extracted text to refine
        api_key: Gemini API key (uses global GEMINI_API_KEY if not provided)
        user_prompt: Optional custom prompt to use instead of the default
        verbose: If True, print more details; otherwise, print only important notice

    Returns:
        Refined text with improved accuracy while preserving layout
    """
    if not api_key:
        api_key = GEMINI_API_KEY

    if not api_key:
        if verbose:
            print("No Gemini API key provided. Returning original text.")
        else:
            print("Notice: No Gemini API key provided.")
        return text

    # Split text into smaller chunks if too large (Gemini has input limits)
    max_chunk_size = 30000  # Conservative limit
    if len(text) <= max_chunk_size:
        chunks = [text]
    else:
        paragraphs = text.split('\n\n')
        chunks = []
        current_chunk = ""
        for paragraph in paragraphs:
            if len(current_chunk) + len(paragraph) + 2 <= max_chunk_size:
                if current_chunk:
                    current_chunk += '\n\n' + paragraph
                else:
                    current_chunk = paragraph
            else:
                if current_chunk:
                    chunks.append(current_chunk)
                current_chunk = paragraph
        if current_chunk:
            chunks.append(current_chunk)

    refined_chunks = []

    max_attempts = 3
    base_sleep_seconds = 5

    for i, chunk in enumerate(chunks):
        if verbose:
            print(f"[Gemini] Refining chunk {i+1}/{len(chunks)} with Gemini API...")
        else:
            print(f"Refining chunk {i+1}/{len(chunks)} with Gemini API...")

        if user_prompt:
            prompt = user_prompt.format(text=chunk)
        else:
            prompt = f"""You are an expert text correction assistant. Please refine the following OCR-extracted text by:

1. Correcting OCR errors (misrecognized characters, words)
2. Fixing obvious spelling mistakes
3. Improving readability while preserving the original meaning
4. IMPORTANT: Maintain the exact layout, line breaks, and spacing structure
5. Keep all dates, numbers, and proper nouns as accurate as possible
6. If the text appears to be in Vietnamese, correct Vietnamese-specific OCR errors
7. If text contains student names, student IDs (8-digit numbers), or dates (d/m/yyyy format), pay special attention to accuracy

Please return ONLY the corrected text without any explanations or additional comments.

Text to refine:
{chunk}"""

        refined_text = None
        model_name = _normalize_gemini_model_name(GEMINI_DEFAULT_MODEL)
        if not model_name:
            model_name = GEMINI_DEFAULT_MODEL
        for attempt in range(1, max_attempts + 1):
            try:
                url = f"https://generativelanguage.googleapis.com/v1beta/models/{model_name}:generateContent?key={api_key}"
                headers = {'Content-Type': 'application/json'}
                data = {
                    "contents": [
                        {
                            "parts": [
                                {
                                    "text": prompt
                                }
                            ]
                        }
                    ]
                }
                response = requests.post(url, headers=headers, json=data, timeout=120)
                if _is_rate_limited(response=response):
                    retry_after = _get_retry_after_seconds(response)
                    if verbose:
                        print(f"[Gemini] Rate limited for chunk {i+1}. Retry {attempt}/{max_attempts} after {retry_after or base_sleep_seconds * attempt}s.")
                    fallback_model = _pick_fallback_model("gemini", current_model=model_name, target_capability="generateContent")
                    if fallback_model:
                        model_name = _normalize_gemini_model_name(fallback_model)
                        if verbose:
                            print(f"[Gemini] Switching to fallback model: {model_name}")
                    if attempt < max_attempts:
                        _sleep_with_backoff(attempt, base_seconds=base_sleep_seconds, retry_after=retry_after)
                        continue
                response.raise_for_status()
                result = response.json()
                if 'candidates' in result and len(result['candidates']) > 0:
                    refined_text = result['candidates'][0]['content']['parts'][0]['text'].strip()
                    _set_last_ai_model("gemini", model_name)
                    if verbose:
                        print(f"[Gemini] Chunk {i+1} refined successfully.")
                else:
                    if verbose:
                        print(f"[Gemini] No response from Gemini API for chunk {i+1}. Using original text.")
                    else:
                        print(f"Notice: No response from Gemini API for chunk {i+1}.")
                    refined_text = chunk
                break
            except Exception as e:
                if verbose:
                    print(f"[Gemini] Error calling Gemini API for chunk {i+1} (attempt {attempt}/{max_attempts}): {e}")
                else:
                    print(f"Notice: Error calling Gemini API for chunk {i+1}.")
                fallback_model = _pick_fallback_model("gemini", current_model=model_name, target_capability="generateContent")
                if fallback_model:
                    model_name = _normalize_gemini_model_name(fallback_model)
                    if verbose:
                        print(f"[Gemini] Switching to fallback model: {model_name}")
                if attempt < max_attempts:
                    _sleep_with_backoff(attempt, base_seconds=base_sleep_seconds)
                    continue
                refined_text = chunk
        refined_chunks.append(refined_text if refined_text is not None else chunk)

    refined_text = '\n\n'.join(refined_chunks)
    return refined_text

def refine_text_with_huggingface(
    text, 
    api_key=HUGGINGFACE_API_KEY, 
    model="meta-llama/llama-3.1-8b-instruct",
    user_prompt=None,
    verbose=False
):
    """
    Use Hugging Face Chat API (novita endpoint) to refine OCR-extracted text while preserving layout.
    Corrects OCR errors, improves readability, but maintains the original structure.

    Args:
        text: The OCR-extracted text to refine
        api_key: Hugging Face API key (required)
        model: Model name on Hugging Face Hub (default: meta-llama/llama-3.1-8b-instruct)
        user_prompt: Optional custom prompt to use instead of the default
        verbose: If True, print more details; otherwise, print only important notice

    Returns:
        Refined text with improved accuracy while preserving layout
    """
    if not api_key:
        if verbose:
            print("No Hugging Face API key provided. Returning original text.")
        else:
            print("Notice: No Hugging Face API key provided.")
        return text

    url = "https://router.huggingface.co/novita/v3/openai/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }

    # Split text into chunks if too large (Hugging Face has input limits)
    max_chunk_size = 3000
    if len(text) <= max_chunk_size:
        chunks = [text]
    else:
        paragraphs = text.split('\n\n')
        chunks = []
        current_chunk = ""
        for paragraph in paragraphs:
            if len(current_chunk) + len(paragraph) + 2 <= max_chunk_size:
                if current_chunk:
                    current_chunk += '\n\n' + paragraph
                else:
                    current_chunk = paragraph
            else:
                if current_chunk:
                    chunks.append(current_chunk)
                current_chunk = paragraph
        if current_chunk:
            chunks.append(current_chunk)

    refined_chunks = []
    max_attempts = 3
    base_sleep_seconds = 5
    for i, chunk in enumerate(chunks):
        if verbose:
            print(f"[HuggingFace] Refining chunk {i+1}/{len(chunks)} with Hugging Face Chat API (novita endpoint)...")
        else:
            print(f"Refining chunk {i+1}/{len(chunks)} with Hugging Face Chat API (novita endpoint)...")
        if user_prompt:
            prompt = user_prompt.format(text=chunk)
        else:
            prompt = f"""You are an expert text correction assistant. Please refine the following OCR-extracted text by:

1. Correcting OCR errors (misrecognized characters, words)
2. Fixing obvious spelling mistakes
3. Improving readability while preserving the original meaning
4. IMPORTANT: Maintain the exact layout, line breaks, and spacing structure
5. Keep all dates, numbers, and proper nouns as accurate as possible
6. If the text appears to be in Vietnamese, correct Vietnamese-specific OCR errors
7. If text contains student names, student IDs (8-digit numbers), or dates (d/m/yyyy format), pay special attention to accuracy

Please return ONLY the corrected text without any explanations or additional comments.

Text to refine:
{chunk}"""

        payload = {
            "messages": [
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            "model": model,
            "stream": False
        }
        refined_text = None
        model_name = model
        for attempt in range(1, max_attempts + 1):
            try:
                payload["model"] = model_name
                response = requests.post(url, headers=headers, json=payload, timeout=120)
                if _is_rate_limited(response=response):
                    retry_after = _get_retry_after_seconds(response)
                    if verbose:
                        print(f"[HuggingFace] Rate limited for chunk {i+1}. Retry {attempt}/{max_attempts} after {retry_after or base_sleep_seconds * attempt}s.")
                    fallback_model = _pick_fallback_model(
                        "huggingface",
                        current_model=model_name,
                        target_capability="text-generation",
                    )
                    if fallback_model:
                        model_name = fallback_model
                        if verbose:
                            print(f"[HuggingFace] Switching to fallback model: {model_name}")
                    if attempt < max_attempts:
                        _sleep_with_backoff(attempt, base_seconds=base_sleep_seconds, retry_after=retry_after)
                        continue
                response.raise_for_status()
                result = response.json()
                # The result should have 'choices' with 'message'->'content'
                if isinstance(result, dict) and "choices" in result and result["choices"]:
                    refined_text = result["choices"][0]["message"]["content"].strip()
                    _set_last_ai_model("huggingface", model_name)
                    if verbose:
                        print(f"[HuggingFace] Chunk {i+1} refined successfully.")
                elif isinstance(result, dict) and "error" in result:
                    error_text = str(result["error"])
                    if _is_rate_limited(error_text=error_text):
                        if attempt < max_attempts:
                            if verbose:
                                print(f"[HuggingFace] Rate limit error for chunk {i+1}. Retry {attempt}/{max_attempts}.")
                            fallback_model = _pick_fallback_model("huggingface", current_model=model_name)
                            if fallback_model:
                                model_name = fallback_model
                                if verbose:
                                    print(f"[HuggingFace] Switching to fallback model: {model_name}")
                            _sleep_with_backoff(attempt, base_seconds=base_sleep_seconds)
                            continue
                    if verbose:
                        print(f"[HuggingFace] API error: {result['error']}")
                    else:
                        print(f"Notice: Hugging Face API error for chunk {i+1}.")
                    refined_text = chunk
                else:
                    if verbose:
                        print(f"[HuggingFace] No valid response from Hugging Face API for chunk {i+1}. Using original text.")
                    else:
                        print(f"Notice: No valid response from Hugging Face API for chunk {i+1}.")
                    refined_text = chunk
                break
            except Exception as e:
                if verbose:
                    print(f"[HuggingFace] Error calling Hugging Face API for chunk {i+1} (attempt {attempt}/{max_attempts}): {e}")
                else:
                    print(f"Notice: Error calling Hugging Face API for chunk {i+1}.")
                fallback_model = _pick_fallback_model("huggingface", current_model=model_name)
                if fallback_model:
                    model_name = fallback_model
                    if verbose:
                        print(f"[HuggingFace] Switching to fallback model: {model_name}")
                if attempt < max_attempts:
                    _sleep_with_backoff(attempt, base_seconds=base_sleep_seconds)
                    continue
                refined_text = chunk
        refined_chunks.append(refined_text if refined_text is not None else chunk)

    refined_text = '\n\n'.join(refined_chunks)
    return refined_text


def _run_local_llm(prompt, model=None, verbose=False):
    command = LOCAL_LLM_COMMAND
    if not command:
        return "", "Missing LOCAL_LLM_COMMAND."
    cmd_parts = shlex.split(command)
    if not cmd_parts:
        return "", "Invalid LOCAL_LLM_COMMAND."
    model_name = model or LOCAL_LLM_MODEL
    if not model_name:
        return "", "Missing LOCAL_LLM_MODEL."
    extra_args = shlex.split(LOCAL_LLM_ARGS) if LOCAL_LLM_ARGS else []
    if "ollama" in os.path.basename(cmd_parts[0]).lower() and "run" not in cmd_parts:
        cmd_parts.append("run")
    cmd = cmd_parts + [model_name] + extra_args
    try:
        result = subprocess.run(
            cmd,
            input=prompt,
            text=True,
            capture_output=True,
            timeout=LOCAL_LLM_TIMEOUT,
            encoding="utf-8",
            errors="replace",
        )
        if result.returncode != 0:
            err = result.stderr.strip() or f"Command failed ({result.returncode})."
            return "", err
        return result.stdout.strip(), ""
    except FileNotFoundError:
        return "", f"Command not found: {cmd_parts[0]}"
    except Exception as e:
        if verbose:
            print(f"[LocalLLM] Error: {e}")
        return "", f"Error: {e}"


def refine_text_with_local_llm(text, model=None, user_prompt=None, verbose=False):
    """
    Use a locally installed LLM (default: Ollama) to refine OCR-extracted text.
    """
    model_name = model or LOCAL_LLM_MODEL
    if not model_name:
        if verbose:
            print("No local model specified. Returning original text.")
        else:
            print("Notice: No local model specified.")
        return text

    max_chunk_size = 4000
    if len(text) <= max_chunk_size:
        chunks = [text]
    else:
        paragraphs = text.split('\n\n')
        chunks = []
        current_chunk = ""
        for paragraph in paragraphs:
            if len(current_chunk) + len(paragraph) + 2 <= max_chunk_size:
                if current_chunk:
                    current_chunk += '\n\n' + paragraph
                else:
                    current_chunk = paragraph
            else:
                if current_chunk:
                    chunks.append(current_chunk)
                current_chunk = paragraph
        if current_chunk:
            chunks.append(current_chunk)

    refined_chunks = []
    for i, chunk in enumerate(chunks):
        if verbose:
            print(f"[LocalLLM] Refining chunk {i+1}/{len(chunks)} with local model...")
        else:
            print(f"Refining chunk {i+1}/{len(chunks)} with local model...")
        if user_prompt:
            prompt = user_prompt.format(text=chunk)
        else:
            prompt = f"""You are an expert text correction assistant. Please refine the following OCR-extracted text by:

1. Correcting OCR errors (misrecognized characters, words)
2. Fixing obvious spelling mistakes
3. Improving readability while preserving the original meaning
4. IMPORTANT: Maintain the exact layout, line breaks, and spacing structure
5. Keep all dates, numbers, and proper nouns as accurate as possible
6. If the text appears to be in Vietnamese, correct Vietnamese-specific OCR errors
7. If text contains student names, student IDs (8-digit numbers), or dates (d/m/yyyy format), pay special attention to accuracy

Please return ONLY the corrected text without any explanations or additional comments.

Text to refine:
{chunk}"""
        refined_text, error = _run_local_llm(prompt, model=model_name, verbose=verbose)
        if error:
            if verbose:
                print(f"[LocalLLM] {error} Using original text.")
            else:
                print("Notice: Local LLM failed. Using original text.")
            refined_chunks.append(chunk)
        else:
            _set_last_ai_model("local", model_name)
            refined_chunks.append(refined_text if refined_text else chunk)

    return '\n\n'.join(refined_chunks)

def refine_text_with_ai(text, method=DEFAULT_AI_METHOD, verbose=False, **kwargs):
    """
    Refine OCR-extracted text using an AI model.
    method: one of ALL_AI_METHODS
    kwargs: extra arguments for the underlying function
    If verbose is True, print more details; otherwise, print only important notice.
    """
    if method not in ALL_AI_METHODS:
        if verbose:
            print(f"[AIRefine] Unknown AI refinement method: {method}. Supported methods: {ALL_AI_METHODS}. Returning original text.")
        else:
            print(f"Notice: Unknown AI refinement method: {method}.")
        return text
    user_prompt = kwargs.get("user_prompt")
    if method == "gemini":
        return refine_text_with_gemini(text, user_prompt=user_prompt, verbose=verbose)
    elif method == "huggingface":
        api_key = kwargs.get("api_key", HUGGINGFACE_API_KEY)
        model = kwargs.get("model", "meta-llama/llama-3.1-8b-instruct")
        return refine_text_with_huggingface(text, api_key=api_key, model=model, user_prompt=user_prompt, verbose=verbose)
    elif method == "local":
        model = kwargs.get("model", LOCAL_LLM_MODEL)
        return refine_text_with_local_llm(text, model=model, user_prompt=user_prompt, verbose=verbose)
    # Fallback (should not reach here)
    return text


def test_ai_models(methods=None, verbose=False, model_override=None):
    """
    Verify AI model connectivity and credentials with a minimal test prompt.
    Returns a dict: {method: {"ok": bool, "message": str}}.
    """
    available = ALL_AI_METHODS if ALL_AI_METHODS else ["gemini", "huggingface", "local"]
    if not methods or methods == "all":
        methods_to_test = available
    elif isinstance(methods, (list, tuple, set)):
        methods_to_test = list(methods)
    else:
        methods_to_test = [methods]

    def extract_rate_limit_headers(headers):
        rate_info = {}
        for key, value in headers.items():
            lower = key.lower()
            if "rate" in lower or "limit" in lower:
                rate_info[key] = value
        return rate_info

    results = {}
    prompt = "Reply with exactly: OK"
    if isinstance(model_override, dict):
        override_map = model_override
    else:
        override_map = {}

    for method in methods_to_test:
        method_override = override_map.get(method) if override_map else model_override
        if verbose:
            print(f"[AITest] Starting test for {method}...")
        if method == "gemini":
            if not GEMINI_API_KEY:
                results[method] = {"ok": False, "message": "Missing GEMINI_API_KEY."}
                continue
            if not GEMINI_DEFAULT_MODEL:
                results[method] = {"ok": False, "message": "Missing GEMINI_DEFAULT_MODEL."}
                continue
            model_name = _normalize_gemini_model_name(method_override) if method_override else GEMINI_DEFAULT_MODEL
            if verbose:
                print(f"[AITest] Gemini model: {model_name}")
            url = f"https://generativelanguage.googleapis.com/v1beta/models/{model_name}:generateContent?key={GEMINI_API_KEY}"
            if verbose:
                print(f"[AITest] Gemini endpoint: {url}")
            headers = {'Content-Type': 'application/json'}
            data = {
                "contents": [
                    {
                        "parts": [
                            {
                                "text": prompt
                            }
                        ]
                    }
                ]
            }
            if verbose:
                print(f"[AITest] Gemini test prompt: {prompt}")
            try:
                resp = requests.post(url, headers=headers, json=data, timeout=120)
                rate_info = extract_rate_limit_headers(resp.headers)
                if verbose and rate_info:
                    print(f"[AITest] Gemini rate limit headers: {rate_info}")
                if _is_rate_limited(response=resp):
                    if method_override:
                        results[method] = {
                            "ok": False,
                            "message": f"Rate limited (HTTP {resp.status_code}).",
                            "model": model_name,
                            "rate_limit": rate_info or None,
                        }
                        continue
                    fallback_model = _pick_fallback_model(
                        "gemini",
                        current_model=_normalize_gemini_model_name(GEMINI_DEFAULT_MODEL),
                        target_capability="generateContent",
                    )
                    if fallback_model:
                        model_name = _normalize_gemini_model_name(fallback_model)
                        if verbose:
                            print(f"[AITest] Gemini rate limited, retrying with fallback model: {model_name}")
                        fallback_url = f"https://generativelanguage.googleapis.com/v1beta/models/{model_name}:generateContent?key={GEMINI_API_KEY}"
                        if verbose:
                            print(f"[AITest] Gemini fallback endpoint: {fallback_url}")
                        resp = requests.post(fallback_url, headers=headers, json=data, timeout=120)
                        rate_info = extract_rate_limit_headers(resp.headers)
                        if verbose and rate_info:
                            print(f"[AITest] Gemini fallback rate limit headers: {rate_info}")
                        if not _is_rate_limited(response=resp):
                            resp.raise_for_status()
                            result = resp.json()
                            response = ""
                            if 'candidates' in result and len(result['candidates']) > 0:
                                response = result['candidates'][0]['content']['parts'][0]['text'].strip()
                            if verbose:
                                print(f"[AITest] Gemini fallback response: {response}")
                            ok = isinstance(response, str) and response.strip().upper().startswith("OK")
                            results[method] = {
                                "ok": ok,
                                "message": "OK (fallback model)" if ok else f"Unexpected response: {str(response).strip()}",
                                "model": model_name,
                                "rate_limit": rate_info or None,
                            }
                            _update_model_cache("gemini", [model_name])
                            continue
                    results[method] = {
                        "ok": False,
                        "message": f"Rate limited (HTTP {resp.status_code}).",
                        "model": model_name,
                        "rate_limit": rate_info or None,
                    }
                    continue
                resp.raise_for_status()
                result = resp.json()
                response = ""
                if 'candidates' in result and len(result['candidates']) > 0:
                    response = result['candidates'][0]['content']['parts'][0]['text'].strip()
                if verbose:
                    print(f"[AITest] Gemini response: {response}")
                ok = isinstance(response, str) and response.strip().upper().startswith("OK")
                results[method] = {
                    "ok": ok,
                    "message": "OK" if ok else f"Unexpected response: {str(response).strip()}",
                    "model": model_name,
                    "rate_limit": rate_info or None,
                }
                _update_model_cache(
                    "gemini",
                    [_normalize_gemini_model_name(model_name)],
                    capabilities={_normalize_gemini_model_name(model_name): ["generateContent"]},
                )
            except Exception as e:
                results[method] = {
                    "ok": False,
                    "message": f"Error: {e}",
                    "model": model_name,
                    "rate_limit": None,
                }
        elif method == "huggingface":
            if not HUGGINGFACE_API_KEY:
                results[method] = {"ok": False, "message": "Missing HUGGINGFACE_API_KEY."}
                continue
            url = "https://router.huggingface.co/novita/v3/openai/chat/completions"
            if verbose:
                print(f"[AITest] HuggingFace endpoint: {url}")
            headers = {
                "Authorization": f"Bearer {HUGGINGFACE_API_KEY}",
                "Content-Type": "application/json"
            }
            model_name = method_override if method_override else "meta-llama/llama-3.1-8b-instruct"
            if verbose:
                print(f"[AITest] HuggingFace model: {model_name}")
            if verbose:
                print(f"[AITest] HuggingFace test prompt: {prompt}")
            payload = {
                "messages": [
                    {
                        "role": "user",
                        "content": prompt
                    }
                ],
                "model": model_name,
                "stream": False
            }
            try:
                resp = requests.post(url, headers=headers, json=payload, timeout=120)
                rate_info = extract_rate_limit_headers(resp.headers)
                if verbose and rate_info:
                    print(f"[AITest] HuggingFace rate limit headers: {rate_info}")
                if _is_rate_limited(response=resp):
                    if method_override:
                        results[method] = {
                            "ok": False,
                            "message": f"Rate limited (HTTP {resp.status_code}).",
                            "model": payload["model"],
                            "rate_limit": rate_info or None,
                        }
                        continue
                    fallback_model = _pick_fallback_model(
                        "huggingface",
                        current_model=payload["model"],
                        target_capability="text-generation",
                    )
                    if fallback_model:
                        if verbose:
                            print(f"[AITest] HuggingFace rate limited, retrying with fallback model: {fallback_model}")
                        payload["model"] = fallback_model
                        resp = requests.post(url, headers=headers, json=payload, timeout=120)
                        rate_info = extract_rate_limit_headers(resp.headers)
                        if verbose and rate_info:
                            print(f"[AITest] HuggingFace fallback rate limit headers: {rate_info}")
                        if not _is_rate_limited(response=resp):
                            resp.raise_for_status()
                            result = resp.json()
                            response = ""
                            if isinstance(result, dict) and "choices" in result and result["choices"]:
                                response = result["choices"][0]["message"]["content"].strip()
                            if verbose:
                                print(f"[AITest] HuggingFace fallback response: {response}")
                            ok = isinstance(response, str) and response.strip().upper().startswith("OK")
                            results[method] = {
                                "ok": ok,
                                "message": "OK (fallback model)" if ok else f"Unexpected response: {str(response).strip()}",
                                "model": payload["model"],
                                "rate_limit": rate_info or None,
                            }
                            _update_model_cache("huggingface", [payload["model"]])
                            continue
                    results[method] = {
                        "ok": False,
                        "message": f"Rate limited (HTTP {resp.status_code}).",
                        "model": payload["model"],
                        "rate_limit": rate_info or None,
                    }
                    continue
                resp.raise_for_status()
                result = resp.json()
                response = ""
                if isinstance(result, dict) and "choices" in result and result["choices"]:
                    response = result["choices"][0]["message"]["content"].strip()
                if verbose:
                    print(f"[AITest] HuggingFace response: {response}")
                ok = isinstance(response, str) and response.strip().upper().startswith("OK")
                results[method] = {
                    "ok": ok,
                    "message": "OK" if ok else f"Unexpected response: {str(response).strip()}",
                    "model": payload["model"],
                    "rate_limit": rate_info or None,
                }
                _update_model_cache(
                    "huggingface",
                    [payload["model"]],
                    capabilities={payload["model"]: ["text-generation"]},
                )
            except Exception as e:
                results[method] = {
                    "ok": False,
                    "message": f"Error: {e}",
                    "model": payload["model"],
                    "rate_limit": None,
                }
        elif method == "local":
            model_name = method_override or LOCAL_LLM_MODEL
            if verbose:
                print(f"[AITest] Local command: {LOCAL_LLM_COMMAND}")
                print(f"[AITest] Local model: {model_name}")
                if LOCAL_LLM_ARGS:
                    print(f"[AITest] Local args: {LOCAL_LLM_ARGS}")
            if verbose:
                print(f"[AITest] Local test prompt: {prompt}")
            response, error = _run_local_llm(prompt, model=model_name, verbose=verbose)
            if verbose and response:
                print(f"[AITest] Local response: {response}")
            if error:
                results[method] = {"ok": False, "message": error, "model": model_name, "rate_limit": None}
            else:
                ok = isinstance(response, str) and response.strip().upper().startswith("OK")
                results[method] = {
                    "ok": ok,
                    "message": "OK" if ok else f"Unexpected response: {str(response).strip()}",
                    "model": model_name,
                    "rate_limit": None,
                }
        else:
            results[method] = {"ok": False, "message": f"Unknown method: {method}."}
            continue

    return results


def list_ai_models(methods=None, verbose=False):
    """
    List available AI models for the provided API keys (when supported by the provider).
    Returns a dict: {method: {"ok": bool, "message": str, "models": list, "rate_limit": dict|None}}.
    """
    available = ALL_AI_METHODS if ALL_AI_METHODS else ["gemini", "huggingface", "local"]
    if not methods or methods == "all":
        methods_to_test = available
    elif isinstance(methods, (list, tuple, set)):
        methods_to_test = list(methods)
    else:
        methods_to_test = [methods]

    results = {}
    for method in methods_to_test:
        if method == "gemini":
            if not GEMINI_API_KEY:
                results[method] = {
                    "ok": False,
                    "message": "Missing GEMINI_API_KEY.",
                    "models": [],
                    "rate_limit": None,
                    "total": 0,
                    "truncated": False,
                }
                continue
            url = f"https://generativelanguage.googleapis.com/v1beta/models?key={GEMINI_API_KEY}"
            try:
                resp = requests.get(url, timeout=120)
                rate_info = {}
                for key, value in resp.headers.items():
                    if "rate" in key.lower() or "limit" in key.lower():
                        rate_info[key] = value
                if _is_rate_limited(response=resp):
                    results[method] = {
                        "ok": False,
                        "message": f"Rate limited (HTTP {resp.status_code}).",
                        "models": [],
                        "rate_limit": rate_info or None,
                    }
                    continue
                resp.raise_for_status()
                payload = resp.json()
                models = []
                capabilities = {}
                for entry in payload.get("models", []) if isinstance(payload, dict) else []:
                    if not isinstance(entry, dict):
                        continue
                    name = entry.get("name")
                    methods = entry.get("supportedGenerationMethods", [])
                    if name and isinstance(methods, list) and "generateContent" in methods:
                        models.append(name)
                        capabilities[_normalize_gemini_model_name(name)] = methods
                total = len(models)
                if total > 50:
                    models = models[:50]
                    truncated = True
                else:
                    truncated = False
                _update_model_cache("gemini", [_normalize_gemini_model_name(m) for m in models], capabilities=capabilities)
                results[method] = {
                    "ok": True,
                    "message": "OK",
                    "models": models,
                    "rate_limit": rate_info or None,
                    "total": total,
                    "truncated": truncated,
                }
            except Exception as e:
                results[method] = {
                    "ok": False,
                    "message": f"Error: {e}",
                    "models": [],
                    "rate_limit": None,
                    "total": 0,
                    "truncated": False,
                }
        elif method == "huggingface":
            if not HUGGINGFACE_API_KEY:
                results[method] = {
                    "ok": False,
                    "message": "Missing HUGGINGFACE_API_KEY.",
                    "models": [],
                    "rate_limit": None,
                    "total": 0,
                    "truncated": False,
                }
                continue
            url = "https://huggingface.co/api/models"
            params = {
                "pipeline_tag": "text-generation",
                "limit": 50,
                "sort": "downloads",
            }
            headers = {"Authorization": f"Bearer {HUGGINGFACE_API_KEY}"}
            try:
                resp = requests.get(url, headers=headers, params=params, timeout=120)
                rate_info = {}
                for key, value in resp.headers.items():
                    if "rate" in key.lower() or "limit" in key.lower():
                        rate_info[key] = value
                if _is_rate_limited(response=resp):
                    results[method] = {
                        "ok": False,
                        "message": f"Rate limited (HTTP {resp.status_code}).",
                        "models": [],
                        "rate_limit": rate_info or None,
                    }
                    continue
                resp.raise_for_status()
                payload = resp.json()
                models = []
                capabilities = {}
                for entry in payload if isinstance(payload, list) else []:
                    model_id = entry.get("modelId") if isinstance(entry, dict) else None
                    if model_id:
                        models.append(model_id)
                total = len(models)
                truncated = total >= 50
                for model_id in models:
                    capabilities[model_id] = ["text-generation"]
                _update_model_cache("huggingface", models, capabilities=capabilities)
                results[method] = {
                    "ok": True,
                    "message": "OK (top public models; availability depends on provider).",
                    "models": models,
                    "rate_limit": rate_info or None,
                    "total": total,
                    "truncated": truncated,
                }
            except Exception as e:
                results[method] = {
                    "ok": False,
                    "message": f"Error: {e}",
                    "models": [],
                    "rate_limit": None,
                    "total": 0,
                    "truncated": False,
                }
        elif method == "local":
            cmd_parts = shlex.split(LOCAL_LLM_COMMAND) if LOCAL_LLM_COMMAND else []
            if not cmd_parts:
                results[method] = {
                    "ok": False,
                    "message": "Missing LOCAL_LLM_COMMAND.",
                    "models": [],
                    "rate_limit": None,
                    "total": 0,
                    "truncated": False,
                }
                continue
            if "ollama" in os.path.basename(cmd_parts[0]).lower() and "list" not in cmd_parts:
                cmd_parts.append("list")
            try:
                resp = subprocess.run(
                    cmd_parts,
                    capture_output=True,
                    text=True,
                    timeout=LOCAL_LLM_TIMEOUT,
                    encoding="utf-8",
                    errors="replace",
                )
                if resp.returncode != 0:
                    message = resp.stderr.strip() or f"Command failed ({resp.returncode})."
                    results[method] = {
                        "ok": False,
                        "message": message,
                        "models": [],
                        "rate_limit": None,
                        "total": 0,
                        "truncated": False,
                    }
                    continue
                models = []
                for idx, line in enumerate(resp.stdout.splitlines()):
                    if idx == 0 and "NAME" in line.upper():
                        continue
                    parts = line.strip().split()
                    if not parts:
                        continue
                    models.append(parts[0])
                total = len(models)
                results[method] = {
                    "ok": True,
                    "message": "OK",
                    "models": models,
                    "rate_limit": None,
                    "total": total,
                    "truncated": False,
                }
            except Exception as e:
                results[method] = {
                    "ok": False,
                    "message": f"Error: {e}",
                    "models": [],
                    "rate_limit": None,
                    "total": 0,
                    "truncated": False,
                }
        else:
            results[method] = {
                "ok": False,
                "message": f"Unknown method: {method}.",
                "models": [],
                "rate_limit": None,
                "total": 0,
                "truncated": False,
            }

    return results


def detect_local_ai_models(verbose=False):
    """
    Detect locally installed AI models (Ollama or llama.cpp compatible).
    Returns a dict with command, models, and status.
    """
    cmd_parts = shlex.split(LOCAL_LLM_COMMAND) if LOCAL_LLM_COMMAND else []
    if not cmd_parts:
        return {
            "ok": False,
            "message": "Missing LOCAL_LLM_COMMAND.",
            "command": LOCAL_LLM_COMMAND,
            "models": [],
        }
    command_base = os.path.basename(cmd_parts[0]).lower()
    if "ollama" in command_base:
        if "list" not in cmd_parts:
            cmd_parts.append("list")
        try:
            resp = subprocess.run(
                cmd_parts,
                capture_output=True,
                text=True,
                timeout=LOCAL_LLM_TIMEOUT,
                encoding="utf-8",
                errors="replace",
            )
            if resp.returncode != 0:
                message = resp.stderr.strip() or f"Command failed ({resp.returncode})."
                return {
                    "ok": False,
                    "message": message,
                    "command": " ".join(cmd_parts),
                    "models": [],
                }
            models = []
            for idx, line in enumerate(resp.stdout.splitlines()):
                if idx == 0 and "NAME" in line.upper():
                    continue
                parts = line.strip().split()
                if not parts:
                    continue
                models.append(parts[0])
            return {
                "ok": True,
                "message": "OK",
                "command": " ".join(cmd_parts),
                "models": models,
            }
        except Exception as e:
            if verbose:
                print(f"[LocalDetect] Error: {e}")
            return {
                "ok": False,
                "message": f"Error: {e}",
                "command": " ".join(cmd_parts),
                "models": [],
            }
    if "llama" in command_base:
        help_cmd = cmd_parts + ["--help"]
        try:
            resp = subprocess.run(
                help_cmd,
                capture_output=True,
                text=True,
                timeout=LOCAL_LLM_TIMEOUT,
                encoding="utf-8",
                errors="replace",
            )
            if resp.returncode != 0:
                message = resp.stderr.strip() or f"Command failed ({resp.returncode})."
                return {
                    "ok": False,
                    "message": message,
                    "command": " ".join(help_cmd),
                    "models": [],
                }
            gguf_dir = LOCAL_LLM_GGUF_DIR
            gguf_models = []
            if gguf_dir and os.path.isdir(gguf_dir):
                for root, _, files in os.walk(gguf_dir):
                    for filename in files:
                        if filename.lower().endswith(".gguf"):
                            gguf_models.append(os.path.join(root, filename))
            message = "OK (llama.cpp detected; provide a .gguf model path in LOCAL_LLM_MODEL)."
            if gguf_models:
                message = f"OK (llama.cpp detected; found {len(gguf_models)} .gguf model(s))."
            return {
                "ok": True,
                "message": message,
                "command": " ".join(help_cmd),
                "models": gguf_models,
            }
        except Exception as e:
            if verbose:
                print(f"[LocalDetect] Error: {e}")
            return {
                "ok": False,
                "message": f"Error: {e}",
                "command": " ".join(help_cmd),
                "models": [],
            }
    return {
        "ok": False,
        "message": "Unknown local LLM command (expected ollama or llama.cpp).",
        "command": " ".join(cmd_parts),
        "models": [],
    }

def calculate_cc_ck_gk(student, gradebook_csv_path="canvas_gradebook.csv", override_file="override_grades.xlsx", verbose=False):
    """
    Calculate CC (Chuyên cần), CK (Cuối kỳ), GK (Giữa kỳ) for a student object using Canvas gradebook CSV.
    If gradebook_csv_path is provided, extract scores from the CSV using the student's Canvas ID.
    Returns a dict: {"CC": cc, "CK": ck, "GK": gk, "details": {...}}
    """
    overrides = get_override_grades(file_path=override_file, verbose=verbose)
    override_entry = None
    student_id = _normalize_student_id(getattr(student, "Student ID", ""))
    student_name = str(getattr(student, "Name", "")).strip()
    if overrides.get("by_id") and student_id in overrides["by_id"]:
        override_entry = overrides["by_id"][student_id]
    elif overrides.get("by_name") and student_name:
        override_entry = overrides["by_name"].get(_normalize_vietnamese_name(student_name))
    override_cc = override_entry.get("CC") if override_entry and _is_grade_provided(override_entry.get("CC")) else None
    override_gk = override_entry.get("GK") if override_entry and _is_grade_provided(override_entry.get("GK")) else None
    override_ck = override_entry.get("CK") if override_entry and _is_grade_provided(override_entry.get("CK")) else None
    override_fields = []
    if override_cc is not None:
        override_fields.append("CC")
    if override_gk is not None:
        override_fields.append("GK")
    if override_ck is not None:
        override_fields.append("CK")
    if verbose and override_entry:
        print(
            "[OverrideGrades] Found entry for "
            f"{student_name} ({student_id}): CC={override_cc}, GK={override_gk}, CK={override_ck}, "
            f"reason={override_entry.get('reason', '')}"
        )
    if verbose and not override_entry:
        print(f"[OverrideGrades] No entry found for {student_name} ({student_id}).")

    if override_cc is not None and override_gk is not None and override_ck is not None:
        details = {
            "diem_danh": None,
            "quiz": None,
            "bai_tap": None,
            "GK": override_gk,
            "CK": override_ck,
            "CC": override_cc,
        }
        override_reason = override_entry.get("reason", "") if override_entry else ""
        if verbose:
            print("[OverrideGrades] All scores overridden; skipping gradebook/attribute calculation.")
        return {
            "CC": override_cc,
            "CK": override_ck,
            "GK": override_gk,
            "details": details,
            "override_reason": override_reason,
            "override_fields": override_fields,
        }

    def to_scale_10(val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return 0.0
        try:
            v = float(val)
            if pd.isna(v):
                return 0.0
            # Normalize if on 100-point scale
            if v > 10:
                return round(v / 10, 2)
            return v
        except Exception:
            return 0.0

    def to_scale_10_optional(val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return None
        if isinstance(val, str) and not val.strip():
            return None
        try:
            v = float(val)
            if pd.isna(v):
                return None
            if v > 10:
                return round(v / 10, 2)
            return v
        except Exception:
            return None

    def normalize_text(s):
        if s is None:
            return ""
        s = str(s).strip().lower()
        # Remove accents to match variants
        s = unicodedata.normalize('NFD', s)
        s = ''.join(c for c in s if not unicodedata.combining(c))
        return s

    # Keywords for CC (attendance/participation/assignment group weight may vary)
    cc_keywords = [
        "chuyen can",           # chuyên cần
        "attendance",
        "participation",
        "diem danh",
        "cc",                   # shorthand
    ]
    # "Final Score" variants to prefer (Canvas style)
    final_keywords = [
        "final score",
        "unposted final score",
        "final points",
        "final",
    ]
    disallowed_score_keywords = [
        "possible score",
        "points possible",
        "possible points",
    ]

    # Also support English/VN for midterm/final
    gk_keywords = ["giua ky", "midterm", "gk"]
    gk_exclude_keywords = [
        "attendance",
        "participation",
        "diem danh",
        "chuyen can",
        "quiz",
        "assignment",
        "bai tap",
    ]
    ck_keywords = ["cuoi ky", "final exam", "ck"]
    ck_exclude_keywords = [
        "attendance",
        "participation",
        "diem danh",
        "chuyen can",
        "quiz",
        "assignment",
        "bai tap",
    ]

    # Helper: best matching column finder
    def find_best_col(columns, primary_keywords, prefer_final=True, exclude_keywords=None):
        """
        Find a column that matches any keyword in primary_keywords.
        If prefer_final=True, only accept columns that also include a 'final' keyword.
        Returns the first best match or None.
        """
        normalized_cols = [(col, normalize_text(col)) for col in columns]
        exclude_keywords = exclude_keywords or []
        # First pass: both primary and final keyword
        if prefer_final:
            for col, ncol in normalized_cols:
                if any(bad in ncol for bad in disallowed_score_keywords):
                    continue
                if any(ex in ncol for ex in exclude_keywords):
                    continue
                if any(k in ncol for k in primary_keywords) and any(fk in ncol for fk in final_keywords):
                    return col
            return None
        # Second pass: any primary keyword (excluding possible score columns)
        for col, ncol in normalized_cols:
            if any(bad in ncol for bad in disallowed_score_keywords):
                continue
            if any(ex in ncol for ex in exclude_keywords):
                continue
            if any(k in ncol for k in primary_keywords):
                return col
        return None

    def find_unposted_col(columns, primary_keywords, exclude_keywords=None):
        normalized_cols = [(col, normalize_text(col)) for col in columns]
        exclude_keywords = exclude_keywords or []
        for col, ncol in normalized_cols:
            if any(bad in ncol for bad in disallowed_score_keywords):
                continue
            if any(ex in ncol for ex in exclude_keywords):
                continue
            if any(k in ncol for k in primary_keywords) and "unposted final score" in ncol:
                return col
        return None

    def pick_score(final_val, unposted_val):
        if final_val is None:
            return unposted_val if unposted_val is not None else 0.0
        if final_val == 0.0 and unposted_val not in (None, 0.0):
            return unposted_val
        return final_val

    # If gradebook_csv_path is provided, use it to extract scores
    if gradebook_csv_path:
        try:
            df = pd.read_csv(gradebook_csv_path)
            # Try to find the Canvas ID column
            canvas_id_col = None
            for col in df.columns:
                ncol = normalize_text(col)
                if ncol in ("canvas user id", "user id", "id"):
                    canvas_id_col = col
                    break
            if not canvas_id_col:
                raise ValueError("Canvas ID column not found in gradebook CSV.")

            student_canvas_id = getattr(student, "Canvas ID", None)
            if not student_canvas_id:
                raise ValueError("Student Canvas ID not found in student object.")

            # Find the row for this student
            row = df[df[canvas_id_col] == int(student_canvas_id)]
            if row.empty:
                raise ValueError(f"Student Canvas ID {student_canvas_id} not found in gradebook CSV.")
            row = row.iloc[0]

            # Detect CC column robustly
            cc_col = find_best_col(df.columns, cc_keywords, prefer_final=True)
            cc_unposted_col = find_unposted_col(df.columns, cc_keywords)

            # Fallback known component columns (Canvas groups)
            diem_danh_col = find_best_col(df.columns, ["diem danh", "attendance"], prefer_final=True) or "Điểm danh Final Score"
            diem_danh_unposted_col = find_unposted_col(df.columns, ["diem danh", "attendance"])
            quiz_col = find_best_col(df.columns, ["quiz"], prefer_final=True) or "Quiz Final Score"
            quiz_unposted_col = find_unposted_col(df.columns, ["quiz"])
            bai_tap_col = find_best_col(df.columns, ["bai tap", "assignment"], prefer_final=True) or "Bài tập Final Score"
            bai_tap_unposted_col = find_unposted_col(df.columns, ["bai tap", "assignment"])

            # Compute CC unless overridden.
            if override_cc is not None:
                CC = override_cc
                if verbose:
                    print("[OverrideGrades] Using override CC from override_grades.xlsx.")
            elif cc_col or cc_unposted_col:
                cc_final = to_scale_10_optional(row.get(cc_col, None)) if cc_col else None
                cc_unposted = to_scale_10_optional(row.get(cc_unposted_col, None)) if cc_unposted_col else None
                CC = pick_score(cc_final, cc_unposted)
                if verbose and cc_final in (None, 0.0) and cc_unposted not in (None, 0.0):
                    print("[calculate_cc_ck_gk] Using unposted CC score from gradebook.")
            else:
                diem_danh = pick_score(
                    to_scale_10_optional(row.get(diem_danh_col, None)),
                    to_scale_10_optional(row.get(diem_danh_unposted_col, None)),
                )
                quiz = pick_score(
                    to_scale_10_optional(row.get(quiz_col, None)),
                    to_scale_10_optional(row.get(quiz_unposted_col, None)),
                )
                bai_tap = pick_score(
                    to_scale_10_optional(row.get(bai_tap_col, None)),
                    to_scale_10_optional(row.get(bai_tap_unposted_col, None)),
                )
                CC = round(0.25 * diem_danh + 0.25 * quiz + 0.5 * bai_tap, 1)

            # Detect GK/CK columns robustly
            gk_col = find_best_col(df.columns, gk_keywords, prefer_final=True) or "Giữa kỳ Final Score"
            gk_unposted_col = find_unposted_col(df.columns, gk_keywords, exclude_keywords=gk_exclude_keywords)
            ck_col = find_best_col(df.columns, ck_keywords, prefer_final=True, exclude_keywords=ck_exclude_keywords) or "Cuối kỳ Final Score"
            ck_unposted_col = find_unposted_col(df.columns, ck_keywords, exclude_keywords=ck_exclude_keywords)

            if override_gk is not None:
                GK = override_gk
            else:
                gk_final = to_scale_10_optional(row.get(gk_col, None))
                gk_unposted = to_scale_10_optional(row.get(gk_unposted_col, None))
                GK = pick_score(gk_final, gk_unposted)
                if verbose and gk_final in (None, 0.0) and gk_unposted not in (None, 0.0):
                    print("[calculate_cc_ck_gk] Using unposted GK score from gradebook.")
            if override_ck is not None:
                CK = override_ck
            else:
                ck_final = to_scale_10_optional(row.get(ck_col, None))
                ck_unposted = to_scale_10_optional(row.get(ck_unposted_col, None))
                CK = pick_score(ck_final, ck_unposted)
                if verbose and ck_final in (None, 0.0) and ck_unposted not in (None, 0.0):
                    print("[calculate_cc_ck_gk] Using unposted CK score from gradebook.")
            if verbose and override_gk is not None:
                print("[OverrideGrades] Using override GK from override_grades.xlsx.")
            if verbose and override_ck is not None:
                print("[OverrideGrades] Using override CK from override_grades.xlsx.")

            details = {
                "diem_danh": pick_score(
                    to_scale_10_optional(row.get(diem_danh_col, None)),
                    to_scale_10_optional(row.get(diem_danh_unposted_col, None)),
                ),
                "quiz": pick_score(
                    to_scale_10_optional(row.get(quiz_col, None)),
                    to_scale_10_optional(row.get(quiz_unposted_col, None)),
                ),
                "bai_tap": pick_score(
                    to_scale_10_optional(row.get(bai_tap_col, None)),
                    to_scale_10_optional(row.get(bai_tap_unposted_col, None)),
                ),
                "GK": GK,
                "CK": CK,
                "CC": CC,
            }
            CC, CK, GK, override_reason = apply_override_grades(student, CC, CK, GK, overrides)
            details["CC"] = CC
            details["GK"] = GK
            details["CK"] = CK
            return {
                "CC": CC,
                "CK": CK,
                "GK": GK,
                "details": details,
                "override_reason": override_reason,
                "override_fields": override_fields,
            }
        except Exception as e:
            if verbose:
                print(f"[calculate_cc_ck_gk] Error extracting from gradebook CSV: {e}")
            # Fallback to default method below

    # Fallback: use student object attributes (legacy logic)
    def get_avg_score_by_title(student, keyword):
        scores = []
        for attr in student.__dict__:
            if attr.startswith("Assignment: "):
                title = attr[len("Assignment: "):]
                if keyword.lower() in title.lower():
                    val = getattr(student, attr)
                    try:
                        score = float(val)
                        if pd.isna(score):
                            continue
                        scores.append(score)
                    except Exception:
                        continue
        if scores:
            return round(sum(scores) / len(scores), 2)
        return None

    # Try to detect CC attribute robustly unless overridden
    if override_cc is not None:
        CC = override_cc
        if verbose:
            print("[OverrideGrades] Using override CC from override_grades.xlsx.")
    else:
        CC = None
        cc_final_attr = None
        cc_unposted_attr = None
        for attr in student.__dict__:
            nattr = normalize_text(attr)
            if any(k in nattr for k in cc_keywords):
                if "unposted final score" in nattr:
                    if not cc_unposted_attr:
                        cc_unposted_attr = attr
                elif any(fk in nattr for fk in final_keywords):
                    if not cc_final_attr:
                        cc_final_attr = attr
        if cc_final_attr or cc_unposted_attr:
            cc_final = to_scale_10_optional(getattr(student, cc_final_attr, None)) if cc_final_attr else None
            cc_unposted = to_scale_10_optional(getattr(student, cc_unposted_attr, None)) if cc_unposted_attr else None
            CC = pick_score(cc_final, cc_unposted)
        if CC is None or CC == 0.0:
            # Try multiple variants
            diem_danh = pick_score(
                to_scale_10_optional(getattr(student, "??i???m danh Final Score", None)) or to_scale_10_optional(getattr(student, "Attendance Final Score", None)),
                to_scale_10_optional(getattr(student, "??i???m danh Unposted Final Score", None)) or to_scale_10_optional(getattr(student, "Attendance Unposted Final Score", None)),
            )
            quiz = pick_score(
                to_scale_10_optional(getattr(student, "Quiz Final Score", None)),
                to_scale_10_optional(getattr(student, "Quiz Unposted Final Score", None)),
            )
            bai_tap = pick_score(
                to_scale_10_optional(getattr(student, "BA?i t??-p Final Score", None)) or to_scale_10_optional(getattr(student, "Assignment Final Score", None)),
                to_scale_10_optional(getattr(student, "BA?i t??-p Unposted Final Score", None)) or to_scale_10_optional(getattr(student, "Assignment Unposted Final Score", None)),
            )
            CC = round(0.25 * (diem_danh or 0) + 0.25 * (quiz or 0) + 0.5 * (bai_tap or 0), 1)

    if override_gk is not None:
        GK = override_gk
        if verbose:
            print("[OverrideGrades] Using override GK from override_grades.xlsx.")
    else:
        GK = get_avg_score_by_title(student, "giữa kỳ")
        if GK is None:
            midterm_id = str(CANVAS_MIDTERM_ASSIGNMENT_ID).strip()
            if midterm_id:
                GK = getattr(student, f"Assignment: {midterm_id}", None)
            if GK is None:
                gk_final = to_scale_10_optional(getattr(student, "Midterm Final Score", None))
                gk_unposted = to_scale_10_optional(getattr(student, "Midterm Unposted Final Score", None))
                GK = pick_score(gk_final, gk_unposted)

    if override_ck is not None:
        CK = override_ck
        if verbose:
            print("[OverrideGrades] Using override CK from override_grades.xlsx.")
    else:
        CK = get_avg_score_by_title(student, "cuối kỳ")
        if CK is None:
            final_id = str(CANVAS_FINAL_ASSIGNMENT_ID).strip()
            if final_id:
                CK = getattr(student, f"Assignment: {final_id}", None)
            if CK is None:
                ck_final = to_scale_10_optional(getattr(student, "Final Final Score", None))
                ck_unposted = to_scale_10_optional(getattr(student, "Final Unposted Final Score", None))
                CK = pick_score(ck_final, ck_unposted)

    CC = 0.0 if CC is None or (isinstance(CC, float) and pd.isna(CC)) else CC
    CK = 0.0 if CK is None or (isinstance(CK, float) and pd.isna(CK)) else CK
    GK = 0.0 if GK is None or (isinstance(GK, float) and pd.isna(GK)) else GK

    details = {
        "diem_danh": pick_score(
            to_scale_10_optional(getattr(student, "??i???m danh Final Score", None)) or to_scale_10_optional(getattr(student, "Attendance Final Score", None)),
            to_scale_10_optional(getattr(student, "??i???m danh Unposted Final Score", None)) or to_scale_10_optional(getattr(student, "Attendance Unposted Final Score", None)),
        ),
        "quiz": pick_score(
            to_scale_10_optional(getattr(student, "Quiz Final Score", None)),
            to_scale_10_optional(getattr(student, "Quiz Unposted Final Score", None)),
        ),
        "bai_tap": pick_score(
            to_scale_10_optional(getattr(student, "BA?i t??-p Final Score", None)) or to_scale_10_optional(getattr(student, "Assignment Final Score", None)),
            to_scale_10_optional(getattr(student, "BA?i t??-p Unposted Final Score", None)) or to_scale_10_optional(getattr(student, "Assignment Unposted Final Score", None)),
        ),
        "GK": GK,
        "CK": CK,
        "CC": CC,
    }
    CC, CK, GK, override_reason = apply_override_grades(student, CC, CK, GK, overrides)
    details["CC"] = CC
    details["GK"] = GK
    details["CK"] = CK
    return {
        "CC": CC,
        "CK": CK,
        "GK": GK,
        "details": details,
        "override_reason": override_reason,
        "override_fields": override_fields,
    }

def export_grade_diff_csv(rows, output_path, verbose=False):
    if not rows:
        if verbose:
            print("[GradeDiff] No changes to export.")
        return None
    if not output_path:
        output_path = os.path.join(os.getcwd(), "grade_diff.csv")
    if DRY_RUN:
        print(f"[GradeDiff] Dry run: would write diff CSV to {output_path}")
        return output_path
    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["Student ID", "Name", "Field", "Old", "New", "Source", "Match Type"],
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    if verbose:
        print(f"[GradeDiff] Wrote grade diff CSV to {output_path}")
    else:
        print(f"Grade diff CSV saved to {output_path}")
    append_run_report(
        "export-grade-diff",
        details=f"rows={len(rows)}",
        outputs=output_path,
        verbose=verbose,
    )
    return output_path


def generate_final_evaluations(
    students,
    results_dir=None,
    override_file="override_grades.xlsx",
    dry_run=False,
    verbose=False,
):
    """Generate per-student final evaluation TXT reports.

    Writes reports to `results_dir` (default: ./final_evaluations) unless `dry_run`.
    Returns a list of dicts with CC/GK/CK/total and other metadata.
    """
    if results_dir is None:
        results_dir = os.path.join(os.getcwd(), "final_evaluations")

    if not dry_run:
        os.makedirs(results_dir, exist_ok=True)

    w_cc = float(WEIGHT_CC)
    w_gk = float(WEIGHT_GK)
    w_ck = float(WEIGHT_CK)
    total_weight = w_cc + w_gk + w_ck
    if abs(total_weight - 1.0) > 1e-6:
        raise ValueError(
            f"Invalid weights: WEIGHT_CC + WEIGHT_GK + WEIGHT_CK must sum to 1.0 (got {total_weight:.6f})."
        )

    evaluations = []
    for s in students:
        sid = _normalize_student_id(getattr(s, "Student ID", ""))
        if not sid:
            inferred_sid = _infer_student_id_from_email(getattr(s, "Email", ""))
            if inferred_sid:
                sid = inferred_sid
                if not getattr(s, "Student ID", ""):
                    setattr(s, "Student ID", sid)

        name = str(getattr(s, "Name", "")).strip()
        name_display = name.replace("/", "_").replace("\\", "_").replace(" ", "_")

        scores = calculate_cc_ck_gk(s, override_file=override_file, verbose=verbose)
        CC = scores["CC"]
        CK = scores["CK"]
        GK = scores["GK"]
        details = scores["details"]
        override_reason = scores.get("override_reason", "")
        override_fields = scores.get("override_fields", [])

        cc_value = float(CC) if CC is not None else 0.0
        gk_value = float(GK) if GK else 0.0
        ck_value = float(CK) if CK else 0.0
        total_score = round(
            (w_cc * cc_value) + (w_gk * gk_value) + (w_ck * ck_value),
            1,
        )

        result_lines = []
        result_lines.append(f"Student ID (Mã sinh viên): {sid}")
        result_lines.append(f"Name (Họ và Tên): {getattr(s, 'Name', '')}")
        result_lines.append(f"CC (Chuyên cần): {CC}")
        result_lines.append(f"GK (Giữa kỳ / Midterm): {GK}")
        result_lines.append(f"CK (Cuối kỳ / Final): {CK}")
        result_lines.append(
            f"Formula (Công thức): Total (Tổng điểm) = {w_cc:.2f}*CC + {w_gk:.2f}*GK + {w_ck:.2f}*CK"
        )

        group_scores = [details.get("diem_danh"), details.get("quiz"), details.get("bai_tap")]
        has_group_scores = False
        for score in group_scores:
            try:
                if score is not None and not pd.isna(score) and float(score) != 0.0:
                    has_group_scores = True
                    break
            except Exception:
                if score:
                    has_group_scores = True
                    break
        if has_group_scores:
            result_lines.append("Assignment group scores (Điểm thành phần):")
            result_lines.append(f"  Attendance (Điểm danh): {details['diem_danh']}")
            result_lines.append(f"  Quiz (Trắc nghiệm): {details['quiz']}")
            result_lines.append(f"  Assignment (Bài tập): {details['bai_tap']}")
        if override_fields:
            result_lines.append(f"Overridden scores (Điểm được ghi đè): {', '.join(override_fields)}")
        if override_reason:
            result_lines.append(f"Override reason (Lý do ghi đè): {override_reason}")
        result_lines.append(f"Total score (scale 10) (Tổng điểm thang 10): {total_score}")

        report_text = "\n".join(result_lines)
        _reset_last_ai_model()
        if REPORT_REFINE_METHOD:
            try:
                report_text = refine_text_with_ai(report_text, method=REPORT_REFINE_METHOD, verbose=verbose)
            except Exception as e:
                if verbose:
                    print(f"[ReportRefine] Failed to refine report with {REPORT_REFINE_METHOD}: {e}")
        model_used = get_last_ai_model_used()
        if REPORT_REFINE_METHOD and model_used:
            report_text += f"\nAI model used (Mô hình AI): {model_used}"
        default_model = ""
        if REPORT_REFINE_METHOD == "gemini":
            default_model = GEMINI_DEFAULT_MODEL
        elif REPORT_REFINE_METHOD == "huggingface":
            default_model = "meta-llama/llama-3.1-8b-instruct"
        elif REPORT_REFINE_METHOD == "local":
            default_model = LOCAL_LLM_MODEL
        if REPORT_REFINE_METHOD and default_model:
            report_text += f"\nDefault model (Mô hình mặc định): {default_model}"

        filename_sid = sid or "unknown"
        result_filename = f"{filename_sid}_{name_display}_results.txt"
        result_path = os.path.join(results_dir, result_filename)
        if not dry_run:
            with open(result_path, "w", encoding="utf-8") as f:
                f.write(report_text)

        evaluations.append(
            {
                "student_id": sid,
                "name": name,
                "CC": CC,
                "GK": GK,
                "CK": CK,
                "total_score": total_score,
                "details": details,
                "override_reason": override_reason,
                "override_fields": override_fields,
                "report_path": result_path,
            }
        )

    if verbose:
        if dry_run:
            print("[FinalEvals] Dry run: skipped writing evaluation reports.")
        else:
            print(f"[FinalEvals] Saved {len(evaluations)} evaluation reports to {results_dir}")

    return evaluations


def update_mat_excel_grades(file_path, students, output_path=None, diff_output_path=None, verbose=False):
    """
    Update the columns CC, CK, GK (Attendance, Final, Midterm) in a MAT*.xlsx file.
    - Create a temp copy, unmerge all merged cells, detect header and end rows.
    - For each student, detect row via student id or student name (allowing Vietnamese name matching).
    - Calculate CC, CK, GK from database.
    - Create a new copy in the current folder with the same layout as the original.
    - Fill the corresponding cells in the new copy with calculated values.
    - Returns the path to the updated file.
    The header row must contain CC, CK, GK.
    """
    global _override_grades_cache
    global _override_grades_cache_path

    def normalize_vietnamese_name(name):
        """Normalize Vietnamese name by removing diacritics and converting to lowercase."""
        if not name:
            return ""
        # Normalize to NFD (decomposed form)
        name = unicodedata.normalize('NFD', name)
        # Remove combining characters (diacritics)
        name = ''.join(c for c in name if not unicodedata.combining(c))
        # Convert to lowercase and strip whitespace
        return name.lower().strip()

    def names_match(name1, name2):
        """Check if two names match after normalization."""
        return normalize_vietnamese_name(name1) == normalize_vietnamese_name(name2)

    # Prefer override_grades.xlsx alongside the MAT file, if present.
    # Temporarily swap the override cache so per-course files are respected.
    override_cache_backup = _override_grades_cache
    override_cache_path_backup = _override_grades_cache_path
    override_path = os.path.join(os.path.dirname(os.path.abspath(file_path)), "override_grades.xlsx")
    override_file = "override_grades.xlsx"
    if os.path.exists(override_path):
        _override_grades_cache = load_override_grades(file_path=override_path, verbose=verbose)
        _override_grades_cache_path = override_path
        override_file = override_path
        if verbose:
            override_count = len(_override_grades_cache.get("by_id", {})) + len(_override_grades_cache.get("by_name", {}))
            print(f"[OverrideGrades] Using override file: {override_file} (entries: {override_count})")
    elif verbose:
        print(f"[OverrideGrades] No override file found at: {override_path}")

    dry_run = DRY_RUN
    # Step 1: Create temp copy and unmerge all merged cells
    temp_path = None
    if dry_run:
        wb_temp = openpyxl.load_workbook(file_path)
    else:
        temp_path = file_path + "_temp.xlsx"
        shutil.copy2(file_path, temp_path)
        wb_temp = openpyxl.load_workbook(temp_path)
    ws_temp = wb_temp.active

    merged_ranges = list(ws_temp.merged_cells.ranges)
    if verbose:
        print(f"[update_mat_excel_grades] Unmerging {len(merged_ranges)} merged cell ranges in temp copy...")
    for merged_range in merged_ranges:
        min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
        value = ws_temp.cell(row=min_row, column=min_col).value
        ws_temp.unmerge_cells(str(merged_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws_temp.cell(row=row, column=col).value = value

    # Step 2: Detect header row and columns
    header_row_idx = None
    for i in range(9, 20):
        row = [str(ws_temp.cell(row=i + 1, column=j + 1).value or "").strip().lower() for j in range(ws_temp.max_column)]
        has_cc = any(cell == "cc" or "chuyên cần" in cell for cell in row)
        has_ck = any(cell == "ck" or "cuối kỳ" in cell for cell in row)
        has_gk = any(cell == "gk" or "giữa kỳ" in cell for cell in row)
        if has_cc and has_ck and has_gk:
            header_row_idx = i + 1
            break
    if header_row_idx is None:
        raise ValueError("Header row with CC, CK, GK columns could not be detected.")
    if verbose:
        print(f"[update_mat_excel_grades] Detected header row at index {header_row_idx} (Excel row {header_row_idx})")

    header = [str(ws_temp.cell(row=header_row_idx, column=j + 1).value or "").strip() for j in range(ws_temp.max_column)]
    if verbose:
        print(f"[update_mat_excel_grades] Header columns: {header}")
    col_map = {}
    for idx, col in enumerate(header):
        col_lower = col.lower()
        if (
            ("mã" in col_lower and "sinh" in col_lower) or
            ("mã" in col_lower and "sv" in col_lower) or
            ("mssv" in col_lower) or
            ("student" in col_lower and "id" in col_lower)
        ):
            col_map["Student ID"] = idx + 1
        elif "họ" in col_lower and "tên" in col_lower or "tên" == col_lower or "name" in col_lower:
            col_map["Name"] = idx + 1
        elif col_lower == "cc" or "chuyên cần" in col_lower:
            col_map["CC"] = idx + 1
        elif col_lower == "ck" or "cuối kỳ" in col_lower:
            col_map["CK"] = idx + 1
        elif col_lower == "gk" or "giữa kỳ" in col_lower:
            col_map["GK"] = idx + 1
    for required in ["CC", "CK", "GK"]:
        if required not in col_map:
            raise ValueError(f"Column '{required}' not found in header row.")
    if verbose:
        print(f"[update_mat_excel_grades] Column mapping: {col_map}")

    # Step 3: Detect end row (before "Tổng số sinh viên")
    end_row = ws_temp.max_row
    for i in range(header_row_idx + 1, ws_temp.max_row + 1):
        row_str = " ".join(str(ws_temp.cell(row=i, column=j + 1).value or "").lower() for j in range(ws_temp.max_column))
        if "tổng số sinh viên" in row_str:
            end_row = i - 1
            if verbose:
                print(f"[update_mat_excel_grades] Detected end row at index {end_row} (Excel row {end_row})")
            break

    # Step 4: Build maps for both student ID and name
    sid_to_cells = {}
    name_to_cells = {}
    row_to_info = {}  # row_idx -> (sid, name)
    
    for i in range(header_row_idx + 1, end_row + 1):
        sid_cell = ws_temp.cell(row=i, column=col_map.get("Student ID", 1))
        sid = str(sid_cell.value).strip() if sid_cell.value else ""
        
        name = ""
        if "Name" in col_map:
            name_cell = ws_temp.cell(row=i, column=col_map["Name"])
            name = str(name_cell.value).strip() if name_cell.value else ""
        
        if not sid and not name:
            continue
            
        cell_info = {}
        for field in ["CC", "CK", "GK"]:
            if field in col_map:
                cell = ws_temp.cell(row=i, column=col_map[field])
                cell_info[field] = cell.coordinate
        
        row_to_info[i] = (sid, name)
        
        if sid:
            sid_to_cells[sid] = cell_info
        if name:
            name_to_cells[normalize_vietnamese_name(name)] = cell_info
            
    if verbose:
        print(f"[update_mat_excel_grades] Found {len(sid_to_cells)} student rows by ID and {len(name_to_cells)} by name in the Excel file.")

    # Step 5: Calculate CC, CK, GK for each student using the helper function
    sid_to_values = {}
    results_dir = os.path.join(os.getcwd(), "final_evaluations")
    if not dry_run:
        os.makedirs(results_dir, exist_ok=True)

    w_cc = float(WEIGHT_CC)
    w_gk = float(WEIGHT_GK)
    w_ck = float(WEIGHT_CK)
    total_weight = w_cc + w_gk + w_ck
    if abs(total_weight - 1.0) > 1e-6:
        raise ValueError(
            f"Invalid weights: WEIGHT_CC + WEIGHT_GK + WEIGHT_CK must sum to 1.0 (got {total_weight:.6f})."
        )

    evaluations = generate_final_evaluations(
        students,
        results_dir=results_dir,
        override_file=override_file,
        dry_run=dry_run,
        verbose=verbose,
    )

    for ev in evaluations:
        sid = ev.get("student_id", "")
        name = ev.get("name", "")

        # Try to find matching row by ID first, then by name
        cell_info = None
        match_type = None
        if sid and sid in sid_to_cells:
            cell_info = sid_to_cells[sid]
            match_type = "ID"
        elif name:
            normalized_name = normalize_vietnamese_name(name)
            if normalized_name in name_to_cells:
                cell_info = name_to_cells[normalized_name]
                match_type = "name"

        if cell_info:
            key = sid or f"NAME:{normalize_vietnamese_name(name)}"
            sid_to_values[key] = {
                "student_id": sid,
                "CC": ev.get("CC"),
                "CK": ev.get("CK"),
                "GK": ev.get("GK"),
                "cell_info": cell_info,
                "match_type": match_type,
                "name": name,
                "override_fields": ev.get("override_fields", []),
            }
            if verbose:
                print(f"[update_mat_excel_grades] Matched student {name} ({sid}) by {match_type}")
        else:
            if verbose:
                print(f"[update_mat_excel_grades] Warning: Could not find Excel row for student {name} ({sid})")

    # Legacy inline generation loop disabled (replaced by generate_final_evaluations)
    for s in []:
        sid = _normalize_student_id(getattr(s, "Student ID", ""))
        if not sid:
            inferred_sid = _infer_student_id_from_email(getattr(s, "Email", ""))
            if inferred_sid:
                sid = inferred_sid
                if not getattr(s, "Student ID", ""):
                    setattr(s, "Student ID", sid)
        name = str(getattr(s, "Name", "")).strip()
        name_display = name.replace("/", "_").replace("\\", "_").replace(" ", "_")
        
        scores = calculate_cc_ck_gk(s, override_file=override_file, verbose=verbose)
        CC = scores["CC"]
        CK = scores["CK"]
        GK = scores["GK"]
        details = scores["details"]
        override_reason = scores.get("override_reason", "")
        override_fields = scores.get("override_fields", [])
        cc_value = float(CC) if CC is not None else 0.0
        gk_value = float(GK) if GK else 0.0
        ck_value = float(CK) if CK else 0.0
        total_score = round(
            (w_cc * cc_value) + (w_gk * gk_value) + (w_ck * ck_value),
            1,
        )
        
        # Try to find matching row by ID first, then by name
        cell_info = None
        match_type = None
        if sid and sid in sid_to_cells:
            cell_info = sid_to_cells[sid]
            match_type = "ID"
        elif name:
            normalized_name = normalize_vietnamese_name(name)
            if normalized_name in name_to_cells:
                cell_info = name_to_cells[normalized_name]
                match_type = "name"
        
        if cell_info:
            sid_to_values[sid] = {
                "CC": CC,
                "CK": CK,
                "GK": GK,
                "cell_info": cell_info,
                "match_type": match_type,
                "name": name,
                "override_fields": override_fields,
            }
            if verbose:
                print(f"[update_mat_excel_grades] Matched student {name} ({sid}) by {match_type}")
        else:
            if verbose:
                print(f"[update_mat_excel_grades] Warning: Could not find Excel row for student {name} ({sid})")

        # Save results to TXT file
        result_lines = []
        result_lines.append(f"Student ID (M\u00e3 sinh vi\u00ean): {sid}")
        result_lines.append(f"Name (H\u1ecd v\u00e0 T\u00ean): {getattr(s, 'Name', '')}")
        result_lines.append(f"CC (Chuy\u00ean c\u1ea7n): {CC}")
        result_lines.append(f"GK (Gi\u1eefa k\u1ef3 / Midterm): {GK}")
        result_lines.append(f"CK (Cu\u1ed1i k\u1ef3 / Final): {CK}")
        result_lines.append(
            f"Formula (Công thức): Total (Tổng điểm) = {w_cc:.2f}*CC + {w_gk:.2f}*GK + {w_ck:.2f}*CK"
        )
        group_scores = [details.get("diem_danh"), details.get("quiz"), details.get("bai_tap")]
        has_group_scores = False
        for score in group_scores:
            try:
                if score is not None and not pd.isna(score) and float(score) != 0.0:
                    has_group_scores = True
                    break
            except Exception:
                if score:
                    has_group_scores = True
                    break
        if has_group_scores:
            result_lines.append("Assignment group scores (\u0110i\u1ec3m th\u00e0nh ph\u1ea7n):")
            result_lines.append(f"  Attendance (\u0110i\u1ec3m danh): {details['diem_danh']}")
            result_lines.append(f"  Quiz (Tr\u1eafc nghi\u1ec7m): {details['quiz']}")
            result_lines.append(f"  Assignment (B\u00e0i t\u1eadp): {details['bai_tap']}")
        if override_fields:
            result_lines.append(f"Overridden scores (\u0110i\u1ec3m \u0111\u01b0\u1ee3c ghi \u0111\u00e8): {', '.join(override_fields)}")
        if override_reason:
            result_lines.append(f"Override reason (L\u00fd do ghi \u0111\u00e8): {override_reason}")
        result_lines.append(f"Total score (scale 10) (T\u1ed5ng \u0111i\u1ec3m thang 10): {total_score}")

        report_text = "\n".join(result_lines)
        _reset_last_ai_model()
        if REPORT_REFINE_METHOD:
            try:
                report_text = refine_text_with_ai(report_text, method=REPORT_REFINE_METHOD, verbose=verbose)
            except Exception as e:
                if verbose:
                    print(f"[ReportRefine] Failed to refine report with {REPORT_REFINE_METHOD}: {e}")
        model_used = get_last_ai_model_used()
        if REPORT_REFINE_METHOD and model_used:
            report_text += f"\nAI model used (M\u00f4 h\u00ecnh AI): {model_used}"
        default_model = ""
        if REPORT_REFINE_METHOD == "gemini":
            default_model = GEMINI_DEFAULT_MODEL
        elif REPORT_REFINE_METHOD == "huggingface":
            default_model = "meta-llama/llama-3.1-8b-instruct"
        elif REPORT_REFINE_METHOD == "local":
            default_model = LOCAL_LLM_MODEL
        if REPORT_REFINE_METHOD and default_model:
            report_text += f"\nDefault model (M\u00f4 h\u00ecnh m\u1eb7c \u0111\u1ecbnh): {default_model}"

        filename_sid = sid or "unknown"
        result_filename = f"{filename_sid}_{name_display}_results.txt"
        if not dry_run:
            result_path = os.path.join(results_dir, result_filename)
            with open(result_path, "w", encoding="utf-8") as f:
                f.write(report_text)

    if verbose:
        print(f"[update_mat_excel_grades] Calculated CC, CK, GK values for {len(sid_to_values)} students from Canvas assignment scores.")
        if not dry_run:
            print(f"[update_mat_excel_grades] Saved individual results to {results_dir}")
        else:
            print("[update_mat_excel_grades] Dry run: skipped writing individual results.")

    # Step 6: Create a new copy in the current folder with the same layout as the original
    if not output_path:
        base = os.path.splitext(os.path.basename(file_path))[0]
        ext = os.path.splitext(file_path)[1]
        output_path = os.path.join(os.getcwd(), f"{base}_updated{ext}")
    if dry_run:
        ws_out = ws_temp
    else:
        shutil.copy2(file_path, output_path)
        wb_out = openpyxl.load_workbook(output_path)
        ws_out = wb_out.active

    # Step 7: Fill the corresponding cells in the new copy
    updated_count = 0
    diff_rows = []
    for _, value_info in sid_to_values.items():
        sid = value_info.get("student_id", "")
        cell_info = value_info["cell_info"]
        match_type = value_info["match_type"]
        name = value_info.get("name", "")
        override_fields = set(value_info.get("override_fields") or [])
        
        for field in ["CC", "CK", "GK"]:
            if field in cell_info and field in value_info:
                cell_addr = cell_info[field]
                cell = ws_out[cell_addr]
                old_val = cell.value
                new_val = value_info[field]
                if new_val is None:
                    new_val = 0.0
                if not dry_run:
                    cell.value = new_val
                updated_count += 1
                diff_rows.append({
                    "Student ID": sid,
                    "Name": name,
                    "Field": field,
                    "Old": old_val,
                    "New": new_val,
                    "Source": "override" if field in override_fields else "computed",
                    "Match Type": match_type,
                })
                if verbose:
                    print(f"[update_mat_excel_grades] Student ID {sid} (matched by {match_type}), {field}: {old_val} -> {new_val}, cell: {cell_addr}")

    if not dry_run:
        wb_out.save(output_path)
    if temp_path and os.path.exists(temp_path):
        os.remove(temp_path)
    # Restore global cache to avoid side effects on other operations.
    _override_grades_cache = override_cache_backup
    _override_grades_cache_path = override_cache_path_backup
    if diff_output_path:
        export_grade_diff_csv(diff_rows, diff_output_path, verbose=verbose)
    if verbose:
        if dry_run:
            print(f"[update_mat_excel_grades] Dry run: would update {updated_count} cells in {output_path}")
        else:
            print(f"[update_mat_excel_grades] Updated {updated_count} cells in {output_path}")
    else:
        if dry_run:
            print(f"Dry run: would update {updated_count} cells in {output_path}")
        else:
            print(f"Updated {updated_count} cells in {output_path}")
    append_run_report(
        "update-mat-excel",
        details=f"cells={updated_count}",
        outputs=[output_path, diff_output_path] if diff_output_path else output_path,
        verbose=verbose,
    )
    return output_path

def read_students_from_excel_csv(file_path, db_path=None, verbose=False, preview_only=False, preview_rows=5):
    """
    Read students from Excel or CSV file, detect and normalize header row, deduplicate, and update database if db_path is given.
    Ignores metadata rows (e.g., rows containing 'ĐẠI HỌC QUỐC GIA HÀ NỘI', 'TRƯỜNG ĐẠI HỌC KHOA HỌC TỰ NHIÊN', etc.).
    Supports both .xlsx/.xls and .csv files.
    Handles cases where the header row is not the first row or is multi-row.
    Unmerges all merged cells (horizontal and vertical) for header detection.
    For files matching "MAT*.xlsx" pattern, starts reading from row 10 forward.
    For MAT*.xlsx files, also ignores all rows after the row containing "Tổng số sinh viên" or similar metadata.
    Extracts registered class ID info (Canvas section) if available in the file.
    Extracts GitHub usernames from columns named "GitHub username" or similar variants.
    If verbose is True, print more details; otherwise, print only important notice.
    """
    if verbose:
        print(f"[ExcelCSV] Reading file: {file_path}")
    else:
        print("Reading file...")

    # Check if the file matches the MAT*.xlsx pattern
    filename = os.path.basename(file_path)
    is_mat_file = filename.startswith("MAT") and filename.lower().endswith(".xlsx")

    ext = os.path.splitext(file_path)[1].lower()
    # For Excel: unmerge merged cells before loading into DataFrame
    if ext in [".xlsx", ".xls"]:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        # Unmerge all merged cells (horizontal and vertical)
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
            value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(merged_range))
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).value = value

        # Read all rows into a list of lists
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        df_raw = pd.DataFrame(data)
    elif ext == ".csv":
        df_raw = pd.read_csv(file_path, header=None)
    else:
        raise ValueError("Unsupported file type. Please provide an Excel (.xlsx, .xls) or CSV (.csv) file.")

    # Patterns for metadata rows to ignore
    metadata_patterns = [
        r"đại học quốc gia hà nội",
        r"trường đại học khoa học tự nhiên",
        r"bảng điểm đánh giá học phần",
        r"học kỳ.*năm học",
        r"giảng viên",
        r"lớp hp",
        r"tên hp",
        r"trọng số điểm",
        r"tổng số sinh viên.*sinh viên",  # Adding the new pattern
    ]

    def is_metadata_row(row):
        for cell in row:
            cell_str = str(cell).lower()
            for pat in metadata_patterns:
                if re.search(pat, cell_str):
                    return True
        return False

    # For MAT*.xlsx files, start from row 10 and ignore all rows after "Tổng số sinh viên"
    if is_mat_file:
        if verbose:
            print(f"[ExcelCSV] Detected MAT*.xlsx file format, starting from row 10")
        else:
            print("Detected MAT*.xlsx file format, starting from row 10")
        first_data_idx = 9  # 0-based index for row 10
        df_raw = df_raw.iloc[first_data_idx:].reset_index(drop=True)

        # Find the row containing "Tổng số sinh viên" and ignore all rows after it
        end_idx = None
        for i in range(len(df_raw)):
            row = df_raw.iloc[i]
            for cell in row:
                if isinstance(cell, str) and "tổng số sinh viên" in cell.lower():
                    end_idx = i
                    break
            if end_idx is not None:
                break
        # Additional: ignore all rows from the row having content "Tổng số sinh viên: 40 sinh viên"
        if end_idx is None:
            for i in range(len(df_raw)):
                row = df_raw.iloc[i]
                for cell in row:
                    if isinstance(cell, str) and "tổng số sinh viên: 40 sinh viên" in cell.lower():
                        end_idx = i
                        break
                if end_idx is not None:
                    break
        if end_idx is not None:
            df_raw = df_raw.iloc[:end_idx].reset_index(drop=True)
            if verbose:
                print(f"[ExcelCSV] Ignored all rows from row {end_idx+1} containing 'Tổng số sinh viên' or 'Tổng số sinh viên: 40 sinh viên'")

        # Now check for additional metadata rows like "Tổng số sinh viên"
        filtered_rows = []
        for i in range(len(df_raw)):
            if not is_metadata_row(df_raw.iloc[i]):
                filtered_rows.append(i)

        if filtered_rows:
            df_raw = df_raw.iloc[filtered_rows].reset_index(drop=True)
            if verbose:
                print(f"[ExcelCSV] Removed {len(df_raw) - len(filtered_rows)} metadata rows from MAT*.xlsx file")

        header_row_idx = 0  # First non-metadata row is the header
        header_rows = [header_row_idx]
    else:
        # Patterns for header detection (Vietnamese and English)
        header_patterns = [
            r"(họ.*tên|tên|full\s*name|name)",  # Name
            r"(mã.*sinh.*viên|mã\s*sv|mssv|student\s*id)",  # Student ID
            r"(email)",  # Email
            r"(lớp|class)",  # Class
            r"(ngày\s*sinh|date\s*of\s*birth|dob)",  # DOB
            # Added patterns for class/section ID
            r"(mã\s*lớp|lớp\s*đăng\s*ký|mã\s*section|section\s*id|canvas\s*section|section)",  # Section/Class ID
            r"(nhóm|group)",  # Group/Section
            r"(mã\s*đăng\s*ký|registration\s*id)",  # Registration ID
            r"(registered\s*class\s*id|class\s*registration)",  # Registered Class ID
            # Added patterns for GitHub username
            r"(github\s*username|github\s*account|github\s*id|github\s*handle|github)",  # GitHub username
        ]

        # Remove metadata rows at the top
        first_data_idx = 0
        for i in range(len(df_raw)):
            row = df_raw.iloc[i]
            if is_metadata_row(row):
                continue
            # If row is not metadata and not all empty, break
            if not all(str(cell).strip() == "" or pd.isna(cell) for cell in row):
                first_data_idx = i
                break
        df_raw = df_raw.iloc[first_data_idx:].reset_index(drop=True)

        # Find header row index (best match with header_patterns)
        header_row_idx = None
        max_matches = 0
        for i in range(min(10, len(df_raw))):  # Only check first 10 rows
            row = df_raw.iloc[i]
            matches = 0
            for cell in row:
                cell_str = str(cell).lower()
                for pat in header_patterns:
                    if re.search(pat, cell_str):
                        matches += 1
                        break
            if matches > max_matches:
                max_matches = matches
                header_row_idx = i
            # If enough matches, break early
            if matches >= 2:
                break

        if header_row_idx is None:
            # Fallback: assume first row is header
            header_row_idx = 0

        # Handle multi-row header: merge rows if next row(s) are also header-like
        header_rows = [header_row_idx]
        for i in range(header_row_idx + 1, header_row_idx + 3):
            if i < len(df_raw):
                row = df_raw.iloc[i]
                matches = 0
                for cell in row:
                    cell_str = str(cell).lower()
                    for pat in header_patterns:
                        if re.search(pat, cell_str):
                            matches += 1
                            break
                if matches >= 2:
                    header_rows.append(i)
                else:
                    break

    # Build header by joining multi-row header cells
    if len(header_rows) == 1:
        header = [str(x) for x in df_raw.iloc[header_rows[0]].values]
    else:
        header = []
        for col in range(df_raw.shape[1]):
            parts = []
            for hidx in header_rows:
                val = df_raw.iloc[hidx, col]
                if pd.notna(val) and str(val).strip():
                    parts.append(str(val).strip())
            header.append(" ".join(parts))

    # Read data after header rows
    df = pd.DataFrame(df_raw.values[len(header_rows):], columns=header)

    # Remove completely empty columns
    df = df.dropna(axis=1, how='all')
    # Remove completely empty rows
    df = df.dropna(axis=0, how='all')

    df = normalize_columns(df, verbose=verbose)

    if is_mat_file:
        mat_drop_cols = {
            "Attendance", "Midterm", "Final", "Participation", "Assignment", "Quiz",
            "Total", "Total Score", "Total Score (DB)", "Total Score (Canvas)", "Total Final Score",
            "Total Grade", "Total Final Grade",
            "CC", "GK", "CK",
            "Tổng điểm", "Điểm học phần", "Điểm tổng", "Điểm tổng kết", "Tổng kết",
        }
        mat_drop_cols_lower = {str(c).strip().lower() for c in mat_drop_cols}
        drop_cols = []
        for col in df.columns:
            col_lower = str(col).strip().lower()
            if col in mat_drop_cols or col_lower in mat_drop_cols_lower:
                drop_cols.append(col)
        if drop_cols:
            df = df.drop(columns=drop_cols)
            if verbose:
                print(f"[ExcelCSV] Dropped score columns for MAT*.xlsx import: {drop_cols}")
            else:
                print(f"Dropped {len(drop_cols)} score columns for MAT*.xlsx import.")

    def render_import_preview():
        columns = list(df.columns)
        missing = [col for col in ["Name", "Student ID"] if col not in columns]
        display_cols = columns[:8]
        extra_cols = len(columns) - len(display_cols)
        print("Import preview")
        print("-" * 14)
        print(f"Columns detected: {', '.join(display_cols)}" + (f" (+{extra_cols} more)" if extra_cols > 0 else ""))
        if missing:
            print(f"Missing required columns: {', '.join(missing)}")
        print(f"Total rows (post-cleaning): {len(df)}")
        sample = df.head(preview_rows)
        if not sample.empty:
            print("Sample rows:")
            for _, row in sample.iterrows():
                row_data = {col: row.get(col, "") for col in display_cols}
                print(f"  {row_data}")
        else:
            print("No rows found after cleaning.")
        return {
            "columns": columns,
            "missing_required": missing,
            "total_rows": len(df),
            "sample_rows": sample.to_dict(orient="records"),
        }

    if preview_only:
        return render_import_preview()

    # Check for section/class ID columns
    section_columns = []
    section_column_names = ["Section ID", "Canvas Section", "Registration ID", "Class ID", "Registered Class ID"]
    for col in df.columns:
        col_lower = str(col).lower()
        if any(term in col_lower for term in ["section", "nhóm", "registration", "đăng ký", "mã lớp", "registered class", "class registration"]):
            section_columns.append(col)
            if verbose:
                print(f"[ExcelCSV] Detected potential section column: {col}")
    
    # Check for GitHub username columns
    github_columns = []
    github_column_names = ["GitHub Username", "GitHub Account", "GitHub ID", "GitHub Handle", "GitHub"]
    for col in df.columns:
        col_lower = str(col).lower()
        if "github" in col_lower and not any(name_term in col_lower for name_term in ["fullname", "họ tên", "họ và tên"]):
            github_columns.append(col)
            if verbose:
                print(f"[ExcelCSV] Detected GitHub username column: {col}")

    # Helper functions for name/id validation
    def is_strange_name(name):
        if not name or not isinstance(name, str):
            return True
        n = name.strip().lower()
        # Remove names with numbers or dates or strange patterns
        if re.search(r"\d", n):
            return True
        if re.search(r"\bngày\b|\btháng\b|\bnăm\b", n):
            return True
        if re.search(r"\(.*\)|\[.*\]|\{.*\}|<.*>", n):  # Brackets or similar symbols
            return True
        if re.search(r"\(ký.*họ tên\)", n):
            return True
        # Remove names that look like dates
        if re.search(r"\d{1,2}\s*(tháng|/)\s*\d{1,2}\s*(năm|/)\s*\d{4}", n):
            return True
        # Remove names that are too short or generic
        if len(n) < 3 or n in {"", "họ và tên", "test student", "possible points", "student name", "full name", "name", "tên", "points possible", "student", "test", "sample"}:
            return True
        return False

    def is_valid_student_id(sid):
        if not sid or not isinstance(sid, str):
            return False
        sid = sid.strip()
        return bool(re.fullmatch(r"\d{8}", sid))

    new_students = []
    for _, row in tqdm(df.iterrows(), total=len(df), desc="Processing rows"):
        student_dict = row.to_dict()
        sid = student_dict.get("Student ID", None)
        name = student_dict.get("Name", None)
        email = student_dict.get("Email", None)
        
        # Strip leading and trailing spaces from name, but preserve spaces between words
        if name:
            name = str(name).strip()
        student_dict["Name"] = name
        
        # Add default email if student has student id but no email
        if sid and (not email or str(email).strip() == ""):
            student_dict["Email"] = f"{str(sid).strip()}@hus.edu.vn"
        
        # Extract section/class ID information
        section_id = None
        for col in section_columns:
            if col in student_dict and student_dict[col]:
                section_value = str(student_dict[col]).strip()
                if section_value:
                    section_id = section_value
                    break
        
        # If section ID found, store it with a standardized name
        if section_id:
            student_dict["Canvas Section"] = section_id
            if verbose and student_dict.get("Student ID"):
                print(f"[ExcelCSV] Found section ID for student {student_dict.get('Student ID')}: {section_id}")
        
        # Extract GitHub username information
        github_username = None
        for col in github_columns:
            if col in student_dict and student_dict[col]:
                username_value = str(student_dict[col]).strip()
                if username_value:
                    github_username = username_value
                    break
        
        # If GitHub username found, store it with a standardized name
        if github_username:
            # Clean up GitHub username (remove @ prefix if present)
            if github_username.startswith('@'):
                github_username = github_username[1:]
            student_dict["GitHub Username"] = github_username
            if verbose and student_dict.get("Student ID"):
                print(f"[ExcelCSV] Found GitHub username for student {student_dict.get('Student ID')}: {github_username}")
        
        # Remove rows with strange names or invalid student id
        if is_strange_name(name):
            continue
        if sid and not is_valid_student_id(str(sid)):
            continue
        
        new_students.append(Student(**student_dict))

    if db_path and os.path.exists(db_path):
        existing_students = load_database(db_path)
        all_students = existing_students + new_students
        before_count = len(existing_students)
    else:
        all_students = new_students
        before_count = 0

    # Helper functions
    email_pattern = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")

    def get_email(student):
        for k, v in student.__dict__.items():
            v_str = str(v).strip()
            if v_str and email_pattern.match(v_str):
                return v_str.lower()
        return None

    def get_name(student):
        name = getattr(student, "Name", None)
        if name:
            return str(name).strip().lower()
        for k in student.__dict__:
            if "name" in k.lower():
                return str(getattr(student, k)).strip().lower()
        return None

    def get_student_id(student):
        sid = getattr(student, "Student ID", None)
        if sid:
            return str(sid).strip()
        return None

    def get_section_id(student):
        for field_name in ["Canvas Section", "Section ID", "Registration ID", "Class ID", "Registered Class ID"]:
            section = getattr(student, field_name, None)
            if section:
                return str(section).strip()
        return None
    
    def get_github_username(student):
        for field_name in ["GitHub Username", "GitHub Account", "GitHub ID", "GitHub Handle", "GitHub"]:
            username = getattr(student, field_name, None)
            if username:
                return str(username).strip()
        return None

    # Deduplication logic with update/merge
    unique_students = []
    for s in all_students:
        sid_s = get_student_id(s)
        name_s = get_name(s)
        email_s = get_email(s)
        section_s = get_section_id(s)
        github_s = get_github_username(s)
        found_duplicate = False
        for u in unique_students:
            sid_u = get_student_id(u)
            name_u = get_name(u)
            email_u = get_email(u)
            is_duplicate = False
            if sid_s and sid_u and sid_s == sid_u:
                is_duplicate = True
            elif (sid_s and not sid_u) or (not sid_s and sid_u):
                if email_s and email_u and email_s == email_u:
                    is_duplicate = True
                elif name_s and name_u and name_s == name_u:
                    is_duplicate = True
            elif email_s and email_u and email_s == email_u:
                is_duplicate = True
            elif (email_s and not email_u) or (not email_s and email_u):
                if name_s and name_u and name_s == name_u:
                    is_duplicate = True
            if is_duplicate:
                # Update u with any new fields from s
                for k, v in s.__dict__.items():
                    if k not in u.__dict__ or not u.__dict__[k]:
                        u.__dict__[k] = v
                    # Special handling for section ID: keep both if different
                    elif k in ["Canvas Section", "Section ID", "Registration ID", "Class ID", "Registered Class ID"] and u.__dict__[k] != v and v:
                        u.__dict__[f"Additional {k}"] = v
                found_duplicate = True
                break
        if not found_duplicate:
            unique_students.append(s)

    after_count = len(unique_students)
    if verbose:
        print(f"[ExcelCSV] Summary: {before_count} students before, {after_count} students after reading and deduplication.")
        section_count = sum(1 for s in unique_students if get_section_id(s))
        print(f"[ExcelCSV] {section_count} students with section/class ID information.")
        github_count = sum(1 for s in unique_students if get_github_username(s))
        print(f"[ExcelCSV] {github_count} students with GitHub username information.")
    else:
        print(f"Summary: {before_count} students before, {after_count} students after reading and deduplication.")
        section_count = sum(1 for s in unique_students if get_section_id(s))
        print(f"{section_count} students with section/class ID information.")
        github_count = sum(1 for s in unique_students if get_github_username(s))
        print(f"{github_count} students with GitHub username information.")

    if db_path:
        save_database(unique_students, db_path, audit_source="import-excel-csv")
        append_run_report(
            "import-students",
            details=f"source={file_path} before={before_count} after={after_count}",
            outputs=db_path,
            verbose=verbose,
        )
    return unique_students

def read_students_from_pdf(pdf_path, db_path=None, lang="auto", service=DEFAULT_OCR_METHOD, verbose=False):
    """
    Extract student info from a scanned PDF file using OCR.
    Returns a list of Student objects with fields: Name, Student ID, Dob, Class (if found).
    Updates the database if db_path is provided (deduplication and merge logic similar to Excel/CSV).
    Only lines matching the format: <some number> <student id> <full name> <dob> <class> are used.
    Excludes metadata lines (e.g., containing 'ĐẠI HỌC QUỐC GIA HÀ NỘI', 'TRƯỜNG ĐẠI HỌC KHOA HỌC TỰ NHIÊN', etc.).
    If verbose is True, print more details; otherwise, print only important notice.
    """
    # Step 1: Extract text from PDF using OCR
    if verbose:
        print(f"[PDF] Extracting text from PDF using {service}...")
    else:
        print("Extracting text from PDF...")
    txt_path = extract_text_from_scanned_pdf(pdf_path, lang=lang, service=service)
    if not txt_path or not os.path.exists(txt_path):
        if verbose:
            print("[PDF] Failed to extract text from PDF.")
        else:
            print("Failed to extract text from PDF.")
        return []

    with open(txt_path, "r", encoding="utf-8") as f:
        text = f.read()

    # Step 2: Parse text to extract student info
    ocr_digit_map = {
        '0': '0', 'O': '0', 'o': '0', 'Q': '0',
        '1': '1', 'I': '1', 'l': '1', '|': '1', '!': '1',
        '2': '2', 'Z': '2',
        '3': '3',
        '4': '4', 'A': '4',
        '5': '5', 'S': '5', 's': '5',
        '6': '6', 'G': '6',
        '7': '7', 'T': '7',
        '8': '8', 'B': '8',
        '9': '9', 'g': '9', 'q': '9',
    }
    def normalize_id(s):
        s = ''.join(c for c in s if c.isalnum() or c in '|!')
        mapped = []
        for c in s:
            mapped.append(ocr_digit_map.get(c, c))
        norm = ''.join(mapped)
        return re.findall(r'(\d{8})', norm)

    # Pattern: <some number> <student id> <full name> <dob> <class>
    dob_regex = r'(\d{1,2})\s*/\s*(\d{1,2})\s*/\s*(\d{4})'
    line_pattern = re.compile(
        r'^\s*(\S+)\s+(\S+)\s+([^\d]+?)\s+' + dob_regex + r'(?:\s+(.*))?$'
    )

    # Metadata patterns to exclude (same as Excel/CSV)
    metadata_patterns = [
        r"đại học quốc gia hà nội",
        r"trường đại học khoa học tự nhiên",
        r"bảng điểm đánh giá học phần",
        r"học kỳ.*năm học",
        r"giảng viên",
        r"lớp hp",
        r"tên hp",
        r"trọng số điểm",
    ]
    def is_metadata_line(line):
        line_lower = line.lower()
        for pat in metadata_patterns:
            if pat in line_lower:
                return True
        return False

    # Helper to check if name is valid (no digits, no special chars)
    def is_valid_name(name):
        if not name or len(name.split()) < 2:
            return False
        if re.search(r'\d', name):
            return False
        if re.search(r'[^A-Za-zÀ-ỹà-ỹ\s\'\-]', name):
            return False
        return True

    new_students = []
    lines = text.splitlines()
    for line in lines:
        if is_metadata_line(line):
            continue
        m = line_pattern.match(line)
        if not m:
            continue
        some_number = m.group(1)
        sid_raw = m.group(2)
        name = m.group(3).strip()
        dob = f"{int(m.group(4))}/{int(m.group(5))}/{m.group(6)}"
        class_name = m.group(7).strip() if m.group(7) else ""

        # Normalize student id
        ids = normalize_id(sid_raw)
        if not ids:
            continue
        sid = ids[0]

        # Name must be valid (no digits, no special chars, at least 2 words)
        if not is_valid_name(name):
            continue

        student_dict = {"Student ID": sid, "Name": name, "Dob": dob}
        if class_name:
            student_dict["Class"] = class_name
        new_students.append(Student(**student_dict))
    if verbose:
        print(f"[PDF] Extracted {len(new_students)} students from PDF.")
    else:
        print(f"Extracted {len(new_students)} students from PDF.")

    # Deduplication and database update logic (similar to Excel/CSV)
    if db_path and os.path.exists(db_path):
        existing_students = load_database(db_path)
        all_students = existing_students + new_students
        before_count = len(existing_students)
    else:
        all_students = new_students
        before_count = 0

    def get_name(student):
        name = getattr(student, "Name", None)
        if name:
            return str(name).strip().lower()
        for k in student.__dict__:
            if "name" in k.lower():
                return str(getattr(student, k)).strip().lower()
        return None

    def get_student_id(student):
        sid = getattr(student, "Student ID", None)
        if sid:
            return str(sid).strip()
        return None

    unique_students = []
    for s in all_students:
        sid_s = get_student_id(s)
        name_s = get_name(s)
        found_duplicate = False
        for u in unique_students:
            sid_u = get_student_id(u)
            name_u = get_name(u)
            is_duplicate = False
            if sid_s and sid_u and sid_s == sid_u:
                is_duplicate = True
            elif name_s and name_u and name_s == name_u:
                is_duplicate = True
            if is_duplicate:
                # Update u with any new fields from s
                for k, v in s.__dict__.items():
                    if k not in u.__dict__ or not u.__dict__[k]:
                        u.__dict__[k] = v
                found_duplicate = True
                break
        if not found_duplicate:
            unique_students.append(s)

    after_count = len(unique_students)
    if verbose:
        print(f"[PDF] Summary: {before_count} students before, {after_count} students after reading and deduplication.")
    else:
        print(f"Summary: {before_count} students before, {after_count} students after reading and deduplication.")

    if db_path:
        save_database(unique_students, db_path, verbose=verbose)
    return unique_students


def ensure_student_emails(students, default_domain="hus.edu.vn", verbose=False):
    """
    Ensure each student has an email. If missing and Student ID exists, create a default email.

    Returns:
        int: number of emails added.
    """
    if not students:
        return 0
    updated = 0
    for student in students:
        email = getattr(student, "Email", None)
        if not email:
            for key, value in getattr(student, "__dict__", {}).items():
                if "email" in key.lower() and value:
                    email = value
                    break
        if email:
            continue
        sid = _normalize_student_id(getattr(student, "Student ID", "") or "")
        if not sid:
            continue
        email = f"{sid}@{default_domain}"
        setattr(student, "Email", email)
        updated += 1
        if verbose:
            print(f"[EmailFallback] Added default email for {sid}: {email}")
    return updated


def import_students_from_file(
    file_path,
    db_path=None,
    verbose=False,
    ocr_service=DEFAULT_OCR_METHOD,
    ocr_lang="auto",
):
    """
    Import students from CSV/XLSX/PDF into the database (if db_path provided).

    Returns:
        list: List of Student objects after import.
    """
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()
    if ext in {".csv", ".xlsx", ".xls"}:
        return read_students_from_excel_csv(file_path, db_path=db_path, verbose=verbose)
    if ext == ".pdf":
        return read_students_from_pdf(
            file_path,
            db_path=db_path,
            lang=ocr_lang or "auto",
            service=ocr_service or DEFAULT_OCR_METHOD,
            verbose=verbose,
        )
    raise ValueError("Unsupported file type. Use CSV, XLSX, XLS, or PDF.")

def sort_students_by_firstname(students, verbose=False):
    """
    Sort students by first name (last word in Name), then by family name (all but last word), then by student id.
    Sorting is by Vietnamese alphabetical order: 
    A, Ă, Â, B, C, D, Đ, E, Ê, G, H, I, K, L, M, N, O, Ô, Ơ, P, Q, R, S, T, U, Ư, V, X, Y
    If verbose is True, print more details; otherwise, print only important notice.
    """
    # Vietnamese alphabet order map
    vn_order = [
        'a', 'ă', 'â', 'b', 'c', 'd', 'đ', 'e', 'ê', 'g', 'h', 'i', 'k', 'l', 'm', 'n',
        'o', 'ô', 'ơ', 'p', 'q', 'r', 's', 't', 'u', 'ư', 'v', 'x', 'y'
    ]
    vn_order_map = {char: idx for idx, char in enumerate(vn_order)}
    # For multi-char letters, map to their order as well
    vn_order_map.update({
        'ch': vn_order_map['c'],
        'gh': vn_order_map['g'],
        'ng': vn_order_map['n'],
        'ngh': vn_order_map['n'],
        'nh': vn_order_map['n'],
        'ph': vn_order_map['p'],
        'th': vn_order_map['t'],
        'tr': vn_order_map['t'],
        'gi': vn_order_map['g'],
        'kh': vn_order_map['k'],
        'qu': vn_order_map['q'],
    })

    def vn_norm(x):
        # Normalize to NFD, lower, replace đ with đ, remove combining marks except for ̆, ̂, ̛
        x = unicodedata.normalize("NFD", x).lower()
        x = x.replace('đ', 'đ')  # keep đ for correct order
        # Remove all combining marks except ̆ (U+0306), ̂ (U+0302), ̛ (U+031B)
        x = ''.join(
            c for c in x
            if not unicodedata.combining(c) or ord(c) in (0x0306, 0x0302, 0x031B)
        )
        return x

    def vn_sortkey_word(word):
        word = vn_norm(word)
        # Try to match the first character (or digraph) to the order
        if not word:
            return (len(vn_order), word)
        # Try 2-char and 3-char digraphs first
        for l in (3, 2, 1):
            prefix = word[:l]
            if prefix in vn_order_map:
                return (vn_order_map[prefix], word)
        # Fallback: use first char
        return (vn_order_map.get(word[0], len(vn_order)), word)

    def vn_sort_key(s):
        name = getattr(s, "Name", "")
        sid = str(getattr(s, "Student ID", "")).strip()
        parts = name.strip().split()
        if not parts:
            first = ""
            family = ""
        else:
            first = parts[-1]
            family = " ".join(parts[:-1])
        # Sort by first name (Vietnamese order), then family name, then sid
        return (vn_sortkey_word(first), vn_sortkey_word(family), sid)

    sorted_students = sorted(students, key=vn_sort_key)
    if verbose:
        print("[SortStudents] Sorted students by Vietnamese alphabetical order of first name, then family name, then student id.")
        for idx, s in enumerate(sorted_students, 1):
            print(f"  {idx}. {getattr(s, 'Name', '')} ({getattr(s, 'Student ID', '')})")
    else:
        print(f"Sorted {len(sorted_students)} students by Vietnamese alphabetical order.")
    return sorted_students

def refine_database(students, verbose=False):
    """
    Remove students with strange names like 'họ và tên', 'test student', 'possible points', etc.
    Handles names with multiple spaces between words.
    Sort students by Vietnamese alphabetical order: first by first name (last word), then by family name, then by student id.
    Returns the refined list of students.
    """

    # List of strange names (lowercase, stripped, normalized spaces)
    bad_names = {
        "họ và tên", "họ tên", "test student", "possible points", "student name",
        "full name", "name", "tên", "points possible", "student", "test", "sample",
        "họ tên lớp", "mã sv", "ho tên lớp mã sv"
    }
    # Compile regex patterns for matching strange names (allow multiple spaces)
    patterns = [
        re.compile(r"^\s*(họ\s+và\s+tên|họ\s+tên|test\s+student|possible\s+points|student\s+name|full\s+name|name|tên|points\s+possible|student|test|sample|họ\s+tên\s+lớp|mã\s+sv)\s*$", re.IGNORECASE),
        re.compile(r"possible\s+points", re.IGNORECASE),
        re.compile(r"test", re.IGNORECASE),
        re.compile(r"sample", re.IGNORECASE),
        re.compile(r"^\s*student\s*$", re.IGNORECASE),
        re.compile(r"^\s*sv\s*$", re.IGNORECASE),
        re.compile(r"^\s*hocvien\s*$", re.IGNORECASE),
        re.compile(r"ho\s+tên\s+lớp\s+mã\s+sv", re.IGNORECASE),
        re.compile(r"[•\*\-▪■●○□◆◇]", re.IGNORECASE),  # Any name containing a bullet point character
    ]

    def normalize_spaces(s):
        return re.sub(r'\s+', ' ', str(s)).strip().lower()

    def is_strange_name(name):
        if not name:
            return True
        n = normalize_spaces(name)
        if n in bad_names or len(n) < 3:
            return True
        for pat in patterns:
            if pat.search(n):
                return True
        return False

    refined = [s for s in students if not is_strange_name(getattr(s, "Name", ""))]
    removed = len(students) - len(refined)
    if removed > 0:
        if verbose:
            print(f"[RefineDatabase] Removed {removed} student(s) with strange names from database.")
        else:
            print(f"Notice: Removed {removed} student(s) with strange names from database.")

    # Sort students by Vietnamese alphabetical order: first name, then family name, then student id
    def vn_sort_key(s):
        name = getattr(s, "Name", "")
        sid = str(getattr(s, "Student ID", "")).strip()
        parts = name.strip().split()
        if not parts:
            first = ""
            family = ""
        else:
            first = parts[-1]
            family = " ".join(parts[:-1])
        return (first.lower(), family.lower(), sid)
    refined_sorted = sorted(refined, key=vn_sort_key)
    return refined_sorted

def _silent_load_database(db_path):
    try:
        class _LegacyStudentUnpickler(pickle.Unpickler):
            def find_class(self, module, name):
                if module == "__main__" and name == "Student":
                    from .models import Student
                    return Student
                return super().find_class(module, name)

        with open(db_path, "rb") as f:
            return _LegacyStudentUnpickler(f).load()
    except Exception:
        return []


def _values_equal(old_value, new_value):
    if old_value is None and new_value is None:
        return True
    if isinstance(old_value, str) and not old_value.strip() and new_value is None:
        return True
    if isinstance(new_value, str) and not new_value.strip() and old_value is None:
        return True
    try:
        old_num = float(str(old_value).replace(",", ".")) if old_value is not None else None
        new_num = float(str(new_value).replace(",", ".")) if new_value is not None else None
        if old_num is not None and new_num is not None:
            return abs(old_num - new_num) < 1e-6
    except Exception:
        pass
    return str(old_value).strip() == str(new_value).strip()


def _append_grade_audit(students, db_path, audit_source=None, verbose=False):
    if not GRADE_AUDIT_ENABLED:
        return
    if not os.path.exists(db_path):
        return
    previous_students = _silent_load_database(db_path)
    if not previous_students:
        return
    prev_by_id = {}
    prev_by_name = {}
    for s in previous_students:
        sid = _normalize_student_id(getattr(s, "Student ID", ""))
        name = _normalize_vietnamese_name(getattr(s, "Name", ""))
        if sid:
            prev_by_id[sid] = s
        if name:
            prev_by_name[name] = s
    source = audit_source or "unknown"
    timestamp = datetime.now().isoformat(timespec="seconds")
    updates = 0
    for student in students:
        sid = _normalize_student_id(getattr(student, "Student ID", ""))
        name = _normalize_vietnamese_name(getattr(student, "Name", ""))
        previous = prev_by_id.get(sid) or prev_by_name.get(name)
        if not previous:
            continue
        for field in GRADE_AUDIT_FIELDS:
            old_val = getattr(previous, field, None)
            new_val = getattr(student, field, None)
            if _values_equal(old_val, new_val):
                continue
            audit_list = getattr(student, "Grade Audit", None)
            if not isinstance(audit_list, list):
                audit_list = []
            audit_list.append({
                "timestamp": timestamp,
                "field": field,
                "old": old_val,
                "new": new_val,
                "source": source,
            })
            setattr(student, "Grade Audit", audit_list)
            updates += 1
    if verbose and updates:
        print(f"[GradeAudit] Recorded {updates} grade change(s).")


def _list_db_backups(backup_dir, base_name, ext):
    pattern = os.path.join(backup_dir, f"{base_name}_backup_*{ext}")
    return sorted(glob.glob(pattern), key=lambda p: os.path.getmtime(p))


def _cleanup_db_backups(backup_dir, base_name, ext, keep=5, verbose=False):
    if keep is None:
        return []
    try:
        keep = int(keep)
    except (TypeError, ValueError):
        return []
    backups = _list_db_backups(backup_dir, base_name, ext)
    if keep < 0 or len(backups) <= keep:
        return []
    to_remove = backups[:len(backups) - keep]
    removed = []
    for path in to_remove:
        try:
            os.remove(path)
            removed.append(path)
        except OSError as e:
            if verbose:
                print(f"[DBBackup] Failed to remove old backup {path}: {e}")
    return removed


def backup_database(db_path=None, backup_dir=None, keep=None, verbose=False):
    if not db_path:
        db_path = get_default_db_path()
    if not os.path.exists(db_path):
        if verbose:
            print(f"[DBBackup] Database not found at {db_path}")
        else:
            print(f"Database not found at {db_path}")
        return None
    backup_dir = backup_dir or os.path.dirname(db_path) or "."
    os.makedirs(backup_dir, exist_ok=True)
    base = os.path.splitext(os.path.basename(db_path))[0]
    ext = os.path.splitext(db_path)[1]
    now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(backup_dir, f"{base}_backup_{now_str}{ext}")
    if DRY_RUN:
        print(f"[DBBackup] Dry run: would back up database to {backup_path}")
        return backup_path
    try:
        shutil.copy2(db_path, backup_path)
        if verbose:
            print(f"[DBBackup] Backed up database to {backup_path}")
        else:
            print(f"Database backup created at {backup_path}")
        append_run_report("backup-db", outputs=backup_path, verbose=verbose)
    except OSError as e:
        print(f"[DBBackup] Failed to back up database: {e}")
        return None
    _cleanup_db_backups(backup_dir, base, ext, keep=keep if keep is not None else DB_BACKUP_KEEP, verbose=verbose)
    return backup_path


def restore_database(db_path=None, backup_path=None, verbose=False):
    if not db_path:
        db_path = get_default_db_path()
    backup_dir = os.path.dirname(db_path) or "."
    base = os.path.splitext(os.path.basename(db_path))[0]
    ext = os.path.splitext(db_path)[1]
    if not backup_path or backup_path == "latest":
        backups = _list_db_backups(backup_dir, base, ext)
        if not backups:
            print("No database backups found.")
            return None
        backup_path = backups[-1]
    if not os.path.exists(backup_path):
        print(f"Database backup not found: {backup_path}")
        return None
    if DRY_RUN:
        print(f"[DBBackup] Dry run: would restore database from {backup_path}")
        return backup_path
    os.makedirs(os.path.dirname(db_path) or ".", exist_ok=True)
    try:
        shutil.copy2(backup_path, db_path)
    except OSError as e:
        print(f"[DBBackup] Failed to restore database: {e}")
        return None
    if verbose:
        print(f"[DBBackup] Restored database from {backup_path}")
    else:
        print(f"Database restored from {backup_path}")
    append_run_report("restore-db", outputs=backup_path, verbose=verbose)
    return backup_path


def save_database(students, db_path, verbose=False, audit_source=None):
    # Backup old database if it exists
    if DRY_RUN:
        print(f"[SaveDatabase] Dry run: would save database to {db_path}")
        return
    if os.path.exists(db_path):
        backup_database(db_path, verbose=verbose)

    if verbose:
        print(f"[SaveDatabase] Saving database to {db_path}...")
    else:
        print(f"Saving database to {db_path}...")
    # Refine database before saving
    _append_grade_audit(students, db_path, audit_source=audit_source, verbose=verbose)
    students_refined = refine_database(students, verbose=verbose)
    with open(db_path, 'wb') as f:
        pickle.dump(students_refined, f)
    append_run_report(
        "save-database",
        details=f"students={len(students_refined)}",
        outputs=db_path,
        verbose=verbose,
    )

def load_database(db_path, verbose=False):
    if not os.path.exists(db_path):
        if verbose:
            print(f"[LoadDatabase] Database file not found at {db_path}. Returning empty list.")
        else:
            print(f"Notice: Database file not found at {db_path}.")
        return []
    if verbose:
        print(f"[LoadDatabase] Loading database from {db_path}...")
    else:
        print(f"Loading database from {db_path}...")
    class _LegacyStudentUnpickler(pickle.Unpickler):
        def find_class(self, module, name):
            if module == "__main__" and name == "Student":
                from .models import Student
                return Student
            return super().find_class(module, name)

    with open(db_path, 'rb') as f:
        return _LegacyStudentUnpickler(f).load()


def generate_data_validation_report(students=None, db_path=None, output_path=None, verbose=False):
    """
    Validate student data for missing IDs, duplicate IDs/names, invalid dates, and out-of-range grades.
    Returns a dict report and optionally writes a text report to disk.
    """
    if db_path:
        students = load_database(db_path, verbose=verbose)
    students = students or []
    if not students:
        if verbose:
            print("[ValidateData] No students available for validation.")
        else:
            print("No students available for validation.")
        return {}

    missing_ids = []
    ids_to_names = {}
    names_to_ids = {}
    invalid_dobs = []
    out_of_range = []

    grade_fields = ["CC", "GK", "CK", "Attendance", "Midterm", "Final"]
    dob_fields = ["Dob", "Date of Birth", "Ngày sinh"]
    date_formats = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%m/%d/%Y"]

    for student in students:
        sid_raw = getattr(student, "Student ID", "")
        name_raw = getattr(student, "Name", "")
        sid = _normalize_student_id(sid_raw)
        name = str(name_raw).strip()
        if not sid:
            missing_ids.append({"name": name, "student_id": sid_raw})
        else:
            ids_to_names.setdefault(sid, set()).add(name or "(missing name)")
        if name:
            names_to_ids.setdefault(_normalize_vietnamese_name(name), set()).add(sid or "(missing id)")

        dob_value = None
        for field in dob_fields:
            if hasattr(student, field):
                dob_value = getattr(student, field)
                if dob_value:
                    break
        if dob_value:
            parsed = False
            for fmt in date_formats:
                try:
                    datetime.strptime(str(dob_value).strip(), fmt)
                    parsed = True
                    break
                except Exception:
                    continue
            if not parsed:
                invalid_dobs.append({"name": name, "student_id": sid, "value": dob_value})

        for field in grade_fields:
            if not hasattr(student, field):
                continue
            value = getattr(student, field)
            if value is None or (isinstance(value, str) and not value.strip()):
                continue
            try:
                num = float(str(value).replace(",", "."))
            except Exception:
                out_of_range.append({"name": name, "student_id": sid, "field": field, "value": value})
                continue
            if num < 0 or num > 10:
                out_of_range.append({"name": name, "student_id": sid, "field": field, "value": value})

    duplicate_ids = {sid: sorted(list(names)) for sid, names in ids_to_names.items() if len(names) > 1}
    duplicate_names = {name: sorted(list(ids)) for name, ids in names_to_ids.items() if len(ids) > 1}

    report = {
        "total_students": len(students),
        "missing_ids": missing_ids,
        "duplicate_ids": duplicate_ids,
        "duplicate_names": duplicate_names,
        "invalid_dobs": invalid_dobs,
        "out_of_range_grades": out_of_range,
    }

    lines = []
    lines.append("Data validation report")
    lines.append("-" * 24)
    lines.append(f"Total students: {len(students)}")
    lines.append(f"Missing IDs: {len(missing_ids)}")
    lines.append(f"Duplicate IDs: {len(duplicate_ids)}")
    lines.append(f"Duplicate names: {len(duplicate_names)}")
    lines.append(f"Invalid DOBs: {len(invalid_dobs)}")
    lines.append(f"Out-of-range grades: {len(out_of_range)}")
    lines.append("")

    if missing_ids:
        lines.append("Missing Student IDs:")
        for item in missing_ids if verbose else missing_ids[:10]:
            lines.append(f"  - {item.get('name') or '(missing name)'} | {item.get('student_id')}")
        if not verbose and len(missing_ids) > 10:
            lines.append(f"  ... ({len(missing_ids) - 10} more)")
        lines.append("")

    if duplicate_ids:
        lines.append("Duplicate Student IDs:")
        for sid, names in duplicate_ids.items():
            lines.append(f"  - {sid}: {', '.join(names)}")
        lines.append("")

    if duplicate_names:
        lines.append("Duplicate Names:")
        for name_key, ids in duplicate_names.items():
            lines.append(f"  - {name_key}: {', '.join(ids)}")
        lines.append("")

    if invalid_dobs:
        lines.append("Invalid DOBs:")
        for item in invalid_dobs if verbose else invalid_dobs[:10]:
            lines.append(f"  - {item.get('name') or '(missing name)'} ({item.get('student_id')}) -> {item.get('value')}")
        if not verbose and len(invalid_dobs) > 10:
            lines.append(f"  ... ({len(invalid_dobs) - 10} more)")
        lines.append("")

    if out_of_range:
        lines.append("Out-of-range grades:")
        for item in out_of_range if verbose else out_of_range[:10]:
            lines.append(f"  - {item.get('name') or '(missing name)'} ({item.get('student_id')}): {item.get('field')}={item.get('value')}")
        if not verbose and len(out_of_range) > 10:
            lines.append(f"  ... ({len(out_of_range) - 10} more)")
        lines.append("")

    report_text = "\n".join(lines)
    print(report_text)
    if output_path is None:
        output_path = os.path.join(os.getcwd(), "data_validation_report.txt")
    if DRY_RUN:
        print(f"[ValidateData] Dry run: would write report to {output_path}")
        return report
    try:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(report_text)
        if verbose:
            print(f"[ValidateData] Report saved to {output_path}")
    except Exception as e:
        print(f"[ValidateData] Failed to write report: {e}")
        return report
    append_run_report(
        "validate-data",
        details=f"students={len(students)}",
        outputs=output_path,
        verbose=verbose,
    )
    return report

def print_student_details(students, identifier, db_path=None, verbose=False):
    """
    Print details of a student by name, student id, or email.
    If verbose is True, print more details; otherwise, print only important notice.
    """
    if db_path:
        students = load_database(db_path, verbose=verbose)
    identifier = identifier.strip().lower()
    email_pattern = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")
    found = False
    for s in students:
        name = str(getattr(s, "Name", "")).strip().lower()
        sid = str(getattr(s, "Student ID", "")).strip().lower()
        email = None
        for v in s.__dict__.values():
            v_str = str(v).strip().lower()
            if email_pattern.match(v_str):
                email = v_str
                break
        if identifier == name or identifier == sid or (email and identifier == email):
            if verbose:
                print("[PrintStudentDetails] Student details:")
            else:
                print("Student details:")
            for k, v in s.__dict__.items():
                print(f"{k}: {v}")
            found = True
            break
    if not found:
        if verbose:
            print("[PrintStudentDetails] No student found with the given identifier.")
        else:
            print("No student found with the given identifier.")

def en_to_vn_field(field, verbose=False):
    """
    Map English field names to Vietnamese.
    Also handles "Blackboard Count: <date>" fields.
    Includes more Canvas/grade field translations.
    Translates Canvas section (registered class ID) and GitHub username fields.
    If verbose is True, print mapping details; otherwise, print only important notice if field is unmapped.
    """
    field_vn_map = {
        "Name": "Họ và Tên",
        "Student ID": "Mã sinh viên",
        "Total Blackboard Counts": "Tổng số lần lên bảng",
        "Max Total Blackboard Counts": "Số lần lên bảng nhiều nhất của một sinh viên",
        "Total Blackboard Counts until midterm": "Tổng số lần lên bảng đến giữa kỳ",
        "Max Total Blackboard Counts until midterm": "Số lần lên bảng nhiều nhất đến giữa kỳ của một sinh viên",
        "Midterm Reward Points": "Điểm cộng giữa kỳ",
        "Final Reward Points": "Điểm cộng cuối kỳ",
        "Email": "Email",
        "Class": "Lớp",
        "Phone": "Số điện thoại",
        "Dob": "Ngày sinh",
        "Gender": "Giới tính",
        "Attendance": "Chuyên cần",
        "Midterm": "Giữa kỳ",
        "Final": "Cuối kỳ",
        "Participation": "Điểm danh",
        "Assignment": "Bài tập",
        "Quiz": "Quiz",
        # Canvas/grade field names
        "Bài tập Final Score": "Bài tập (Canvas)",
        "Kiểm tra giữa kỳ Final Score": "Giữa kỳ (Canvas)",
        "Giữa kỳ Final Score": "Giữa kỳ (Canvas)",
        "Cuối kỳ Final Score": "Cuối kỳ (Canvas)",
        "Midterm Final Score": "Giữa kỳ (Canvas)",
        "Attendance Final Score": "Điểm danh (Canvas)",
        "Chuyên cần Final Score": "Chuyên cần (Canvas)",
        "Quiz Final Score": "Quiz (Canvas)",
        "Điểm danh Final Score": "Điểm danh (Canvas)",
        "Assignment Final Score": "Bài tập (Canvas)",
        "Participation Final Score": "Điểm danh (Canvas)",
        "Final Final Score": "Cuối kỳ (Canvas)",
        "Total Final Score": "Tổng điểm (Canvas)",
        "Điểm tổng kết": "Tổng kết",
        "Điểm tổng": "Tổng điểm",
        "Điểm": "Điểm",
        # Canvas section / registered class ID fields
        "Canvas Section": "Lớp học phần",
        "Registration ID": "Lớp học phần",
        "Registered Class ID": "Lớp học phần",
        "Class ID": "Mã lớp",
        "Section ID": "Mã section",
        # GitHub username fields
        "GitHub Username": "Tên người dùng GitHub",
        "GitHub Account": "Tài khoản GitHub",
        "GitHub ID": "ID GitHub",
        "GitHub Handle": "Tên đăng nhập GitHub",
        "GitHub": "GitHub",
        "Google_ID": "Google ID",
        "Google_Classroom_Display_Name": "Google Classroom Display Name",
    }
    if field.startswith("Blackboard Count: "):
        date = field[len("Blackboard Count: "):]
        vn_field = f"Số lần lên bảng: {date}"
        if verbose:
            print(f"[en_to_vn_field] Mapped '{field}' to '{vn_field}'")
        return vn_field
    vn_field = field_vn_map.get(field, field)
    if verbose:
        if field in field_vn_map:
            print(f"[en_to_vn_field] Mapped '{field}' to '{vn_field}'")
        else:
            print(f"[en_to_vn_field] No mapping found for '{field}', using original.")
    elif field not in field_vn_map and not field.startswith("Blackboard Count: "):
        print(f"Notice: No Vietnamese mapping for field '{field}', using original.")
    return vn_field

def print_all_student_details(students, db_path=None, verbose=False, sort_method=None):
    """
    Print details of all students, separated by a series of '-'.
    Field names are translated into Vietnamese.
    If verbose is True, print more details; otherwise, print only important notice.
    """
    if db_path:
        students = load_database(db_path)
    if not students:
        if verbose:
            print("[print_all_student_details] No students to display.")
        else:
            print("No students to display.")
        return
    if verbose:
        print(f"[print_all_student_details] Tổng số sinh viên: {len(students)}")
    else:
        print(f"Tổng số sinh viên: {len(students)}")
    print('-' * 40)
    display_labels = {
        "Google_ID": "Google ID",
        "Google_Classroom_Display_Name": "Google Classroom Display Name",
    }
    vn_to_en = {
        "Họ và Tên": "Name",
        "Mã sinh viên": "Student ID",
        "Ngày sinh": "Date of Birth",
        "Lớp": "Class",
        "Điểm học phần": "Course Grade",
        "Tổng điểm": "Total Score",
        "Tổng điểm (DB)": "Total Score (DB)",
        "Tổng điểm (Canvas)": "Total Score (Canvas)",
    }
    sorted_students = sorted(students, key=lambda s: _get_student_sort_key(s, sort_method))
    for idx, s in enumerate(sorted_students, 1):
        if verbose:
            print(f"[print_all_student_details] Sinh viên {idx}/{len(sorted_students)}:")
        else:
            print(f"Sinh viên {idx}/{len(sorted_students)}:")
        data = s.__dict__
        preferred_keys = [
            "Họ và Tên", "Name",
            "Mã sinh viên", "Student ID",
            "STT",
            "Ngày sinh", "Date of Birth", "DOB",
            "Lớp", "Class",
            "Điểm học phần", "Course Grade",
            "Tổng điểm (DB)", "Total Score",
            "Tổng điểm (Canvas)", "Total Final Score",
            "Email",
            "Google_ID",
            "Google_Classroom_Display_Name",
        ]
        seen = set()
        ordered_keys = []
        for key in preferred_keys:
            if key in data and key not in seen:
                ordered_keys.append(key)
                seen.add(key)
        remaining_keys = [k for k in data.keys() if k not in seen and k not in ("Grades", "Submissions", "Canvas Submission Comments", "Canvas Rubric Evaluations")]
        for key in sorted(remaining_keys):
            ordered_keys.append(key)
        for k in ordered_keys:
            v = data.get(k)
            if isinstance(k, str) and k in display_labels:
                en_k = display_labels[k]
            elif isinstance(k, str) and k in vn_to_en:
                en_k = vn_to_en[k]
            else:
                en_k = k if isinstance(k, str) else str(k)
            vn_k = en_to_vn_field(k, verbose=verbose) if isinstance(k, str) else str(k)
            print(f"{en_k} ({vn_k}): {v}")
        grades = data.get("Grades")
        if isinstance(grades, dict):
            total_gc = 0.0
            total_gc_max = 0.0
            print("Grades (Điểm):")
            for title in sorted(grades.keys()):
                info = grades.get(title) or {}
                if isinstance(info, dict):
                    grade = info.get("grade")
                    max_points = info.get("max_points")
                    if grade is not None and max_points is not None:
                        value = f"{grade}/{max_points}"
                        try:
                            total_gc += float(grade)
                            total_gc_max += float(max_points)
                        except (TypeError, ValueError):
                            pass
                    elif grade is not None:
                        value = f"{grade}"
                        try:
                            total_gc += float(grade)
                        except (TypeError, ValueError):
                            pass
                    else:
                        value = str(info)
                else:
                    value = str(info)
                print(f"  - {title}: {value}")
            if total_gc_max > 0:
                print(f"Total score (Classroom) (Tổng điểm Classroom): {total_gc}/{total_gc_max}")
            elif total_gc > 0:
                print(f"Total score (Classroom) (Tổng điểm Classroom): {total_gc}")
        submissions = data.get("Submissions")
        if isinstance(submissions, dict):
            print("Submissions (Nộp bài):")
            for title in sorted(submissions.keys()):
                state = submissions.get(title)
                print(f"  - {title}: {state}")
        comments = data.get("Canvas Submission Comments")
        if isinstance(comments, dict):
            print("Canvas submission comments (Nhận xét bài nộp Canvas):")
            for title in sorted(comments.keys()):
                items = comments.get(title) or []
                print(f"  - {title}: {len(items)} comment(s)")
                for item in items:
                    author = item.get("author_name") if isinstance(item, dict) else None
                    content = item.get("comment") if isinstance(item, dict) else str(item)
                    posted_at = item.get("posted_at") if isinstance(item, dict) else None
                    meta = f" ({posted_at})" if posted_at else ""
                    label = author or "Unknown"
                    print(f"      * {label}{meta}: {content}")
        rubrics = data.get("Canvas Rubric Evaluations")
        if isinstance(rubrics, dict):
            print("Canvas rubric evaluations (Đánh giá rubrics Canvas):")
            for title in sorted(rubrics.keys()):
                rubric = rubrics.get(title)
                if isinstance(rubric, dict):
                    print(f"  - {title}: {len(rubric)} criterion")
                    for key, info in rubric.items():
                        points = info.get("points") if isinstance(info, dict) else None
                        rating = info.get("rating") if isinstance(info, dict) else None
                        comment = info.get("comments") if isinstance(info, dict) else None
                        parts = []
                        if points is not None:
                            parts.append(f"points={points}")
                        if rating:
                            parts.append(f"rating={rating}")
                        if comment:
                            parts.append(f"comment={comment}")
                        detail = "; ".join(parts) if parts else str(info)
                        print(f"      * {key}: {detail}")
                else:
                    print(f"  - {title}: {rubric}")
        print('-' * 40)

def export_to_excel(students, file_path=None, db_path=None, verbose=False):
    # If db_path is provided, load students from database
    if db_path:
        students = load_database(db_path)
    if not students:
        if verbose:
            print("[ExportExcel] No students to export.")
        else:
            print("No students to export.")
        return
    if not file_path:
        file_path = os.path.join(os.getcwd(), "students_export.xlsx")
    if DRY_RUN:
        print(f"[ExportExcel] Dry run: would export to {file_path}")
        return
    if verbose:
        print(f"[ExportExcel] Exporting {len(students)} students to Excel file: {file_path}")
    else:
        print(f"Exporting {len(students)} students to Excel...")

    # Only export selected fields
    export_fields = [
        "Name",
        "Student ID",
        "Total Blackboard Counts",
        "Max Total Blackboard Counts",
        "Midterm Reward Points",
        "Final Reward Points"
    ]
    # Translate field names to Vietnamese
    export_fields_vn = [en_to_vn_field(field) for field in export_fields]

    rows = []
    for s in tqdm(students, desc="Exporting"):
        row = {en_to_vn_field(field): getattr(s, field, "") for field in export_fields}
        rows.append(row)
    df = pd.DataFrame(rows, columns=export_fields_vn)
    df.to_excel(file_path, index=False)

    # Auto-fit column widths and center numeric columns
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Identify numeric columns (by checking the first non-header row)
    numeric_cols = set()
    for idx, col in enumerate(ws.iter_cols(min_row=2, max_row=ws.max_row), 1):
        for cell in col:
            val = cell.value
            if isinstance(val, (int, float)):
                numeric_cols.add(idx)
                break
            # Also treat strings that look like numbers as numeric
            if isinstance(val, str):
                try:
                    float(val)
                    numeric_cols.add(idx)
                    break
                except Exception:
                    continue

    for col in ws.columns:
        max_length = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        col_idx = col[0].column
        for cell in col:
            try:
                cell_length = len(str(cell.value)) if cell.value is not None else 0
                if cell_length > max_length:
                    max_length = cell_length
                # Center align numeric columns (skip header row)
                if cell.row > 1 and col_idx in numeric_cols:
                    cell.alignment = Alignment(horizontal='center')
            except Exception:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width
    wb.save(file_path)

    if verbose:
        print(f"[ExportExcel] Exported {len(students)} students to {file_path}")
    else:
        print(f"Exported {len(students)} students to {file_path}")
    append_run_report(
        "export-excel",
        details=f"students={len(students)}",
        outputs=file_path,
        verbose=verbose,
    )


def _anonymize_student(name, student_id):
    seed = f"{str(name).strip()}|{str(student_id).strip()}"
    digest = hashlib.sha256(seed.encode("utf-8")).hexdigest()
    anon_name = f"Student-{digest[:6]}"
    anon_id = f"ID-{digest[6:14]}"
    return anon_name, anon_id


def export_anonymized_roster(students, file_path=None, db_path=None, verbose=False):
    if db_path:
        students = load_database(db_path)
    if not students:
        if verbose:
            print("[ExportAnon] No students to export.")
        else:
            print("No students to export.")
        return
    if not file_path:
        file_path = os.path.join(os.getcwd(), "students_anonymized.csv")
    if DRY_RUN:
        print(f"[ExportAnon] Dry run: would export anonymized roster to {file_path}")
        return

    columns = set()
    for s in students:
        columns.update(s.__dict__.keys())
    ordered_columns = []
    for key in ["Name", "Student ID"]:
        if key in columns:
            ordered_columns.append(key)
    ordered_columns.extend(sorted([c for c in columns if c not in ordered_columns]))

    rows = []
    for s in students:
        row = {col: getattr(s, col, "") for col in ordered_columns}
        anon_name, anon_id = _anonymize_student(row.get("Name", ""), row.get("Student ID", ""))
        if "Name" in row:
            row["Name"] = anon_name
        if "Student ID" in row:
            row["Student ID"] = anon_id
        rows.append(row)

    df = pd.DataFrame(rows, columns=ordered_columns)
    df.to_csv(file_path, index=False)
    if verbose:
        print(f"[ExportAnon] Exported {len(students)} students to {file_path}")
    else:
        print(f"Exported {len(students)} students to {file_path}")
    append_run_report(
        "export-anonymized",
        details=f"students={len(students)}",
        outputs=file_path,
        verbose=verbose,
    )

def search_students(students, query, db_path=None, verbose=False):
    """
    Search students by query string in any field.
    If verbose is True, print more details; otherwise, print only important notice.
    """
    # If db_path is provided, load students from database
    if db_path:
        students = load_database(db_path)
    query = query.lower()
    results = []
    for s in tqdm(students, desc="Searching"):
        for value in s.__dict__.values():
            if query in str(value).lower():
                results.append(s)
                break
    if verbose:
        print(f"[search_students] Query: '{query}', found {len(results)} result(s).")
    elif results:
        print(f"Found {len(results)} student(s) matching your query.")
    else:
        print("No student found matching your query.")
    return results

def export_emails_to_txt(students, file_path, db_path=None, verbose=False):
    """
    Export all student emails to a TXT file, avoiding duplicates with previously exported emails.
    If verbose is True, print more details; otherwise, print only important notice.
    """
    # If db_path is provided, load students from database
    if db_path:
        students = load_database(db_path)

    if DRY_RUN:
        print(f"[ExportEmails] Dry run: would export emails to {file_path}")
        return

    # Gather all previously exported emails from emails_*.txt files
    prev_emails = set()
    dir_path = os.path.dirname(file_path) or "."
    for fname in glob.glob(os.path.join(dir_path, "emails_*.txt")):
        try:
            with open(fname, "r", encoding="utf-8") as f:
                content = f.read()
                prev_emails.update([e.strip() for e in content.split(",") if e.strip()])
        except Exception:
            continue

    # Collect emails from current students, skipping those already exported
    emails = []
    email_pattern = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")
    for s in students:
        for attr, value in s.__dict__.items():
            value_str = str(value).strip()
            if value_str and email_pattern.match(value_str):
                if value_str not in prev_emails:
                    emails.append(value_str)
                break  # Only take the first matching email attribute per student

    if not emails:
        if verbose:
            print("[ExportEmails] No new emails to export.")
        else:
            print("No new emails to export.")
        return

    # Write to a new file with today's date, append if exists
    today_str = datetime.now().strftime("%Y%m%d")
    out_path = os.path.join(dir_path, f"emails_{today_str}.txt")
    if os.path.exists(out_path):
        with open(out_path, 'a', encoding='utf-8') as f:
            if os.path.getsize(out_path) > 0:
                f.write(',')
            f.write(','.join(emails))
        if verbose:
            print(f"[ExportEmails] Appended {len(emails)} new emails to {out_path}")
        else:
            print(f"Appended {len(emails)} new emails to {out_path}")
    else:
        with open(out_path, 'w', encoding='utf-8') as f:
            f.write(','.join(emails))
        if verbose:
            print(f"[ExportEmails] Exported {len(emails)} new emails to {out_path}")
        else:
            print(f"Exported {len(emails)} new emails to {out_path}")
    append_run_report(
        "export-emails",
        details=f"emails={len(emails)}",
        outputs=out_path,
        verbose=verbose,
    )

def export_emails_and_names_to_txt(students, file_path=None, db_path=None, verbose=False):
    """
    Export all student (name, email, role, section) 4-tuples to a TXT file, avoiding duplicates.
    Role is always set to "student".
    Section information (Canvas Section) is included if available.
    If verbose is True, print more details; otherwise, print only important notice.
    """
    # If db_path is provided, load students from database
    if db_path:
        students = load_database(db_path)

    if DRY_RUN:
        print(f"[ExportEmailsNames] Dry run: would export to {file_path}")
        return

    # Gather all previously exported emails from emails_*.txt files
    prev_emails = set()
    dir_path = os.path.dirname(file_path) or "."
    for fname in glob.glob(os.path.join(dir_path, "emails_*.txt")):
        try:
            with open(fname, "r", encoding="utf-8") as f:
                for line in f:
                    parts = line.strip().split(",")
                    if len(parts) >= 2:  # Could be name,email or name,email,role,section
                        prev_emails.add(parts[1].strip())
        except Exception:
            continue

    # Collect (name, email, role, section) 4-tuples from current students, skipping those already exported
    tuples = []
    email_pattern = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")
    
    for s in students:
        name = getattr(s, "Name", "")
        email = None
        role = "student"  # Always set role to "student"
        section = ""  # Default empty section
        
        # Find the first valid email
        for attr, value in s.__dict__.items():
            value_str = str(value).strip()
            if value_str and email_pattern.match(value_str):
                email = value_str
                break  # Only take the first matching email attribute per student
        
        # Find section information (check multiple possible field names)
        for section_field in ["Canvas Section", "Section ID", "Registration ID", 
                             "Registered Class ID", "Class ID"]:
            if hasattr(s, section_field) and getattr(s, section_field):
                section = str(getattr(s, section_field)).strip()
                break
                
        if email and email not in prev_emails:
            tuples.append((name, email, role, section))

    if not tuples:
        if verbose:
            print("[ExportEmailsNames] No new emails to export.")
        else:
            print("No new emails to export.")
        return

    # Write to a new file with today's date, append if exists
    today_str = datetime.now().strftime("%Y%m%d")
    if os.path.exists(file_path):
        out_path = file_path
    else:
        out_path = os.path.join(dir_path, f"emails_{today_str}.txt")
    mode = 'a' if os.path.exists(out_path) else 'w'
    with open(out_path, mode, encoding='utf-8') as f:
        for name, email, role, section in tuples:
            f.write(f"{name},{email},{role},{section}\n")
    if verbose:
        print(f"[ExportEmailsNames] Exported {len(tuples)} new name,email,role,section 4-tuples to {out_path}")
    else:
        print(f"Exported {len(tuples)} new name,email,role,section 4-tuples to {out_path}")
    append_run_report(
        "export-emails-names",
        details=f"rows={len(tuples)}",
        outputs=out_path,
        verbose=verbose,
    )

def export_all_details_to_txt(students, file_path=None, db_path=None, verbose=False, sort_method=None):
    # If db_path is provided, load students from database
    if db_path:
        students = load_database(db_path)
    if not students:
        if verbose:
            print("[ExportAllDetails] No students to export.")
        else:
            print("No students to export.")
        return
    if not file_path:
        file_path = os.path.join(os.getcwd(), "students_details.txt")
    if DRY_RUN:
        print(f"[ExportAllDetails] Dry run: would export details to {file_path}")
        return
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(f"Tổng số sinh viên: {len(students)}\n")
        f.write('-' * 40 + '\n')
        display_labels = {
            "Google_ID": "Google ID",
            "Google_Classroom_Display_Name": "Google Classroom Display Name",
        }
        vn_to_en = {
            "Họ và Tên": "Name",
            "Mã sinh viên": "Student ID",
            "Ngày sinh": "Date of Birth",
            "Lớp": "Class",
            "Điểm học phần": "Course Grade",
            "Tổng điểm": "Total Score",
            "Tổng điểm (DB)": "Total Score (DB)",
            "Tổng điểm (Canvas)": "Total Score (Canvas)",
        }
        sorted_students = sorted(students, key=lambda s: _get_student_sort_key(s, sort_method))
        for idx, s in enumerate(sorted_students, 1):
            f.write(f"Sinh viên {idx}/{len(sorted_students)}:\n")
            data = s.__dict__
            preferred_keys = [
                "Họ và Tên", "Name",
                "Mã sinh viên", "Student ID",
                "STT",
                "Ngày sinh", "Date of Birth", "DOB",
                "Lớp", "Class",
                "Điểm học phần", "Course Grade",
                "Tổng điểm (DB)", "Total Score",
                "Tổng điểm (Canvas)", "Total Final Score",
                "Email",
                "Google_ID",
                "Google_Classroom_Display_Name",
            ]
            seen = set()
            ordered_keys = []
            for key in preferred_keys:
                if key in data and key not in seen:
                    ordered_keys.append(key)
                    seen.add(key)
            remaining_keys = [k for k in data.keys() if k not in seen and k not in ("Grades", "Submissions", "Canvas Submission Comments", "Canvas Rubric Evaluations")]
            for key in sorted(remaining_keys):
                ordered_keys.append(key)
            for k in ordered_keys:
                v = data.get(k)
                if isinstance(k, str) and k in display_labels:
                    en_k = display_labels[k]
                elif isinstance(k, str) and k in vn_to_en:
                    en_k = vn_to_en[k]
                else:
                    en_k = k if isinstance(k, str) else str(k)
                vn_k = en_to_vn_field(k, verbose=verbose) if isinstance(k, str) else str(k)
                f.write(f"{en_k} ({vn_k}): {v}\n")
            grades = data.get("Grades")
            if isinstance(grades, dict):
                total_gc = 0.0
                total_gc_max = 0.0
                f.write("Grades (Điểm):\n")
                for title in sorted(grades.keys()):
                    info = grades.get(title) or {}
                    if isinstance(info, dict):
                        grade = info.get("grade")
                        max_points = info.get("max_points")
                        if grade is not None and max_points is not None:
                            value = f"{grade}/{max_points}"
                            try:
                                total_gc += float(grade)
                                total_gc_max += float(max_points)
                            except (TypeError, ValueError):
                                pass
                        elif grade is not None:
                            value = f"{grade}"
                            try:
                                total_gc += float(grade)
                            except (TypeError, ValueError):
                                pass
                        else:
                            value = str(info)
                    else:
                        value = str(info)
                    f.write(f"  - {title}: {value}\n")
                if total_gc_max > 0:
                    f.write(f"Total score (Classroom) (Tổng điểm Classroom): {total_gc}/{total_gc_max}\n")
                elif total_gc > 0:
                    f.write(f"Total score (Classroom) (Tổng điểm Classroom): {total_gc}\n")
            submissions = data.get("Submissions")
            if isinstance(submissions, dict):
                f.write("Submissions (Nộp bài):\n")
                for title in sorted(submissions.keys()):
                    state = submissions.get(title)
                    f.write(f"  - {title}: {state}\n")
            comments = data.get("Canvas Submission Comments")
            if isinstance(comments, dict):
                f.write("Canvas submission comments (Nhận xét bài nộp Canvas):\n")
                for title in sorted(comments.keys()):
                    items = comments.get(title) or []
                    f.write(f"  - {title}: {len(items)} comment(s)\n")
                    for item in items:
                        author = item.get("author_name") if isinstance(item, dict) else None
                        content = item.get("comment") if isinstance(item, dict) else str(item)
                        posted_at = item.get("posted_at") if isinstance(item, dict) else None
                        meta = f" ({posted_at})" if posted_at else ""
                        label = author or "Unknown"
                        f.write(f"      * {label}{meta}: {content}\n")
            rubrics = data.get("Canvas Rubric Evaluations")
            if isinstance(rubrics, dict):
                f.write("Canvas rubric evaluations (Đánh giá rubrics Canvas):\n")
                for title in sorted(rubrics.keys()):
                    rubric = rubrics.get(title)
                    if isinstance(rubric, dict):
                        f.write(f"  - {title}: {len(rubric)} criterion\n")
                        for key, info in rubric.items():
                            points = info.get("points") if isinstance(info, dict) else None
                            rating = info.get("rating") if isinstance(info, dict) else None
                            comment = info.get("comments") if isinstance(info, dict) else None
                            parts = []
                            if points is not None:
                                parts.append(f"points={points}")
                            if rating:
                                parts.append(f"rating={rating}")
                            if comment:
                                parts.append(f"comment={comment}")
                            detail = "; ".join(parts) if parts else str(info)
                            f.write(f"      * {key}: {detail}\n")
                    else:
                        f.write(f"  - {title}: {rubric}\n")
            f.write('-' * 40 + '\n')
    if verbose:
        print(f"[ExportAllDetails] Exported all student details to {file_path}")
    else:
        print(f"Exported all student details to {file_path}")
    append_run_report(
        "export-all-details",
        details=f"students={len(students)}",
        outputs=file_path,
        verbose=verbose,
    )

def _preprocess_ocr_image(image, verbose=False):
    try:
        img = image.convert("L")
        img = ImageOps.autocontrast(img)
        img = img.filter(ImageFilter.MedianFilter(size=3))
        img = img.filter(ImageFilter.UnsharpMask(radius=1, percent=150, threshold=3))
        np_img = np.array(img)
        if np_img.size == 0:
            return img
        hist, _ = np.histogram(np_img.flatten(), bins=256, range=(0, 256))
        total = np_img.size
        sum_total = np.dot(np.arange(256), hist)
        sum_b = 0.0
        w_b = 0.0
        var_max = 0.0
        threshold = 0
        for i in range(256):
            w_b += hist[i]
            if w_b == 0:
                continue
            w_f = total - w_b
            if w_f == 0:
                break
            sum_b += i * hist[i]
            m_b = sum_b / w_b
            m_f = (sum_total - sum_b) / w_f
            var_between = w_b * w_f * (m_b - m_f) ** 2
            if var_between > var_max:
                var_max = var_between
                threshold = i
        bw = (np_img > threshold).astype(np.uint8) * 255
        return Image.fromarray(bw)
    except Exception as exc:
        if verbose:
            print(f"[OCRPreprocess] Failed to preprocess image: {exc}")
        return image


def _write_pdf_chunk(images, output_path, max_size_bytes, verbose=False):
    if not images:
        return None
    scales = [1.0, 0.85, 0.7, 0.6]
    for scale in scales:
        if scale != 1.0:
            resized = []
            for img in images:
                w, h = img.size
                resized.append(img.resize((int(w * scale), int(h * scale)), Image.LANCZOS))
            save_images = resized
        else:
            save_images = images
        save_images = [img.convert("RGB") for img in save_images]
        save_images[0].save(
            output_path,
            "PDF",
            save_all=True,
            append_images=save_images[1:],
            resolution=200.0,
        )
        try:
            size = os.path.getsize(output_path)
        except Exception:
            size = max_size_bytes + 1
        if size <= max_size_bytes:
            return output_path
        if verbose:
            print(f"[OCRPreprocess] Chunk too large at scale {scale}: {size} bytes.")
    return output_path

def extract_text_from_scanned_pdf_ocrspace(pdf_path, txt_output_path=None, lang="auto", simple_text=False, verbose=False):

    """
    Extract text from a scanned PDF file using the free OCR.space API.
    Splits the PDF into smaller PDFs with at least 1 page, at most 3 pages, and each chunk not exceeding 1024 KB,
    sends requests via the API, receives results, and combines all into one text file.
    Keeps the overlay exactly as in the PDF: texts on the same line must stay on the same line.
    Requires pdf2image and PIL.
    Uses the JSON response structure to extract and combine parsed texts.
    Returns the path to the text file.
    """
    lang_map = {
        "eng": "eng",
        "en": "eng",
        "vie": "vie",
        "vi": "vie",
        "vnm": "vie",
        "auto": "auto",
        "ara": "ara", "bul": "bul", "chs": "chs", "cht": "cht", "hrv": "hrv",
        "cze": "cze", "dan": "dan", "dut": "dut", "fin": "fin", "fre": "fre",
        "ger": "ger", "gre": "gre", "hun": "hun", "kor": "kor", "ita": "ita",
        "jpn": "jpn", "pol": "pol", "por": "por", "rus": "rus", "slv": "slv",
        "spa": "spa", "swe": "swe", "tha": "tha", "tur": "tur", "ukr": "ukr"
    }
    ocr_lang = lang_map.get(lang.lower(), "auto")
    api_url = OCRSPACE_API_URL
    api_key = OCRSPACE_API_KEY  # Free API key

    if not api_url:
        raise ValueError("OCRSPACE_API_URL is not set. Update your config.json or settings.")
    if not api_key:
        raise ValueError("OCRSPACE_API_KEY is not set. Update your config.json before using ocrspace.")

    if verbose:
        print(f"[OCRSpace] Reading PDF: {pdf_path}")
        print(f"[OCRSpace] Using OCR language: {ocr_lang}")
        print(f"[OCRSpace] API URL: {api_url}")

    if verbose:
        print("[OCRSpace] Converting PDF to images for preprocessing...")
    images = convert_from_path(pdf_path, dpi=300)
    if not images:
        if verbose:
            print("[OCRSpace] No pages detected in PDF.")
        else:
            print("No pages detected in PDF.")
        return None
    processed_images = [_preprocess_ocr_image(img, verbose=verbose) for img in images]

    # Split into chunks: each chunk has at most 3 pages and <= 1024 KB
    chunk_paths = []
    max_chunk_size = 1024 * 1024  # 1024 KB
    max_chunk_pages = 3
    with tempfile.TemporaryDirectory() as tmpdir:
        total_pages = len(processed_images)
        i = 0
        while i < total_pages:
            n_pages = min(max_chunk_pages, total_pages - i)
            chunk_images = processed_images[i:i + n_pages]
            temp_chunk_path = os.path.join(tmpdir, f"chunk_{i+1}_{i+n_pages}.pdf")
            out_path = _write_pdf_chunk(chunk_images, temp_chunk_path, max_chunk_size, verbose=verbose)
            if out_path:
                if verbose:
                    size = os.path.getsize(out_path) if os.path.exists(out_path) else 0
                    print(f"[OCRSpace] Created chunk: {out_path} ({n_pages} page(s), {size} bytes)")
                chunk_paths.append(out_path)
            i += n_pages

        extracted_text = []
        for idx, chunk_path in enumerate(chunk_paths, 1):
            if verbose:
                print(f"[OCRSpace] Processing chunk {idx}/{len(chunk_paths)}: {chunk_path}")
            else:
                print(f"Processing chunk {idx}/{len(chunk_paths)}")
            with open(chunk_path, "rb") as pdf_file:
                pdf_bytes = pdf_file.read()
                pdf_base64 = base64.b64encode(pdf_bytes).decode("utf-8")

            data = {
                "apikey": api_key,
                "isOverlayRequired": True,
                "base64Image": "data:application/pdf;base64," + pdf_base64,
                "filetype": "pdf",
                "OCREngine": 2,
                "isTable": True,
                "scale": True,
                "detectOrientation": True,  # Enable auto-rotation detection
                "language": ocr_lang,
            }

            if verbose:
                print(f"[OCRSpace] Uploading chunk to OCR.space API...")
            # OCR.space expects a form-encoded payload and returns JSON per chunk.
            response = requests.post(api_url, data=data)
            if not response.ok:
                if verbose:
                    print(f"[OCRSpace] HTTP {response.status_code}: {response.text[:500]}")
                else:
                    print(f"OCR.space request failed with status {response.status_code}.")
                continue
            try:
                result = response.json()
            except Exception as e:
                if verbose:
                    print(f"[OCRSpace] Failed to parse OCR.space response: {e}")
                    print(f"[OCRSpace] Raw response: {response.text[:500]}")
                else:
                    print("Failed to parse OCR.space response.")
                continue

            # Check if result is a dictionary before using .get()
            if not isinstance(result, dict):
                if verbose:
                    print(f"[OCRSpace] Unexpected response format from OCR.space: {result}")
                else:
                    print("Unexpected response format from OCR.space.")
                continue

            if result.get("IsErroredOnProcessing"):
                if verbose:
                    print(f"[OCRSpace] Error: {result.get('ErrorMessage')}")
                else:
                    print(f"Error: {result.get('ErrorMessage')}")
                continue

            parsed_results = result.get("ParsedResults", [])
            if simple_text:
                for page_idx, page in enumerate(parsed_results):
                    overlay = page.get("TextOverlay")
                    if overlay and overlay.get("HasOverlay"):
                        lines = overlay.get("Lines", [])
                        lines_sorted = sorted(lines, key=lambda l: l.get("MinTop", 0))
                        page_text = []
                        for line in lines_sorted:
                            words = line.get("Words", [])
                            words_sorted = sorted(words, key=lambda w: w.get("Left", 0))
                            line_text = " ".join(word.get("WordText", "") for word in words_sorted)
                            page_text.append(line_text)
                        combined = "\n".join(page_text)
                        if combined.strip():
                            extracted_text.append(combined)
                        elif page.get("ParsedText"):
                            extracted_text.append(page.get("ParsedText", "").strip())
                    elif page.get("ParsedText"):
                        extracted_text.append(page.get("ParsedText", "").strip())
            else:
                for page_num, page in enumerate(parsed_results, 1):
                    page_header = f"=== Page {((idx-1)*3) + page_num} ==="
                    overlay = page.get("TextOverlay")
                    if overlay and overlay.get("HasOverlay"):
                        lines = overlay.get("Lines", [])
                        word_objs = []
                        for line in lines:
                            words = line.get("Words", [])
                            for word in words:
                                word_objs.append({
                                    "text": word.get("WordText", ""),
                                    "left": word.get("Left", 0),
                                    "top": word.get("Top", 0),
                                    "width": word.get("Width", 0),
                                    "height": word.get("Height", 0),
                                })
                        if not word_objs:
                            if page.get("ParsedText"):
                                extracted_text.append(f"{page_header}\n{page.get('ParsedText', '').strip()}")
                            continue
                        word_objs.sort(key=lambda w: (w["top"], w["left"]))
                        lines_by_top = []
                        line_thresh = 10  # pixels
                        for word in word_objs:
                            placed = False
                            for line in lines_by_top:
                                if abs(word["top"] - line["top"]) <= line_thresh:
                                    line["words"].append(word)
                                    line["tops"].append(word["top"])
                                    placed = True
                                    break
                            if not placed:
                                lines_by_top.append({"top": word["top"], "words": [word], "tops": [word["top"]]})
                        page_text_lines = []
                        for line in sorted(lines_by_top, key=lambda l: min(l["tops"])):
                            words = sorted(line["words"], key=lambda w: w["left"])
                            min_word_width = min([w["width"] for w in words if w["width"] > 0] or [10])
                            line_str = ""
                            prev_right = None
                            for w in words:
                                if prev_right is not None:
                                    gap = w["left"] - prev_right
                                    n_spaces = max(1, min(8, int(round(gap / (min_word_width / 2))))) if gap > min_word_width / 2 else 1
                                    line_str += " " * n_spaces
                                line_str += w["text"]
                                prev_right = w["left"] + w["width"]
                            page_text_lines.append(line_str.rstrip())
                        combined = "\n".join(page_text_lines)
                        if combined.strip():
                            extracted_text.append(f"{page_header}\n{combined}")
                        elif page.get("ParsedText"):
                            extracted_text.append(f"{page_header}\n{page.get('ParsedText', '').strip()}")
                    elif page.get("ParsedText"):
                        extracted_text.append(f"{page_header}\n{page.get('ParsedText', '').strip()}")

        full_text = "\n\n".join(extracted_text)

        if not txt_output_path:
            base = os.path.splitext(pdf_path)[0]
            txt_output_path = base + "_text_ocrspace.txt"
        with open(txt_output_path, "w", encoding="utf-8") as f:
            f.write(full_text)
        if verbose:
            print(f"[OCRSpace] Saved extracted text to {txt_output_path}")
        else:
            print(f"Saved to {txt_output_path}")

    # Return the text file path
    return txt_output_path

def extract_text_from_scanned_pdf_with_tesseract(pdf_path, txt_output_path=None, lang="vie", simple_text=False, verbose=False):
    """
    Extract text from a scanned PDF file using Tesseract OCR.
    Converts PDF pages to images and applies OCR to extract text.

    Args:
        pdf_path: Path to the PDF file
        txt_output_path: Path to save the extracted text (if None, will use pdf_path with suffix)
        lang: Language code for Tesseract (e.g., "vie" for Vietnamese, "eng" for English)
        simple_text: If True, output raw text without page headers
        verbose: If True, print more details; otherwise, print only important notice

    Returns:
        txt_output_path: Path to the text file containing the extracted text
    """
    if verbose:
        print(f"[TesseractOCR] Extracting text from PDF: {pdf_path}")
    else:
        print("Extracting text from PDF using Tesseract OCR...")

    if not txt_output_path:
        base = os.path.splitext(pdf_path)[0]
        txt_output_path = base + "_text_tesseract.txt"

    try:
        # Convert PDF to images
        if verbose:
            print("[TesseractOCR] Converting PDF to images...")
        images = convert_from_path(pdf_path, dpi=300)

        extracted_text = []

        # Process each page
        for page_num, image in enumerate(tqdm(images, desc="Processing pages"), 1):
            if verbose:
                print(f"[TesseractOCR] Processing page {page_num}/{len(images)}")
            # Apply OCR to the image
            try:
                custom_config = r'--oem 3 --psm 6'
                processed_image = _preprocess_ocr_image(image, verbose=verbose)
                text = pytesseract.image_to_string(processed_image, lang=lang, config=custom_config)
                if text.strip():
                    if simple_text:
                        extracted_text.append(text.strip())
                    else:
                        extracted_text.append(f"=== Page {page_num} ===\n{text.strip()}")
            except Exception as e:
                if verbose:
                    print(f"[TesseractOCR] Error processing page {page_num}: {e}")
                else:
                    print(f"Notice: Error processing page {page_num}.")
                continue

        full_text = "\n\n".join(extracted_text)

        with open(txt_output_path, "w", encoding="utf-8") as f:
            f.write(full_text)

        if verbose:
            print(f"[TesseractOCR] Extracted text saved to {txt_output_path}")
        else:
            print(f"Extracted text saved to {txt_output_path}")
        return txt_output_path

    except Exception as e:
        if verbose:
            print(f"[TesseractOCR] Error extracting text from PDF: {e}")
        else:
            print(f"Error extracting text from PDF: {e}")
        return None

def extract_text_from_scanned_pdf_with_paddleocr(pdf_path, txt_output_path=None, simple_text=False, verbose=False):
    """
    Extract text from a scanned PDF file using PaddleOCR.
    Converts PDF pages to images and applies PaddleOCR to extract text with better structure preservation.

    Args:
        pdf_path: Path to the PDF file
        txt_output_path: Path to save the extracted text (if None, will use pdf_path with suffix)
        simple_text: If True, print raw result without any layout (no page headers)
        verbose: If True, print more details; otherwise, print only important notice

    Returns:
        txt_output_path: Path to the text file containing the extracted text
    """
    if verbose:
        print(f"[PaddleOCR] Extracting text from PDF: {pdf_path}")
    else:
        print("Extracting text from PDF using PaddleOCR...")

    if not txt_output_path:
        base = os.path.splitext(pdf_path)[0]
        txt_output_path = base + "_text_paddleocr.txt"

    try:
        # Initialize PaddleOCR
        if verbose:
            print("[PaddleOCR] Initializing PaddleOCR...")
        ocr = PaddleOCR(
            use_doc_orientation_classify=False, 
            use_doc_unwarping=False, 
            use_textline_orientation=False
        )

        # Convert PDF to images
        if verbose:
            print("[PaddleOCR] Converting PDF to images...")
        images = convert_from_path(pdf_path, dpi=300)

        extracted_text = []

        # Process each page
        for page_num, image in enumerate(tqdm(images, desc="Processing pages"), 1):
            if verbose:
                print(f"[PaddleOCR] Processing page {page_num}/{len(images)}")
            processed_image = _preprocess_ocr_image(image, verbose=verbose)
            image_np = np.array(processed_image)
            try:
                result = ocr.predict(image_np)
                if result is None or len(result) == 0:
                    if verbose:
                        print(f"[PaddleOCR] No text detected on page {page_num}")
                    continue

                page_text_lines = []

                for res in result:
                    with tempfile.NamedTemporaryFile(mode="w+", suffix=".json", delete=False, encoding="utf-8") as tmp_json:
                        res.save_to_json(tmp_json.name)
                        tmp_json_path = tmp_json.name
                    with open(tmp_json_path, "r", encoding="utf-8") as f:
                        text_output = f.read()
                    if text_output:
                        page_text_lines.append(text_output)

                def try_parse_json(s):
                    try:
                        return json.loads(s)
                    except Exception:
                        return None

                page_json = None
                for line in page_text_lines:
                    obj = try_parse_json(line)
                    if obj and isinstance(obj, dict) and "rec_texts" in obj:
                        page_json = obj
                        break

                if simple_text:
                    texts = []
                    if page_json and "rec_texts" in page_json:
                        texts = [t for t in page_json["rec_texts"] if t.strip()]
                    else:
                        texts = []
                        for line in page_text_lines:
                            obj = try_parse_json(line)
                            if obj and "rec_texts" in obj:
                                texts.extend([t for t in obj["rec_texts"] if t.strip()])
                            else:
                                texts.append(line)
                    page_text = "\n".join(texts)
                    if page_text.strip():
                        extracted_text.append(page_text)
                else:
                    if page_json and "rec_texts" in page_json and "rec_boxes" in page_json:
                        rec_texts = page_json["rec_texts"]
                        rec_boxes = page_json["rec_boxes"]
                        box_info = []
                        for i, (text, box) in enumerate(zip(rec_texts, rec_boxes)):
                            if not text.strip():
                                continue
                            x1, y1, x2, y2 = box
                            box_info.append({
                                "text": text,
                                "x1": x1,
                                "y1": y1,
                                "x2": x2,
                                "y2": y2,
                                "idx": i,
                            })
                        box_info.sort(key=lambda b: (b["y1"], b["x1"]))
                        lines = []
                        line = []
                        last_y = None
                        y_thresh = 20
                        for b in box_info:
                            if last_y is None or abs(b["y1"] - last_y) > y_thresh:
                                if line:
                                    lines.append(line)
                                line = [b]
                                last_y = b["y1"]
                            else:
                                line.append(b)
                        if line:
                            lines.append(line)
                        page_lines = []
                        for line in lines:
                            line_sorted = sorted(line, key=lambda b: b["x1"])
                            line_text = "  ".join(b["text"] for b in line_sorted)
                            page_lines.append(line_text)
                        page_text = "\n".join(page_lines)
                        extracted_text.append(f"=== Page {page_num} ===\n{page_text}")
                    else:
                        if page_text_lines:
                            page_text = '\n'.join(page_text_lines)
                            extracted_text.append(f"=== Page {page_num} ===\n{page_text}")

            except Exception as e:
                if verbose:
                    print(f"[PaddleOCR] Error processing page {page_num}: {e}")
                else:
                    print(f"Notice: Error processing page {page_num}.")
                continue

        full_text = "\n\n".join(extracted_text)

        with open(txt_output_path, "w", encoding="utf-8") as f:
            f.write(full_text)

        if verbose:
            print(f"[PaddleOCR] Extracted text saved to {txt_output_path}")
        else:
            print(f"Extracted text saved to {txt_output_path}")
        return txt_output_path

    except Exception as e:
        if verbose:
            print(f"[PaddleOCR] Error extracting text from PDF: {e}")
        else:
            print(f"Error extracting text from PDF with PaddleOCR: {e}")
        return None

def extract_text_from_scanned_pdf(
    pdf_path,
    txt_output_path=None,
    service=DEFAULT_OCR_METHOD,
    lang="auto",
    simple_text=False,
    verbose=False
):
    """
    Extract handwriting text from a scanned PDF file using the specified OCR service.
    Supported services: see ALL_OCR_METHODS.
    If simple_text=True, print raw result without any layout.
    Post-OCR refinement is not supported.
    Returns the path to the output text file (or (txt_path, page_rotations) tuple if available).
    If verbose is True, print more details; otherwise, print only important notice.
    """
    # Dispatch OCR based on service.
    if service is None:
        service = DEFAULT_OCR_METHOD
    service = service.lower()
    if service not in ALL_OCR_METHODS:
        if verbose:
            print(f"[extract_text_from_scanned_pdf] OCR service '{service}' is not supported. Supported: {', '.join(ALL_OCR_METHODS)}.")
        else:
            print(f"OCR service '{service}' is not supported.")
        raise ValueError(f"OCR service '{service}' is not supported. Supported: {', '.join(ALL_OCR_METHODS)}.")
    if verbose:
        print(f"[extract_text_from_scanned_pdf] Using OCR service: {service}")
    else:
        print(f"Using OCR service: {service}")
    if service == "ocrspace":
        return extract_text_from_scanned_pdf_ocrspace(
            pdf_path,
            txt_output_path=txt_output_path,
            lang=lang,
            simple_text=simple_text,
            verbose=verbose
        )
    elif service == "tesseract":
        return extract_text_from_scanned_pdf_with_tesseract(
            pdf_path,
            txt_output_path=txt_output_path,
            lang="vie" if lang == "auto" else lang,
            simple_text=simple_text,
            verbose=verbose
        )
    elif service == "paddleocr":
        return extract_text_from_scanned_pdf_with_paddleocr(
            pdf_path,
            txt_output_path=txt_output_path,
            simple_text=simple_text,
            verbose=verbose
        )

def count_id_between_dates_in_textfile(file_path, student_id=None, verbose=False):
    """
    Detects student ids (8 digits, possibly with OCR errors) and dates (d/m/yyyy) from a text file.
    Reports how many times each id appears between two consecutive dates.
    The number of appearances between two different dates is regarded as the appearances in the first date.
    Each date and id should be in a single line, but the file may contain other data lines as well.
    Handles common OCR errors for digits (e.g., 1 <-> |, 0 <-> O, etc.).
    If student_id is specified, only count for that id.
    If not, detect all possible ids and return a dict: {id: {date: count, ...}, ...}
    Handles the case where one date appears many times: combine all data between same dates.
    Accepts dates with optional spaces and trailing text (e.g., "4 / 7 / 2025", "4 / 7 / 2025 ( Tiếp )").
    Removes extra texts like ( Tiếp ) from date.
    If verbose is True, print more details; otherwise, print only important notice.
    """
    ocr_digit_map = {
        '0': '0', 'O': '0', 'o': '0', 'Q': '0',
        '1': '1', 'I': '1', 'l': '1', '|': '1', '!': '1',
        '2': '2', 'Z': '2',
        '3': '3',
        '4': '4', 'A': '4',
        '5': '5', 'S': '5', 's': '5',
        '6': '6', 'G': '6',
        '7': '7', 'T': '7',
        '8': '8', 'B': '8',
        '9': '9', 'g': '9', 'q': '9',
    }

    def normalize_id(s):
        s = ''.join(c for c in s if c.isalnum() or c in '|!')
        mapped = []
        for c in s:
            mapped.append(ocr_digit_map.get(c, c))
        norm = ''.join(mapped)
        return re.findall(r'(\d{8})', norm)

    date_pattern = re.compile(
        r'\b(\d{1,2})\s*/\s*(\d{1,2})\s*/\s*(\d{4})\b'
    )

    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    date_lines = []
    id_lines_map = {}

    for idx, line in enumerate(lines):
        line = line.strip()
        date_match = date_pattern.search(line)
        if date_match:
            date_str = f"{int(date_match.group(1))}/{int(date_match.group(2))}/{date_match.group(3)}"
            date_lines.append((idx, date_str))
        ids = normalize_id(line)
        for found_id in ids:
            if found_id not in id_lines_map:
                id_lines_map[found_id] = []
            id_lines_map[found_id].append(idx)

    if student_id:
        norm_ids = normalize_id(student_id)
        if not norm_ids:
            if verbose:
                print(f"[count_id_between_dates_in_textfile] Provided student_id '{student_id}' is not a valid 8-digit id.")
            else:
                print("Provided student_id is not a valid 8-digit id.")
            raise ValueError("Provided student_id is not a valid 8-digit id.")
        norm_student_id = norm_ids[0]
        id_lines_map = {norm_student_id: id_lines_map.get(norm_student_id, [])}

    intervals = []
    for i in range(len(date_lines) - 1):
        start_idx = date_lines[i][0]
        end_idx = date_lines[i + 1][0]
        date1 = date_lines[i][1]
        date2 = date_lines[i + 1][1]
        intervals.append((start_idx, end_idx, date1, date2))
    if len(date_lines) >= 1:
        start_idx = date_lines[-1][0]
        end_idx = len(lines)
        date1 = date_lines[-1][1]
        date2 = None
        intervals.append((start_idx, end_idx, date1, date2))

    combined_intervals = {}
    for start_idx, end_idx, date1, date2 in intervals:
        if date1 not in combined_intervals:
            combined_intervals[date1] = []
        combined_intervals[date1].append((start_idx, end_idx))

    results_dict = {}
    for sid, id_lines in id_lines_map.items():
        date_counts = {}
        for date1, idx_ranges in combined_intervals.items():
            count = 0
            for start_idx, end_idx in idx_ranges:
                in_this_range = [id_idx for id_idx in id_lines if start_idx < id_idx < end_idx]
                count += len(in_this_range)
            date_counts[date1] = count
        results_dict[sid] = date_counts

    for sid, date_counts in results_dict.items():
        if verbose:
            print(f"[count_id_between_dates_in_textfile] Student id {sid}:")
            for date1, c in date_counts.items():
                print(f"  {date1}: {c}")
        else:
            print(f"Student id {sid}:")
            for date1, c in date_counts.items():
                print(f"  {date1}: {c}")

    return results_dict

def count_id_between_dates_in_pdf(pdf_path, student_id=None, lang="auto", verbose=False):
    """
    Extracts text from a scanned PDF using OCR, then counts student ids (8 digits, possibly with OCR errors)
    between consecutive dates (d/m/yyyy) in the extracted text.
    Combines extract_text_from_scanned_pdf and count_id_between_dates_in_textfile.
    Returns the same result as count_id_between_dates_in_textfile.
    If verbose is True, print more details; otherwise, print only important notice.
    """
    if verbose:
        print(f"[count_id_between_dates_in_pdf] Extracting text from PDF: {pdf_path}")
    txt_output_path = extract_text_from_scanned_pdf_ocrspace(pdf_path, lang=lang, verbose=verbose)
    if verbose:
        print(f"[count_id_between_dates_in_pdf] Extracted text saved to: {txt_output_path}")
        print(f"[count_id_between_dates_in_pdf] Counting IDs between dates in extracted text...")
    else:
        print("Extracted text and counting IDs between dates...")
    return count_id_between_dates_in_textfile(txt_output_path, student_id=student_id, verbose=verbose)

def add_blackboard_counts_from_pdf(pdf_path, db_path=None, lang="vnm", service=DEFAULT_OCR_METHOD, verbose=False):
    """
    Extracts blackboard counts from the PDF, calculates total and max total counts,
    updates the database with "Total Blackboard Counts", "Max Total Blackboard Counts",
    "Total Blackboard Counts until midterm", "Max total blackboard counts until midterm",
    and also adds per-date fields (e.g., "Blackboard Count: 4/7/2025").
    Also adds "Midterm Reward Points" and "Final Reward Points" for each student.
    If a student has no blackboard counts on a date, sets the value to 0.

    Enhancement: The student name is extracted as the closest name above the student id,
    but the name must not contain any numbers.
    If a student id does not exist in the database, but a student has a matching id and the name is similar,
    update the count on that student entry.
    Also supports merging counts from an extra TXT file named "blackboard_counts_extra.txt" in the same folder as the PDF.
    The extra file uses the same format as the extracted text: each line is a date or a student id, ids between two dates are counted to the first one.
    If verbose is True, print more details; otherwise, print only important notice.

    New: Adds a field "Blackboard Count Hash" to the database to store hashes of blackboard counts files.
    If the hash of the input file is unchanged, prompt user to proceed or quit (default: quit after 60s).
    """

    def file_hash(path):
        h = hashlib.sha256()
        with open(path, "rb") as f:
            while True:
                chunk = f.read(8192)
                if not chunk:
                    break
                h.update(chunk)
        return h.hexdigest()

    def get_input_with_timeout(prompt, timeout=60, default="quit"):
        def timeout_handler(signum, frame):
            print(f"\nTimeout: No response after {timeout} seconds. Using default: {default}")
            raise TimeoutError("User input timeout")
        if hasattr(signal, "SIGALRM"):
            signal.signal(signal.SIGALRM, timeout_handler)
            signal.alarm(timeout)
            try:
                val = input(prompt)
                signal.alarm(0)
                if not val:
                    return default
                return val
            except TimeoutError:
                signal.alarm(0)
                return default
            except KeyboardInterrupt:
                signal.alarm(0)
                print("\nOperation cancelled by user.")
                return "quit"
        else:
            # Windows fallback: no timeout
            try:
                val = input(prompt)
                if not val:
                    return default
                return val
            except KeyboardInterrupt:
                print("\nOperation cancelled by user.")
                return "quit"

    def normalize_name(name):
        if not name:
            return ""
        name = str(name)
        name = re.sub(r"[^a-zA-Z0-9 ]", "", name)
        name = re.sub(r"\s+", " ", name)
        return name.strip().lower()

    def is_name_similar(name1, name2, threshold=0.7):
        n1 = normalize_name(name1)
        n2 = normalize_name(name2)
        if not n1 or not n2:
            return False
        ratio = difflib.SequenceMatcher(None, n1, n2).ratio()
        return ratio >= threshold

    # --- Hash checking logic ---
    pdf_hash = file_hash(pdf_path)
    extra_path = os.path.join(os.path.dirname(pdf_path), "blackboard_counts_extra.txt")
    extra_hash = file_hash(extra_path) if os.path.exists(extra_path) else None
    hash_key = os.path.basename(pdf_path)
    extra_hash_key = os.path.basename(extra_path) if extra_hash else None

    if db_path:
        students = load_database(db_path)
    else:
        students = []

    # Find or create the Blackboard Count Hash field (dict: {filename: hash})
    hash_dict = {}
    for s in students:
        if hasattr(s, "Blackboard Count Hash"):
            try:
                hash_dict = s.__dict__["Blackboard Count Hash"]
                if not isinstance(hash_dict, dict):
                    hash_dict = {}
            except Exception:
                hash_dict = {}
            break
    # If not found, create it on the first student (or create a dummy student if db is empty)
    if not hash_dict:
        hash_dict = {}
        if students:
            students[0].__dict__["Blackboard Count Hash"] = hash_dict
        else:
            dummy = Student(**{"Name": "dummy", "Student ID": "00000000", "Blackboard Count Hash": hash_dict})
            students.append(dummy)

    # Check if hash matches previous run
    prev_pdf_hash = hash_dict.get(hash_key)
    prev_extra_hash = hash_dict.get(extra_hash_key) if extra_hash_key else None
    hash_changed = (pdf_hash != prev_pdf_hash) or (extra_hash and extra_hash != prev_extra_hash)
    if not hash_changed:
        msg = f"File '{os.path.basename(pdf_path)}' has not changed since last run."
        if extra_hash:
            msg += f" (and extra file '{os.path.basename(extra_path)}' has not changed)"
        print(msg)
        ans = get_input_with_timeout("Do you want to proceed anyway? (y/n, or 'q' to quit, default: quit): ", timeout=60, default="quit").strip().lower()
        if ans in ("q", "quit", "n", "no", ""):
            print("Operation cancelled. No changes made.")
            return students
        # else: proceed

    # Update hash_dict with current hashes
    hash_dict[hash_key] = pdf_hash
    if extra_hash_key:
        hash_dict[extra_hash_key] = extra_hash
    # Save hash_dict to all students (for redundancy)
    for s in students:
        s.__dict__["Blackboard Count Hash"] = hash_dict

    if verbose:
        print(f"[add_blackboard_counts_from_pdf] Extracting text from PDF: {pdf_path} using service: {service}")
    txt_output_path = extract_text_from_scanned_pdf(pdf_path, service=service, simple_text=True, verbose=verbose)
    if verbose:
        print(f"[add_blackboard_counts_from_pdf] Extracted text saved to: {txt_output_path}")
        print("[add_blackboard_counts_from_pdf] Parsing text to extract (date, [(id, name_above_id), ...]) structure...")

    ocr_digit_map = {
        '0': '0', 'O': '0', 'o': '0', 'Q': '0',
        '1': '1', 'I': '1', 'l': '1', '|': '1', '!': '1',
        '2': '2', 'Z': '2',
        '3': '3',
        '4': '4', 'A': '4',
        '5': '5', 'S': '5', 's': '5', '[': '5',
        '6': '6', 'G': '6',
        '7': '7', 'T': '7',
        '8': '8', 'B': '8',
        '9': '9', 'g': '9', 'q': '9',
    }
    def normalize_id(s):
        s = ''.join(c for c in s if c.isalnum() or c in '|!')
        mapped = []
        for c in s:
            mapped.append(ocr_digit_map.get(c, c))
        norm = ''.join(mapped)
        return re.findall(r'(\d{8})', norm)

    date_pattern = re.compile(
        r'\b(\d{1,2})\s*/\s*(\d{1,2})\s*/\s*(\d{4})\b'
    )

    # --- Parse extracted text ---
    with open(txt_output_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    date_lines = []
    for idx, line in enumerate(lines):
        if date_pattern.search(line):
            date_match = date_pattern.search(line)
            date_str = f"{int(date_match.group(1))}/{int(date_match.group(2))}/{date_match.group(3)}"
            date_lines.append((idx, date_str))

    date_id_name_map = {}
    for i in range(len(date_lines)):
        start_idx = date_lines[i][0]
        date_str = date_lines[i][1]
        end_idx = date_lines[i + 1][0] if i + 1 < len(date_lines) else len(lines)
        id_name_pairs = []
        for idx in range(start_idx + 1, end_idx):
            line = lines[idx].strip()
            ids = normalize_id(line)
            if ids:
                name = ""
                for j in range(idx - 1, start_idx, -1):
                    prev_line = lines[j].strip()
                    if prev_line and not normalize_id(prev_line) and not date_pattern.search(prev_line):
                        if not re.search(r'\d', prev_line):
                            name = prev_line
                            break
                for found_id in ids:
                    id_name_pairs.append((found_id, name))
        date_id_name_map.setdefault(date_str, []).extend(id_name_pairs)

    counts_dict = {}
    id_date_names = {}
    for date, id_name_pairs in date_id_name_map.items():
        for sid, name in id_name_pairs:
            if sid not in counts_dict:
                counts_dict[sid] = {}
                id_date_names[sid] = {}
            counts_dict[sid][date] = counts_dict[sid].get(date, 0) + 1
            id_date_names[sid][date] = name

    # --- Enhancement: Merge counts from blackboard_counts_extra.txt ---
    if os.path.exists(extra_path):
        if verbose:
            print(f"[add_blackboard_counts_from_pdf] Merging extra counts from: {extra_path}")
        with open(extra_path, "r", encoding="utf-8") as f:
            extra_lines = f.readlines()
        # Parse extra file: each line is a date or a student id, ids between two dates are counted to the first one
        extra_date_lines = []
        for idx, line in enumerate(extra_lines):
            if date_pattern.search(line):
                date_match = date_pattern.search(line)
                date_str = f"{int(date_match.group(1))}/{int(date_match.group(2))}/{date_match.group(3)}"
                extra_date_lines.append((idx, date_str))
        extra_date_id_name_map = {}
        for i in range(len(extra_date_lines)):
            start_idx = extra_date_lines[i][0]
            date_str = extra_date_lines[i][1]
            end_idx = extra_date_lines[i + 1][0] if i + 1 < len(extra_date_lines) else len(extra_lines)
            id_name_pairs = []
            for idx in range(start_idx + 1, end_idx):
                line = extra_lines[idx].strip()
                ids = normalize_id(line)
                if ids:
                    name = ""
                    for j in range(idx - 1, start_idx, -1):
                        prev_line = extra_lines[j].strip()
                        if prev_line and not normalize_id(prev_line) and not date_pattern.search(prev_line):
                            if not re.search(r'\d', prev_line):
                                name = prev_line
                                break
                    for found_id in ids:
                        id_name_pairs.append((found_id, name))
            extra_date_id_name_map.setdefault(date_str, []).extend(id_name_pairs)
        # Merge extra counts into main counts_dict
        for date, id_name_pairs in extra_date_id_name_map.items():
            for sid, name in id_name_pairs:
                if sid not in counts_dict:
                    counts_dict[sid] = {}
                    id_date_names[sid] = {}
                counts_dict[sid][date] = counts_dict[sid].get(date, 0) + 1
                if name:
                    id_date_names[sid][date] = name
        if verbose:
            print("[add_blackboard_counts_from_pdf] Extra counts merged.")

    all_dates = set()
    for date_counts in counts_dict.values():
        all_dates.update(date_counts.keys())
    all_dates = sorted(all_dates)

    def _parse_date(date_str):
        if not date_str:
            return None
        parts = [p for p in date_str.split("/") if p.strip()]
        if len(parts) != 3:
            return None
        try:
            day, month, year = (int(p) for p in parts)
        except ValueError:
            return None
        return [day, month, year]

    def date_leq(date1, date2):
        d1 = _parse_date(date1)
        d2 = _parse_date(date2)
        if not d1 or not d2:
            return False
        return d1[2] < d2[2] or (d1[2] == d2[2] and (d1[1] < d2[1] or (d1[1] == d2[1] and d1[0] <= d2[0])))

    sid_to_total = {sid: sum(date_counts.values()) for sid, date_counts in counts_dict.items()}
    if _parse_date(MIDTERM_DATE):
        sid_to_total_until_midterm = {
            sid: sum(v for k, v in date_counts.items() if date_leq(k, MIDTERM_DATE))
            for sid, date_counts in counts_dict.items()
        }
    else:
        sid_to_total_until_midterm = dict(sid_to_total)
    max_total = max(sid_to_total.values()) if sid_to_total else 0
    max_total_until_midterm = max(sid_to_total_until_midterm.values()) if sid_to_total_until_midterm else 0

    sid_to_student = {}
    name_to_students = {}
    for s in students:
        sid = getattr(s, "Student ID", None)
        name = getattr(s, "Name", None)
        if sid:
            sid_to_student[str(sid).strip()] = s
        if name:
            norm_name = normalize_name(name)
            if norm_name:
                name_to_students.setdefault(norm_name, []).append(s)

    db_id_to_extracted_ids = {}
    for sid in counts_dict.keys():
        student = sid_to_student.get(sid)
        if student:
            db_id_to_extracted_ids.setdefault(sid, []).append(sid)
            continue
        best_id_match = None
        min_mismatches = float('inf')
        for db_sid in sid_to_student.keys():
            if len(sid) == len(db_sid):
                mismatches = sum(1 for a, b in zip(sid, db_sid) if a != b)
                if mismatches < min_mismatches and mismatches <= 2:
                    min_mismatches = mismatches
                    best_id_match = db_sid
        if best_id_match:
            db_id_to_extracted_ids.setdefault(best_id_match, []).append(sid)
            if verbose:
                print(f"[add_blackboard_counts_from_pdf] Matched extracted ID {sid} to database ID {best_id_match} (mismatches: {min_mismatches})")
        else:
            best_match = None
            best_score = 0
            extracted_names = id_date_names.get(sid, {})
            for date, ext_name in extracted_names.items():
                norm_ext_name = normalize_name(ext_name)
                if not norm_ext_name:
                    continue
                for db_norm_name, candidates in name_to_students.items():
                    score = difflib.SequenceMatcher(None, norm_ext_name, db_norm_name).ratio()
                    if score > best_score and score >= 0.7:
                        best_score = score
                        best_match = candidates[0]
                        break
                if best_match:
                    break
            if best_match:
                matched_sid = getattr(best_match, "Student ID", None)
                if matched_sid:
                    db_id_to_extracted_ids.setdefault(matched_sid, []).append(sid)
                    if verbose:
                        print(f"[add_blackboard_counts_from_pdf] Matched extracted ID {sid} to student with name {getattr(best_match, 'Name', '')} (ID: {matched_sid}, name similarity: {best_score:.2f})")
            else:
                db_id_to_extracted_ids.setdefault(sid, []).append(sid)
                if verbose:
                    print(f"[add_blackboard_counts_from_pdf] Extracted ID {sid} could not be matched to any student in the database. Will create new record.")

    for db_sid, extracted_ids in db_id_to_extracted_ids.items():
        student = sid_to_student.get(db_sid)
        if not student:
            best_name = ""
            for ext_sid in extracted_ids:
                extracted_names = id_date_names.get(ext_sid, {})
                for date, name in extracted_names.items():
                    if name and len(name) > len(best_name):
                        best_name = name
            student = Student(**{"Student ID": db_sid, "Name": best_name})
            students.append(student)
            sid_to_student[db_sid] = student
            if best_name:
                name_to_students.setdefault(normalize_name(best_name), []).append(student)
            if verbose:
                print(f"[add_blackboard_counts_from_pdf] Created new student record: ID {db_sid}, Name: {best_name}")
        merged_date_counts = {}
        for ext_sid in extracted_ids:
            date_counts = counts_dict.get(ext_sid, {})
            for date, count in date_counts.items():
                merged_date_counts[date] = merged_date_counts.get(date, 0) + count
        if verbose and len(extracted_ids) > 1:
            print(f"[add_blackboard_counts_from_pdf] Merged counts for student {getattr(student, 'Name', 'Unknown')} (ID: {db_sid}):")
            print(f"  Extracted IDs: {extracted_ids}")
            print(f"  Merged counts: {merged_date_counts}")
        total = sum(merged_date_counts.values())
        total_until_midterm = sum(v for k, v in merged_date_counts.items() if date_leq(k, MIDTERM_DATE))
        setattr(student, "Total Blackboard Counts", total)
        setattr(student, "Total Blackboard Counts until midterm", total_until_midterm)
        for date in all_dates:
            field_name = f"Blackboard Count: {date}"
            count = merged_date_counts.get(date, 0)
            setattr(student, field_name, count)

    for s in students:
        sid = getattr(s, "Student ID", None)
        if sid is None or str(sid).strip() not in counts_dict:
            setattr(s, "Total Blackboard Counts", 0)
            setattr(s, "Total Blackboard Counts until midterm", 0)
            for date in all_dates:
                field_name = f"Blackboard Count: {date}"
                setattr(s, field_name, 0)

    for s in students:
        setattr(s, "Max Total Blackboard Counts", max_total)
        setattr(s, "Max Total Blackboard Counts until midterm", max_total_until_midterm)

    M = max_total
    for s in students:
        total = getattr(s, "Total Blackboard Counts", 0)
        total_until_midterm = getattr(s, "Total Blackboard Counts until midterm", 0)
        if total_until_midterm >= M / 2 and total_until_midterm > 0:
            midterm_reward = 2
        elif total_until_midterm > 0:
            midterm_reward = 1
        else:
            midterm_reward = 0
        setattr(s, "Midterm Reward Points", midterm_reward)
        if total >= M / 2 and total > 0:
            final_reward = 1
        elif total >= 3:
            final_reward = 0.5
        else:
            final_reward = 0
        setattr(s, "Final Reward Points", final_reward)

    # Save hash_dict to all students again (in case new students were added)
    for s in students:
        s.__dict__["Blackboard Count Hash"] = hash_dict

    if db_path:
        save_database(students, db_path)
    if verbose:
        print("[add_blackboard_counts_from_pdf] Total, max, per-date, midterm blackboard counts, and reward points added to database.")
        print(f"[add_blackboard_counts_from_pdf] Processed {len(students)} students. Dates found: {all_dates}")
        print(f"[add_blackboard_counts_from_pdf] Max total blackboard counts: {max_total}, Max until midterm: {max_total_until_midterm}")
    else:
        print("Blackboard counts and reward points updated in database.")
    return students

def print_all_blackboard_counts_by_date(students, db_path=None, verbose=False):
    """
    Print all dates with blackboard counts in a table-like format (Vietnamese if possible).
    Each date is a table: columns are "STT", "Họ và Tên", "Mã sinh viên", "Số lần lên bảng".
    Students are assumed to be already sorted by Vietnamese alphabetical order of full name, then student id.
    Handles Vietnamese full names.
    If verbose is True, print more details; otherwise, print only important notice.
    """
    output_str = get_blackboard_counts_by_date_text(students, db_path=db_path, verbose=verbose)
    print(output_str)

def export_all_blackboard_counts_by_date_to_txt(students, file_path=None, db_path=None, verbose=False):
    """
    Export all dates with blackboard counts to a TXT file (Vietnamese if possible).
    Each date is a table: columns are "STT", "Họ và Tên", "Mã sinh viên", "Số lần lên bảng".
    Students are assumed to be already sorted by Vietnamese alphabetical order of full name, then student id.
    Handles Vietnamese full names.
    If verbose is True, print more details; otherwise, print only important notice.
    """
    if not file_path:  # Handles both None and empty string
        file_path = os.path.join(os.getcwd(), "blackboard_counts_by_date.txt")
    output_str = get_blackboard_counts_by_date_text(students, db_path=db_path, verbose=verbose)
    try:
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(output_str)
        if verbose:
            print(f"[ExportBlackboardCounts] Saved results to {file_path}")
        else:
            print(f"Saved results to {file_path}")
    except Exception as e:
        if verbose:
            print(f"[ExportBlackboardCounts] Could not save file: {e}")
        else:
            print(f"Could not save file: {e}")

def get_blackboard_counts_by_date_text(students, db_path=None, verbose=False):
    """
    Helper function to generate the blackboard counts by date as a string.
    Includes the midterm and final reward points for each student.
    If verbose is True, print more details; otherwise, print only important notice.
    """
    if db_path:
        students = load_database(db_path)
    if not students:
        msg = "No students found in the database."
        if verbose:
            print("[BlackboardCounts] " + msg)
        return msg

    # Find all dates with blackboard counts
    all_dates = set()
    for s in students:
        for attr in s.__dict__:
            if attr.startswith("Blackboard Count: "):
                date = attr[len("Blackboard Count: "):]
                all_dates.add(date)
    
    # Sort dates chronologically (d/m/yyyy format)
    def parse_date(date_str):
        # Handles both d/m/yyyy and dd/mm/yyyy
        parts = date_str.strip().split("/")
        if len(parts) == 3:
            day, month, year = [int(p) for p in parts]
            return (year, month, day)
        return (9999, 99, 99)  # fallback for malformed dates
    sorted_dates = sorted(all_dates, key=parse_date)

    # Sort students by name for consistent output
    sorted_students = sort_students_by_firstname(students)
    
    # Create output
    output_lines = []
    
    # First add max total blackboard counts information
    max_total = 0
    max_students = []
    for s in students:
        total = getattr(s, "Total Blackboard Counts", 0)
        if total > max_total:
            max_total = total
            max_students = [s]
        elif total == max_total and max_total > 0:
            max_students.append(s)

    current_date = datetime.now().strftime("%d/%m/%Y")

    output_lines.append(f"SỐ LẦN LÊN BẢNG TỐI ĐA (Tính đến: {current_date}): {max_total}")
    if verbose:
        print(f"[BlackboardCounts] Max total blackboard counts as of {current_date}: {max_total}")

    # Then add reward points summary
    output_lines.append(f"\nĐIỂM CỘNG CỦA SINH VIÊN (Tính đến: {current_date}):")
    
    header = f"{'STT':<4}    {'Họ và Tên':<30}    {'Mã sinh viên':<12}    {'Điểm cộng giữa kỳ':<18}    {'Điểm cộng cuối kỳ':<18}"
    output_lines.append(header)
    output_lines.append("-" * len(header))
    
    for idx, s in enumerate(sorted_students, 1):
        name = getattr(s, "Name", "")
        sid = getattr(s, "Student ID", "")
        midterm_reward = getattr(s, "Midterm Reward Points", 0)
        final_reward = getattr(s, "Final Reward Points", 0)
        output_lines.append(f"{idx:<4}    {name:<30}    {sid:<12}    {midterm_reward:<18}    {final_reward:<18}")
    if verbose:
        print(f"[BlackboardCounts] Reward points summary for {len(sorted_students)} students.")

    # Then add blackboard counts by date
    output_lines.append("\n\nSỐ LẦN LÊN BẢNG THEO NGÀY:")
    
    # For each date, list all students and their blackboard counts
    for date in sorted_dates:
        output_lines.append(f"\nNgày: {date}")
        date_header = f"{'STT':<4}    {'Họ và Tên':<30}    {'Mã sinh viên':<12}    {'Số lần lên bảng':<16}"
        output_lines.append(date_header)
        output_lines.append("-" * len(date_header))
        
        for idx, s in enumerate(sorted_students, 1):
            name = getattr(s, "Name", "")
            sid = getattr(s, "Student ID", "")
            count = getattr(s, f"Blackboard Count: {date}", 0)
            if count > 0:  # Only show students who have counts for this date
                output_lines.append(f"{idx:<4}    {name:<30}    {sid:<12}    {count:<16}")
        if verbose:
            print(f"[BlackboardCounts] Listed blackboard counts for date {date}.")

    return "\n".join(output_lines)

def export_all_blackboard_counts_by_date_to_markdown(students, file_path=None, db_path=None, verbose=False):
    """
    Export all dates with blackboard counts to a Markdown file.
    The content and structure closely matches the TXT version: includes max total, reward points summary, and per-date tables.
    If verbose is True, print more details; otherwise, print only important notice.
    """
    if db_path:
        students = load_database(db_path)
    if not students:
        if verbose:
            print("[ExportMarkdown] No students to export.")
        else:
            print("No students to export.")
        return
    if not file_path:
        file_path = os.path.join(os.getcwd(), "blackboard_counts_by_date.md")

    # Find all dates with blackboard counts
    all_dates = set()
    for s in students:
        for attr in s.__dict__:
            if attr.startswith("Blackboard Count: "):
                date = attr[len("Blackboard Count: "):]
                all_dates.add(date)
    sorted_dates = sorted(all_dates, key=lambda x: [int(part) for part in x.split("/")])

    sorted_students = sort_students_by_firstname(students)

    # Max total blackboard counts
    max_total = 0
    for s in students:
        total = getattr(s, "Total Blackboard Counts", 0)
        if total > max_total:
            max_total = total

    current_date = datetime.now().strftime("%d/%m/%Y")

    lines = []
    lines.append(f"# SỐ LẦN LÊN BẢNG TỐI ĐA (Tính đến: {current_date}): {max_total}\n")

    # Reward points summary
    lines.append(f"## ĐIỂM CỘNG CỦA SINH VIÊN (Tính đến: {current_date})\n")
    lines.append("| STT | Họ và Tên | Mã sinh viên | Điểm cộng giữa kỳ | Điểm cộng cuối kỳ |")
    lines.append("| --- | --------- | ------------ | ----------------- | ----------------- |")
    for idx, s in enumerate(sorted_students, 1):
        name = getattr(s, "Name", "")
        sid = getattr(s, "Student ID", "")
        midterm_reward = getattr(s, "Midterm Reward Points", 0)
        final_reward = getattr(s, "Final Reward Points", 0)
        lines.append(f"| {idx} | {name} | {sid} | {midterm_reward} | {final_reward} |")
    lines.append("")

    # Blackboard counts by date
    lines.append("## SỐ LẦN LÊN BẢNG THEO NGÀY\n")
    for date in sorted_dates:
        lines.append(f"### Ngày: {date}\n")
        lines.append("| STT | Họ và Tên | Mã sinh viên | Số lần lên bảng |")
        lines.append("| --- | --------- | ------------ | --------------- |")
        for idx, s in enumerate(sorted_students, 1):
            name = getattr(s, "Name", "")
            sid = getattr(s, "Student ID", "")
            count = getattr(s, f"Blackboard Count: {date}", 0)
            if count > 0:
                lines.append(f"| {idx} | {name} | {sid} | {count} |")
        lines.append("")  # Blank line between tables

    try:
        with open(file_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        if verbose:
            print(f"[ExportMarkdown] Saved blackboard counts by date to {file_path}")
        else:
            print(f"Saved blackboard counts by date to {file_path}")
    except Exception as e:
        if verbose:
            print(f"[ExportMarkdown] Could not save markdown file: {e}")
        else:
            print(f"Could not save markdown file: {e}")

def analyze_text_meaningfulness(text, refine_method=DEFAULT_AI_METHOD, average_length=None, verbose=False):
    """
    Analyze the meaningfulness of extracted text using AI with improved heuristics.
    Handles both regular Vietnamese text and mathematical documents with special notation.

    Args:
        text (str): The extracted text to analyze.
        refine_method (str): AI service to use ("gemini", "huggingface", or "local"). Default is "gemini".
        average_length (float): Average length of all submissions for comparison. If None, length factor is ignored.
        verbose (bool): If True, print more details; otherwise, print only important notice.

    Returns:
        float: Meaningfulness score between 0 and 1.
    """
    if not text or not text.strip():
        if verbose:
            print("[Meaningfulness] Empty text, score = 0.0")
        else:
            print("Notice: Empty text, score = 0.0")
        return 0.0

    # Pre-analysis using heuristics to short-circuit obvious cases.
    heuristic_score = _analyze_text_heuristics(text)
    metrics = _compute_text_quality_metrics(text)
    is_math_document = metrics["likely_math"]

    # Length-based adjustment if average_length is provided (compares to cohort).
    if average_length and average_length > 0:
        text_length = len(text.strip())
        length_ratio = text_length / average_length

        # Apply penalty for submissions that are significantly shorter than average
        if length_ratio < QUALITY_LENGTH_RATIO_LOW:  # Less than 30% of average length
            length_penalty = 0.5  # Reduce score by 50%
        elif length_ratio < QUALITY_LENGTH_RATIO_MEDIUM:  # Less than 50% of average length
            length_penalty = 0.7  # Reduce score by 30%
        elif length_ratio < QUALITY_LENGTH_RATIO_HIGH:  # Less than 70% of average length
            length_penalty = 0.85  # Reduce score by 15%
        else:
            length_penalty = 1.0  # No penalty

        heuristic_score *= length_penalty
        if verbose:
            print(f"[Meaningfulness] Length ratio: {length_ratio:.2f}, penalty: {length_penalty:.2f}, heuristic_score after penalty: {heuristic_score:.2f}")

    # Adjust heuristic score if math content is detected (OCR noise is higher).
    if is_math_document:
        heuristic_score = min(1.0, heuristic_score * 1.2)
        if verbose:
            math_density = metrics.get("math_density", 0.0)
            print(f"[Meaningfulness] Math document detected (density: {math_density:.2f}), heuristic_score boosted to {heuristic_score:.2f}")

    # If heuristics show very low or very high confidence, use that
    if heuristic_score <= 0.15 or heuristic_score >= 0.85:
        if verbose:
            print(f"[Meaningfulness] Heuristic score is extreme ({heuristic_score:.2f}), returning directly.")
        else:
            print(f"Notice: Heuristic score is extreme ({heuristic_score:.2f}), returning directly.")
        return heuristic_score

    # Truncate text if too long, but preserve structure
    if len(text) > 5000:
        text = text[:2500] + "\n...[text truncated]...\n" + text[-2500:]
        if verbose:
            print("[Meaningfulness] Text truncated for AI analysis.")

    # Enhanced prompt with more specific criteria
    length_context = f" (Length: {len(text.strip())} chars"
    if average_length:
        try:
            length_context += f", Average: {int(average_length)} chars)"
        except Exception:
            length_context += ")"
    else:
        length_context += ")"
    math_context = ", likely contains mathematical content" if is_math_document else ""
    issues = _summarize_quality_issues(metrics, average_length=average_length)
    issues_text = "; ".join(issues) if issues else "None"
    if verbose:
        print(f"[Meaningfulness] Detected issues: {issues_text}")

    # Escape curly braces in text to avoid f-string format errors
    safe_text = text.replace("{", "{{").replace("}", "}}")
    prompt = (
        "You are an expert text analysis assistant specializing in Vietnamese OCR text evaluation, with particular expertise in mathematical and scientific content. Please analyze the following text extracted from a student's PDF submission and rate its meaningfulness on a scale of 0 to 1.\n\n"
        "Rating criteria:\n"
        "- 0.0-0.1: Completely meaningless (random characters, no Vietnamese words or valid math notation, pure OCR garbage)\n"
        "- 0.1-0.3: Mostly meaningless (few Vietnamese words or math symbols scattered, no coherent sentences, 80%+ OCR errors)\n"
        "- 0.3-0.5: Partially meaningful (some Vietnamese sentences or mathematical expressions identifiable, moderate structure, 50-80% OCR errors)\n"
        "- 0.5-0.7: Moderately meaningful (clear Vietnamese sentences or mathematical content, recognizable structure, 20-50% OCR errors)\n"
        "- 0.7-0.9: Mostly meaningful (well-structured Vietnamese or mathematical content, minor OCR issues, <20% errors)\n"
        "- 0.9-1.0: Highly meaningful (clear, well-structured Vietnamese academic or mathematical content, minimal OCR errors)\n\n"
        "Evaluation factors:\n"
        "1. Vietnamese language coherence (proper grammar, sentence structure) OR mathematical coherence for math documents\n"
        "2. Presence of academic/educational vocabulary or mathematical notation appropriate for student work\n"
        "3. Logical flow and organization of ideas or mathematical reasoning\n"
        "4. OCR error density (character substitutions, missing diacritics, garbled text)\n"
        "5. Ratio of meaningful content to noise/errors\n"
        "6. Context appropriateness for academic or mathematical submission\n"
        "7. Content length and substance relative to expectations\n\n"
        "IMPORTANT: For mathematical content, consider that the following are likely meaningful, NOT errors:\n"
        "- Greek letters (α, β, γ, etc.)\n"
        "- Mathematical symbols (∫, Σ, ∂, ∇, etc.)\n"
        "- Logical operators (∧, ∨, ¬, →, ↔, ⊕, ⊢, ⊨)\n"
        "- Logical equivalence symbols (≡, ≅, ≈, ≃, ⟺, ⇔)\n"
        "- Logical truth values (T, F, True, False)\n"
        "- Set operators (∪, ∩, ⊂, ⊆, ⊃, ⊇, \\, ∈, ∉, ∅)\n"
        "- Set notation and descriptions using curly braces\n"
        "- LaTeX-style notation (may appear as \\alpha, \\beta, \\int, \\sum, etc.)\n"
        "- Equations with symbols like =, ≠, ≈, ≤, ≥\n"
        "- Subscripts and superscripts\n"
        "- Fractions and mathematical expressions\n"
        "- Various brackets, parentheses and special characters used in mathematics\n\n"
        "Vietnamese-specific considerations:\n"
        "- Diacritics (á, à, ả, ã, ạ, ă, ắ, ằ, ẳ, ẵ, ặ, â, ấ, ầ, ẩ, ẫ, ậ, etc.) may be missing or incorrect\n"
        "- Common OCR substitutions: đ→d, ô→o, ư→u, etc.\n"
        "- Tone marks may appear as separate characters\n"
        "- Handwriting OCR often produces character fragments\n\n"
        f"Detected potential issues: {issues_text}.\n"
        f"Based on preliminary analysis, this text appears to have moderate quality (heuristic score: {heuristic_score:.2f}){length_context}{math_context}. Please provide your expert assessment.\n\n"
        "Respond with ONLY a single number between 0 and 1 (format: 0.XX). No explanations needed.\n\n"
        "Text to analyze:\n"
        f"{safe_text}"
    )

    try:
        if refine_method not in ALL_AI_METHODS:
            if verbose:
                print("[Meaningfulness] Unknown AI method, defaulting to gemini.")
            response = refine_text_with_gemini("", user_prompt=prompt)
        elif refine_method == "gemini":
            if verbose:
                print("[Meaningfulness] Calling Gemini for AI analysis...")
            response = refine_text_with_gemini("", user_prompt=prompt)
        elif refine_method == "huggingface":
            if verbose:
                print("[Meaningfulness] Calling HuggingFace for AI analysis...")
            response = refine_text_with_huggingface("", user_prompt=prompt)
        elif refine_method == "local":
            if verbose:
                print("[Meaningfulness] Calling local model for AI analysis...")
            response = refine_text_with_local_llm("", user_prompt=prompt)

        # Extract score from response
        score_match = re.search(r'(\d+\.?\d*)', response.strip())
        if score_match:
            ai_score = float(score_match.group(1))
            ai_score = min(1.0, max(0.0, ai_score))  # Clamp between 0 and 1

            # For math documents, adjust weights to rely more on AI's interpretation
            if is_math_document:
                weight_ai = 0.8
                weight_heuristic = 0.2
            else:
                weight_ai = 0.7
                weight_heuristic = 0.3

            # Combine heuristic and AI scores with weighted average
            final_score = (ai_score * weight_ai) + (heuristic_score * weight_heuristic)
            if verbose:
                print(f"[Meaningfulness] AI score: {ai_score:.2f}, heuristic: {heuristic_score:.2f}, final: {final_score:.2f}")
            else:
                print(f"Meaningfulness score: {final_score:.2f}")
            return min(1.0, max(0.0, final_score))
        else:
            if verbose:
                print(f"[Meaningfulness] Could not parse meaningfulness score from AI response: {response}")
            else:
                print("Notice: Could not parse meaningfulness score from AI response.")
            return heuristic_score  # Fallback to heuristic score

    except Exception as e:
        if verbose:
            print(f"[Meaningfulness] Error analyzing text meaningfulness: {e}")
        else:
            print(f"Error analyzing text meaningfulness: {e}")
        return heuristic_score  # Fallback to heuristic score

def _compute_text_quality_metrics(text):
    text = text.strip()
    total_chars = len(text)
    words = re.findall(r"\w+", text, re.UNICODE)
    total_words = len(words)
    unique_chars = len(set(text)) if total_chars else 0
    unique_char_ratio = unique_chars / total_chars if total_chars else 0.0

    alnum_count = sum(1 for c in text if c.isalnum())
    alnum_ratio = alnum_count / total_chars if total_chars else 0.0
    symbol_count = sum(1 for c in text if not c.isalnum() and not c.isspace())
    symbol_ratio = symbol_count / total_chars if total_chars else 0.0

    repeat_char_runs = len(re.findall(r"(.)\1{3,}", text))
    repeat_char_ratio = repeat_char_runs / max(1, total_chars // 10)

    lines = text.splitlines()
    empty_lines = sum(1 for line in lines if not line.strip())
    line_empty_ratio = empty_lines / max(1, len(lines))

    # Use Unicode ranges to detect Vietnamese letters and avoid mojibake.
    vn_letter_re = re.compile(r"[\u00c0-\u1ef9\u0102\u0103\u00c2\u00e2\u00ca\u00ea\u00d4\u00f4\u01a0\u01a1\u01af\u01b0\u0110\u0111]")
    vn_char_count = len(vn_letter_re.findall(text))
    vn_char_ratio = vn_char_count / total_chars if total_chars else 0.0

    common_vn_words = [
        "và", "là", "của", "trong", "có", "được", "một", "cho", "với", "các",
        "này", "đó", "những", "từ", "theo", "hoặc", "khi", "nếu", "sẽ", "đã",
        "đang", "bài", "sinh", "viên", "giải", "đáp", "phương", "trình", "hàm",
        "số", "định", "nghĩa"
    ]
    text_lower = unicodedata.normalize("NFC", text).lower()
    common_word_count = sum(1 for word in common_vn_words if word in text_lower)

    math_indicators = [
        r"\\[a-zA-Z]+",
        r"\$[^$]*\$", r"\$\$[^$]*\$\$",
        r"[a-zA-Z]\s*[=<>]\s*[0-9a-zA-Z]",
        r"[a-zA-Z]\^[0-9]+", r"[a-zA-Z]_[0-9]+",
        r"(\d+/\d+)", r"(\d+\.\d+)"
    ]
    # Heuristic math signal to avoid penalizing symbol-heavy math submissions.
    math_matches = 0
    for pattern in math_indicators:
        try:
            math_matches += len(re.findall(pattern, text))
        except Exception:
            continue
    math_density = math_matches / max(1, total_words)
    likely_math = math_density > QUALITY_MATH_DENSITY_THRESHOLD

    return {
        "total_chars": total_chars,
        "total_words": total_words,
        "unique_char_ratio": unique_char_ratio,
        "alnum_ratio": alnum_ratio,
        "symbol_ratio": symbol_ratio,
        "repeat_char_ratio": repeat_char_ratio,
        "line_empty_ratio": line_empty_ratio,
        "vn_char_ratio": vn_char_ratio,
        "common_word_count": common_word_count,
        "likely_math": likely_math,
        "math_density": math_density,
    }


def _summarize_quality_issues(metrics, average_length=None):
    issues = []
    if metrics["total_chars"] < QUALITY_MIN_CHARS:
        issues.append("Nội dung quá ngắn")
    if (
        metrics["unique_char_ratio"] < QUALITY_UNIQUE_CHAR_RATIO_MIN
        or metrics["repeat_char_ratio"] > QUALITY_REPEAT_CHAR_RATIO_MAX
    ):
        issues.append("Ký tự lặp bất thường hoặc thiếu đa dạng")
    if metrics["vn_char_ratio"] < QUALITY_VN_CHAR_RATIO_MIN and not metrics["likely_math"]:
        issues.append("Tỷ lệ ký tự tiếng Việt rất thấp")
    if metrics["alnum_ratio"] < QUALITY_ALNUM_RATIO_MIN:
        issues.append("Tỷ lệ ký tự chữ/số thấp")
    if metrics["symbol_ratio"] > QUALITY_SYMBOL_RATIO_MAX and not metrics["likely_math"]:
        issues.append("Quá nhiều ký hiệu hoặc ký tự không chữ/số")
    if metrics["line_empty_ratio"] > QUALITY_EMPTY_LINE_RATIO_MAX:
        issues.append("Nhiều dòng trống")
    if average_length and average_length > 0:
        length_ratio = metrics["total_chars"] / average_length
        if length_ratio < QUALITY_LENGTH_RATIO_LOW:
            issues.append("Độ dài thấp hơn nhiều so với trung bình")
        elif length_ratio < QUALITY_LENGTH_RATIO_MEDIUM:
            issues.append("Độ dài thấp hơn trung bình")
    return issues


def _analyze_text_heuristics(text):
    """
    Analyze text meaningfulness using heuristic methods.

    Args:
        text (str): The text to analyze.

    Returns:
        float: Heuristic meaningfulness score between 0 and 1.
    """
    if not text or not text.strip():
        return 0.0

    metrics = _compute_text_quality_metrics(text)
    total_chars = metrics["total_chars"]
    total_words = metrics["total_words"]

    if total_chars < QUALITY_MIN_CHARS:
        return 0.2

    if (
        metrics["unique_char_ratio"] < QUALITY_UNIQUE_CHAR_RATIO_MIN
        or metrics["repeat_char_ratio"] > QUALITY_REPEAT_CHAR_RATIO_MAX
    ):
        return 0.1

    score = 0.3

    if metrics["vn_char_ratio"] > 0.05:
        score += 0.2
    if metrics["vn_char_ratio"] > 0.1:
        score += 0.2

    if metrics["common_word_count"] > 0:
        score += 0.1
    if metrics["common_word_count"] > 3:
        score += 0.1

    if total_words > 10:
        score += 0.1
    if total_words > 50:
        score += 0.05

    if metrics["alnum_ratio"] < QUALITY_ALNUM_RATIO_MIN:
        score *= 0.8
    if metrics["symbol_ratio"] > QUALITY_SYMBOL_RATIO_MAX and not metrics["likely_math"]:
        score *= 0.7
    if metrics["line_empty_ratio"] > QUALITY_EMPTY_LINE_RATIO_MAX:
        score *= 0.85

    return min(1.0, max(0.0, score))


def generate_low_quality_message(filename, score, text, refine_method=DEFAULT_AI_METHOD, verbose=False):
    """
    Generate a personalized message for students with low quality submissions.

    Args:
        filename (str): The PDF filename.
        score (float): The meaningfulness score.
        text (str): The extracted text (truncated for context).
        refine_method (str): AI service to use.
        verbose (bool): If True, print more details; otherwise, print only important notice.

    Returns:
        str: Generated message in Vietnamese.
    """
    # If refine_method is None, ask user which AI model to use, default to gemini after 60s
    if refine_method is None:
        def timeout_handler(signum, frame):
            print("\nNo response after 60 seconds. Using default: gemini")
            raise TimeoutError("User input timeout")
        # Only use SIGALRM if available (not on Windows)
        if hasattr(signal, "SIGALRM"):
            signal.signal(signal.SIGALRM, timeout_handler)
            signal.alarm(60)
            try:
                while True:
                    choice = input("Which AI model do you want to use to generate the message? (gemini/huggingface/local) [default: gemini]: ").strip().lower()
                    if not choice:
                        refine_method = "gemini"
                        break
                    if choice in ALL_AI_METHODS:
                        refine_method = choice
                        break
                    else:
                        if verbose:
                            print("[Meaningfulness] Please enter 'gemini', 'huggingface', or 'local'.")
                        else:
                            print("Please enter 'gemini', 'huggingface', or 'local'.")
            except TimeoutError:
                refine_method = "gemini"
            finally:
                signal.alarm(0)
        else:
            # Fallback for platforms without SIGALRM (e.g., Windows)
            try:
                while True:
                    choice = input("Which AI model do you want to use to generate the message? (gemini/huggingface/local) [default: gemini]: ").strip().lower()
                    if not choice:
                        refine_method = "gemini"
                        break
                    if choice in ALL_AI_METHODS:
                        refine_method = choice
                        break
                    else:
                        if verbose:
                            print("[Meaningfulness] Please enter 'gemini', 'huggingface', or 'local'.")
                        else:
                            print("Please enter 'gemini', 'huggingface', or 'local'.")
            except KeyboardInterrupt:
                refine_method = "gemini"

    # Truncate text for context
    text_sample = text[:1000] + "..." if len(text) > 1000 else text
    metrics = _compute_text_quality_metrics(text)
    issues = _summarize_quality_issues(metrics, average_length=None)
    issues_text = "; ".join(issues) if issues else "Không phát hiện rõ ràng"

    # The prompt includes explicit criteria and detected issues to keep the message consistent.
    prompt = f"""You are an expert educational assistant. Generate a polite, formal, professional message in Vietnamese to inform a student that their PDF submission has quality issues and needs to be reformatted and resubmitted. Do not add any extra message or sign message. Mention that the message is automatically generated and sent. Do not use "em" for calling a student in Vietnamese, instead use "bạn" for singular or "các bạn" for plural. 

The message should include:
1. A polite greeting
2. Explanation that the system detected low readability in their submission (score: {score:.2f}/1.0)
3. Mention the comprehensive evaluation criteria:
   - Vietnamese language coherence (proper grammar, sentence structure)
   - Presence of academic/educational vocabulary appropriate for student work
   - Logical flow and organization of ideas
   - OCR error density (character substitutions, missing diacritics, garbled text)
   - Ratio of meaningful Vietnamese content to noise/errors
   - Context appropriateness for academic submission
   - Content length and substance relative to expectations
   - Vietnamese-specific issues like missing diacritics and tone marks
   - Detected issues from system heuristics: {issues_text}
4. Mention that this might be due to:
   - Poor scan quality or resolution
   - Handwriting legibility issues
   - Formatting problems
   - Insufficient content length
   - Language coherence issues
5. Clear instructions to improve and resubmit:
   - Use better scan quality (higher resolution, good lighting)
   - Ensure clearer handwriting with proper Vietnamese diacritics
   - Consider typing the text if handwriting is unclear
   - Ensure adequate content length and academic substance
   - Check for proper Vietnamese grammar and sentence structure
6. A deadline reminder to resubmit soon
7. Offer to help if they have questions
8. Also mention that the detection can be wrong due to various factors such as differences in writing style, errors in extracting text from PDF via OCR, and so on.

Context:
- File: {filename}
- Quality score: {score:.2f}/1.0 (threshold for acceptable quality is 0.4)
- Detected issues: {issues_text}
- Sample extracted text: {text_sample[:500]}...

Please write a complete, professional message that is ready to send without any placeholders or additional editing needed. The message should be empathetic but clear about the requirement to resubmit with improved quality."""
    
    try:
        if refine_method not in ALL_AI_METHODS:
            if verbose:
                print("[Meaningfulness] Unknown AI method, defaulting to gemini.")
            message = refine_text_with_gemini("", user_prompt=prompt)
        elif refine_method == "gemini":
            if verbose:
                print("[Meaningfulness] Generating message using Gemini...")
            message = refine_text_with_gemini("", user_prompt=prompt)
        elif refine_method == "huggingface":
            if verbose:
                print("[Meaningfulness] Generating message using HuggingFace...")
            message = refine_text_with_huggingface("", user_prompt=prompt)
        elif refine_method == "local":
            if verbose:
                print("[Meaningfulness] Generating message using local model...")
            message = refine_text_with_local_llm("", user_prompt=prompt)
        
        # Add automatic generation note
        message += "\n\nThông báo này được tạo và gửi tự động bởi hệ thống AI sau khi phát hiện chất lượng bài nộp thấp."
        
        return message

    except Exception as e:
        if verbose:
            print(f"[Meaningfulness] Error generating message: {e}")
        else:
            print(f"Error generating message: {e}")
        # Fallback message with updated criteria
        fallback_message = f"""Chào bạn,

Hệ thống đã phát hiện rằng bài nộp của bạn (file: {filename}) có chất lượng đọc thấp (điểm: {score:.2f}/1.0, ngưỡng chấp nhận: 0.4/1.0).

Hệ thống đánh giá dựa trên các tiêu chí toàn diện sau:
- Tính mạch lạc của tiếng Việt (ngữ pháp, cấu trúc câu đúng)
- Sự hiện diện của từ vựng học thuật phù hợp
- Luồng logic và tổ chức ý tưởng
- Mật độ lỗi OCR (thay thế ký tự, thiếu dấu, văn bản rối)
- Tỷ lệ nội dung tiếng Việt có nghĩa so với nhiễu/lỗi
- Tính phù hợp ngữ cảnh cho bài nộp học thuật
- Độ dài nội dung và chất lượng so với yêu cầu
- Các vấn đề đặc thù tiếng Việt (thiếu dấu, thanh điệu)

Điều này có thể do:
- Chất lượng scan kém hoặc độ phân giải thấp
- Chữ viết tay khó đọc
- Vấn đề định dạng file
- Nội dung quá ngắn hoặc thiếu chất lượng học thuật
- Vấn đề về tính mạch lạc ngôn ngữ

Để cải thiện và nộp lại, vui lòng:
- Sử dụng chất lượng scan tốt hơn (độ phân giải cao, ánh sáng tốt)
- Đảm bảo chữ viết tay rõ ràng với dấu tiếng Việt chính xác
- Cân nhắc đánh máy nếu chữ viết tay không rõ
- Đảm bảo nội dung đủ dài và có chất lượng học thuật
- Kiểm tra ngữ pháp và cấu trúc câu tiếng Việt

Vui lòng nộp lại sớm nhất có thể. Nếu bạn có thắc mắc gì, vui lòng liên hệ với giảng viên.

Trân trọng,
Hệ thống tự động

Thông báo này được tạo và gửi tự động bởi hệ thống AI sau khi phát hiện chất lượng bài nộp thấp."""
        if issues_text:
            fallback_message += f"\n\nC\u00e1c d\u1ea5u hi\u1ec7u h\u1ec7 th\u1ed1ng ph\u00e1t hi\u1ec7n: {issues_text}"
        return fallback_message

def read_multichoice_exam_solutions_from_pdf(
    pdf_path,
    verbose=False
):
    """
    Read a multi-choice exam solution PDF containing multiple sheet codes, each page corresponds to one sheet code.
    For each sheet code (page), extract the corresponding correct solutions for exam questions.
    Saves each sheet's solutions to ./Exams/<EXAM_TYPE>/solutions/<EXAM_TYPE>_<sheet_code>_solutions.txt
    Returns a dict: {sheet_code: {question_number: correct_answer, ...}, ...}

    Args:
        pdf_path (str): Path to the PDF file.
        verbose (bool): Print more details.

    Returns:
        dict: {sheet_code: {question_number: correct_answer, ...}, ...}
    """
    solutions_by_sheet = {}

    # Prepare output folder
    out_dir = os.path.join("Exams", EXAM_TYPE, "solutions")
    os.makedirs(out_dir, exist_ok=True)

    with open(pdf_path, "rb") as f:
        reader = PyPDF2.PdfReader(f)
        num_pages = len(reader.pages)
        for page_num in range(num_pages):
            page = reader.pages[page_num]
            text = page.extract_text() or ""

            # Parse text to extract sheet code and solutions
            sheet_code_pattern = re.compile(r"(Đề\s*số)\s*[:\-]?\s*(\d+)", re.IGNORECASE)
            # Match multiple "number: answer" or "number.answer" or "number answer" in a line
            multi_q_pattern = re.compile(r"0?(\d{1,3})\s*[.:]?\s*([A-D])", re.IGNORECASE)

            current_sheet = None
            solutions = {}
            lines = text.splitlines()
            for line in lines:
                line = line.strip()
                m_sheet = sheet_code_pattern.search(line)
                if m_sheet:
                    current_sheet = m_sheet.group(2)
                    continue
                # Find all question/answer pairs in the line
                for m_q in multi_q_pattern.finditer(line):
                    q_num = int(m_q.group(1))
                    ans = m_q.group(2).upper()
                    solutions[q_num] = ans

            # If no sheet code found, use page number as sheet code
            sheet_code = current_sheet if current_sheet else f"page_{page_num+1}"
            solutions_by_sheet[sheet_code] = solutions

            # Save to TXT file
            out_path = os.path.join(out_dir, f"{EXAM_TYPE}_{sheet_code}_solutions.txt")
            with open(out_path, "w", encoding="utf-8") as out_f:
                out_f.write(f"Sheet code: {sheet_code}\n")
                for q_num in sorted(solutions.keys()):
                    out_f.write(f"Question {q_num}: {solutions[q_num]}\n")

            if verbose:
                print(f"[ExamSolutions] Extracted solutions for sheet code {sheet_code}: {len(solutions)} questions. Saved to {out_path}")
            else:
                print(f"Extracted solutions for sheet code {sheet_code}: {len(solutions)} questions. Saved to {out_path}")

    if verbose:
        print(f"[ExamSolutions] Extracted solutions for {len(solutions_by_sheet)} sheet codes (pages).")
    else:
        print(f"Extracted solutions for {len(solutions_by_sheet)} sheet codes (pages).")

    return solutions_by_sheet

def read_multichoice_answers_from_scanned_pdf(
    pdf_path,
    ocr_service=DEFAULT_OCR_METHOD,
    lang="auto",
    verbose=False,
    db_path=None
):
    """
    Extract student answers from scanned multi-choice exam sheet PDF.
    Each page corresponds to a single sheet code.
    For each page:
      - Extract student id ("Mã sinh viên: <id>" or "Số báo danh: <id>") and sheet code ("Đề số: <sheet code>")
        using the specified OCR service (ocrspace, tesseract, paddleocr).
      - If id cannot be correctly extracted, try to find the closest match in the student database (if db_path is provided).
      - Convert page to PNG named <student_id>_<sheet_code>.png in a temp dir.
    Run omrchecker: omrchecker -a -i <tempdir> -o <tempdir>
    Read latest Results_*.csv in <tempdir>/Results, extract answers from q1, q2, ...
    Saves each student's answers to ./Exams/<EXAM_TYPE>/answers/<EXAM_TYPE>_<student_id>_<sheet_code>_answers.txt
    Returns: list of dicts: {sheet_code, student_id, student_name, answers: {q_num: answer, ...}}
    """

    # Load student database for id correction if provided
    student_db = {}
    if db_path and os.path.exists(db_path):
        try:
            students = load_database(db_path)
            for s in students:
                sid = str(getattr(s, "Student ID", "")).strip()
                name = getattr(s, "Name", "")
                if sid:
                    student_db[sid] = name
        except Exception:
            student_db = {}

    def find_closest_student_id(extracted_id):
        # Try to find the closest match in the student_db with at most one digit difference
        if not student_db or not extracted_id or not extracted_id.isdigit():
            return extracted_id
        for sid in student_db:
            if len(sid) == len(extracted_id):
                diff = sum(a != b for a, b in zip(sid, extracted_id))
                if diff <= 1:
                    return sid
        return extracted_id

    def normalize_student_id(raw_id):
        # Remove all non-digit characters except allow dots and colons between digits, then remove them
        # e.g. "2.3.1.2.0.0.1.2" or "2:3:1:2:0:0:1:2" -> "23120012"
        if not raw_id:
            return ""
        # Remove all except digits, dots, and colons
        s = re.sub(r"[^\d\.\:]", "", raw_id)
        # Remove dots and colons
        s = s.replace(".", "").replace(":", "")
        return s

    # Use ./tmp as temp directory
    tempdir = os.path.join(os.getcwd(), "tmp")
    if os.path.exists(tempdir):
        shutil.rmtree(tempdir, ignore_errors=True)
    os.makedirs(tempdir, exist_ok=True)

    # --- Copy config.json and template.json from ./omrchecker/<EXAM_TYPE>/ to tempdir ---
    omr_config_dir = os.path.join("omrchecker", EXAM_TYPE)
    for fname in ["config.json", "template.json"]:
        src_path = os.path.join(omr_config_dir, fname)
        dst_path = os.path.join(tempdir, fname)
        if os.path.exists(src_path):
            try:
                shutil.copy(src_path, dst_path)
                if verbose:
                    print(f"[OMR] Copied {fname} from {src_path} to {dst_path}")
            except Exception as e:
                if verbose:
                    print(f"[OMR] Failed to copy {fname}: {e}")
        else:
            if verbose:
                print(f"[OMR] {fname} not found in {omr_config_dir}")

    results = []
    try:
        images = convert_from_path(pdf_path, dpi=300)
        page_info = []
        for idx, image in enumerate(images):
            # Step 1: Extract text from image using the specified OCR service
            ocr_text = ""
            if ocr_service == "ocrspace":
                # Save image to temp file
                img_path = os.path.join(tempdir, f"page_{idx+1}.png")
                image.save(img_path, "PNG")
                
                # Reduce image size to less than 1MB, convert to base64, save as PNG
                img_path = os.path.join(tempdir, f"page_{idx+1}.png")
                # Save initial PNG
                image.save(img_path, "PNG")
                # Reduce size if needed
                max_size = 1024 * 1024  # 1MB
                quality = 90
                img_bytes = io.BytesIO()
                image.save(img_bytes, format="PNG", optimize=True)
                while img_bytes.tell() > max_size and quality > 10:
                    img_bytes = io.BytesIO()
                    # Convert to JPEG for better compression if PNG is too large
                    image.convert("RGB").save(img_bytes, format="JPEG", quality=quality, optimize=True)
                    quality -= 10
                # Save the final image (PNG or JPEG) to img_path
                with open(img_path, "wb") as f_img:
                    f_img.write(img_bytes.getvalue())
                # Convert to base64
                with open(img_path, "rb") as f_img:
                    img_data = f_img.read()
                png_base64 = base64.b64encode(img_data).decode("utf-8")
                
                # Use OCR.space API
                with open(img_path, "rb") as img_f:
                    api_url = OCRSPACE_API_URL
                    api_key = OCRSPACE_API_KEY
                    payload = {
                        "apikey": api_key,
                        "isOverlayRequired": True,
                        "base64Image": "data:image/png;base64," + png_base64,
                        "filetype": "png",
                        "OCREngine": 2,
                        "scale": True,
                        "detectOrientation": True,  # Enable auto-rotation detection
                        "language": "auto"
                    }
                    files = {'file': img_f}
                    try:
                        resp = requests.post(api_url, files=files, data=payload)
                        resp.raise_for_status()
                        result = resp.json()
                        ocr_text = result.get("ParsedResults", [{}])[0].get("ParsedText", "")
                        if verbose:
                            print(f"[OMR] OCR.space result for page {idx+1}: {ocr_text[:500]}...")  # Print first 500 chars
                        time.sleep(3)  # Respect OCR.space rate limit
                    except Exception as e:
                        if verbose:
                            print(f"[OMR] OCR.space error: {e}")
                        ocr_text = ""
                # Remove the extracted PNG file after OCR
                try:
                    if os.path.exists(img_path):
                        os.remove(img_path)
                except Exception as e:
                    if verbose:
                        print(f"[OMR] Warning: Could not remove temp PNG file {img_path}: {e}")
            elif ocr_service == "tesseract":
                ocr_text = pytesseract.image_to_string(image, lang=lang if lang != "auto" else "vie")
            elif ocr_service == "paddleocr":
                ocr = PaddleOCR(use_doc_orientation_classify=False, 
                                use_doc_unwarping=False, 
                                use_textline_orientation=False)
                img_np = np.array(image)
                result = ocr.predict(img_np)
                lines = []
                for page in result:
                    for line in page:
                        lines.append(line[1][0])
                ocr_text = "\n".join(lines)
            else:
                # fallback to tesseract
                ocr_text = pytesseract.image_to_string(image, lang=lang if lang != "auto" else "vie")

            # --- Robust extraction for possibly defected text (less spaces) ---
            # Optionally extract student name (robust: allow less spaces)
            name_match = (
                re.search(r"H[ọo][\s_]*v[àa][\s_]*tên[\s_:.\-]*([^\n]+)", ocr_text, re.IGNORECASE)
                or re.search(r"H[ọo]vàtên[\s_:.\-]*([^\n]+)", ocr_text, re.IGNORECASE)
            )
            student_name = name_match.group(1).strip() if name_match else ""
            # Try to find "Mã sinh viên" or "Số báo danh" even if spaces are missing or reduced
            sid_match = (
                re.search(r"(M[ãa][\s_]*sinh[\s_]*viên|S[ốo][\s_]*b[áa][\s_]*danh)[\s_:.\-]*([A-Za-z0-9\.\:\s]+)", ocr_text, re.IGNORECASE)
                or re.search(r"(M[ãa]sinhviên|S[ốo]bádanh)[\s_:.\-]*([A-Za-z0-9\.\:\s]+)", ocr_text, re.IGNORECASE)
                or re.search(r"(M[ãa][\s_]*SV)[\s_:.\-]*([A-Za-z0-9\.\:\s]+)", ocr_text, re.IGNORECASE)
            )
            extracted_id_raw = sid_match.group(2) if sid_match else f"unknown_{idx+1}"
            student_id = normalize_student_id(extracted_id_raw)

            # Ensure student_id has exactly 8 digits
            if not (student_id.isdigit() and len(student_id) == 8):
                # Try to correct id if not found in db or not 8 digits
                corrected_id = None
                if student_db:
                    # Try to find closest id in db with 8 digits
                    for sid in student_db:
                        if len(sid) == 8 and sid.isdigit() and len(student_id) == 8:
                            diff = sum(a != b for a, b in zip(sid, student_id))
                            if diff <= 1:
                                corrected_id = sid
                                break
                    # If not found, try to find closest name in db and extract id
                    if not corrected_id and student_name:
                        best_match = None
                        best_ratio = 0.0
                        for sid, name in student_db.items():
                            ratio = difflib.SequenceMatcher(None, student_name.lower(), name.lower()).ratio()
                            if ratio > best_ratio:
                                best_ratio = ratio
                                best_match = sid
                        if best_match and best_ratio > 0.7:
                            corrected_id = best_match
                            if verbose:
                                print(f"[OMR] Used name match for id: {student_name} -> {student_db[corrected_id]} ({corrected_id})")
                if corrected_id:
                    if verbose and corrected_id != student_id:
                        print(f"[OMR] Corrected student id: {student_id} -> {corrected_id}")
                    student_id = corrected_id
                else:
                    # fallback: mark as unknown if still not valid
                    student_id = f"unknown_{idx+1}"
            # At this point, student_id is either valid or "unknown_x"

            # Find sheet code "Đề số" even if spaces are missing or reduced
            sheet_match = (
                re.search(r"Đ[ềe][\s_]*s[ốo][\s_:.\-]*([A-Za-z0-9]+)", ocr_text, re.IGNORECASE)
                or re.search(r"Đ[ềe]s[ốo][\s_:.\-]*([A-Za-z0-9]+)", ocr_text, re.IGNORECASE)
            )
            sheet_code = sheet_match.group(1) if sheet_match else f"page_{idx+1}"

            # Save image as PNG
            png_filename = f"{student_id}_{sheet_code}.png"
            png_path = os.path.join(tempdir, png_filename)
            image.save(png_path, "PNG")
            page_info.append({
                "student_id": student_id,
                "sheet_code": sheet_code,
                "student_name": student_name,
                "png_path": png_path
            })
            if verbose:
                print(f"[OMR] Saved PNG: {png_path} (Student ID: {student_id}, Sheet code: {sheet_code})")

        # Step 2: Run omrchecker
        omr_cmd = ["omrchecker", "-a", "-i", tempdir, "-o", tempdir]
        if verbose:
            print(f"[OMR] Running: {' '.join(omr_cmd)}")
        subprocess.run(omr_cmd, check=True)

        # Step 3: Find latest Results_*.csv in Results subdir
        results_dir = os.path.join(tempdir, "Results")
        csv_files = sorted(
            glob.glob(os.path.join(results_dir, "Results_*.csv")),
            key=os.path.getmtime,
            reverse=True
        )
        if not csv_files:
            if verbose:
                print("[OMR] No Results_*.csv found in Results directory.")
            return []
        latest_csv = csv_files[0]
        if verbose:
            print(f"[OMR] Reading results from: {latest_csv}")

        # Prepare output folder for answers
        answers_dir = os.path.join("Exams", EXAM_TYPE, "answers")
        os.makedirs(answers_dir, exist_ok=True)

        # Step 4: Parse CSV and extract answers
        with open(latest_csv, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Get PNG filename using file_id column
                file_id = row.get("file_id", "") or row.get("File_ID", "") or row.get("FileId", "")
                base_img = f"{file_id}.png" if file_id else ""
                # Extract student_id and sheet_code from filename
                m = re.match(r"([A-Za-z0-9]+)_([A-Za-z0-9]+)\.png", base_img)
                student_id = m.group(1) if m else ""
                sheet_code = m.group(2) if m else ""
                # Find student_name from database using student_id
                student_name = ""
                if student_db and student_id in student_db:
                    student_name = student_db[student_id]
                else:
                    # Fallback: try to get from page_info if not found in db
                    for info in page_info:
                        if os.path.basename(info["png_path"]) == base_img:
                            student_name = info.get("student_name", "")
                            break
                # Extract answers
                answers = {}
                for k, v in row.items():
                    q_match = re.match(r"q(\d+)", k, re.IGNORECASE)
                    if q_match:
                        q_num = int(q_match.group(1))
                        answers[q_num] = v.strip()
                # Save answers to TXT file
                out_path = os.path.join(
                    answers_dir,
                    f"{EXAM_TYPE}_{student_id}_{sheet_code}_answers.txt"
                )
                with open(out_path, "w", encoding="utf-8") as out_f:
                    out_f.write(f"Sheet code: {sheet_code}\n")
                    out_f.write(f"Student ID: {student_id}\n")
                    out_f.write(f"Student name: {student_name}\n")
                    for q_num in sorted(answers.keys()):
                        out_f.write(f"Question {q_num}: {answers[q_num]}\n")
                if verbose:
                    print(f"[OMR] Saved answers to {out_path}")
                results.append({
                    "sheet_code": sheet_code,
                    "student_id": student_id,
                    "student_name": student_name,
                    "answers": answers
                })
        if verbose:
            print(f"[OMR] Extracted answers for {len(results)} students.")
        return results
    except Exception as e:
        if verbose:
            print(f"[OMR] Error: {e}")
        return []
    finally:
        shutil.rmtree(tempdir, ignore_errors=True)

def evaluate_multichoice_exam_answers(
    exam_type=EXAM_TYPE,
    db_path=None,
    verbose=False
):
    """
    Evaluate multiple-choice exam answers for all students by comparing their answers with the correct solutions.
    If solution TXT files do not exist in Exams/<exam_type>/solutions, extract them from PDF.
    If answer TXT files do not exist in Exams/<exam_type>/answers, extract them from scanned PDF.
    Adds EXAM_TYPE reward points from student database (if available) to the mark.
    If the total points is more than 10, the mark is capped at 10.
    Saves results for each student in ./Exams/<EXAM_TYPE>/evaluations/<EXAM_TYPE>_<student_id>_<sheet_code>_evaluations.txt
    Returns a list of dicts: {student_id, student_name, sheet_code, total_questions, correct_count, answers, solutions, details, mark, reward_points}
    """
    solutions_dir = os.path.join("Exams", exam_type, "solutions")
    answers_dir = os.path.join("Exams", exam_type, "answers")
    evaluations_dir = os.path.join("Exams", exam_type, "evaluations")
    os.makedirs(evaluations_dir, exist_ok=True)

    # Determine point per question based on exam_type
    if str(exam_type).lower() == "midterm":
        point_per_question = 0.25
        reward_field = "Midterm Reward Points"
        mark_field = "Midterm Mark"
    elif str(exam_type).lower() == "final":
        point_per_question = 0.2
        reward_field = "Final Reward Points"
        mark_field = "Final Mark"
    else:
        point_per_question = 0.25  # Default
        reward_field = f"{exam_type.capitalize()} Reward Points"
        mark_field = f"{exam_type.capitalize()} Mark"

    # Helper to read solutions from TXT files
    def read_solutions_from_txt():
        solutions = {}
        for fname in glob.glob(os.path.join(solutions_dir, f"{exam_type}_*_solutions.txt")):
            sheet_code = None
            sheet_solutions = {}
            with open(fname, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line.lower().startswith("sheet code:"):
                        sheet_code = line.split(":", 1)[1].strip()
                    elif line.lower().startswith("question"):
                        m = re.match(r"question\s*(\d+):\s*([A-D])", line, re.IGNORECASE)
                        if m:
                            q_num = int(m.group(1))
                            ans = m.group(2).upper()
                            sheet_solutions[q_num] = ans
            if sheet_code:
                solutions[sheet_code] = sheet_solutions
        return solutions

    # Helper to read answers from TXT files
    def read_answers_from_txt():
        answers = []
        for fname in glob.glob(os.path.join(answers_dir, f"{exam_type}_*_answers.txt")):
            entry = {"answers": {}}
            with open(fname, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line.lower().startswith("sheet code:"):
                        entry["sheet_code"] = line.split(":", 1)[1].strip()
                    elif line.lower().startswith("student id:"):
                        entry["student_id"] = line.split(":", 1)[1].strip()
                    elif line.lower().startswith("student name:"):
                        entry["student_name"] = line.split(":", 1)[1].strip()
                    elif line.lower().startswith("question"):
                        m = re.match(r"question\s*(\d+):\s*([A-D])", line, re.IGNORECASE)
                        if m:
                            q_num = int(m.group(1))
                            ans = m.group(2).upper()
                            entry["answers"][q_num] = ans
            if "student_id" in entry and "sheet_code" in entry:
                answers.append(entry)
        return answers

    # If solution TXT files do not exist, try to extract from PDF
    if not glob.glob(os.path.join(solutions_dir, f"{exam_type}_*_solutions.txt")):
        pdf_candidates = glob.glob(os.path.join("Exams", exam_type, "*.pdf"))
        if pdf_candidates:
            pdf_path = pdf_candidates[0]
            if verbose:
                print(f"[EvalMC] Extracting solutions from PDF: {pdf_path}")
            read_multichoice_exam_solutions_from_pdf(pdf_path, verbose=verbose)
        else:
            if verbose:
                print(f"[EvalMC] No solution PDF found in Exams/{exam_type}/")
    # If answer TXT files do not exist, try to extract from scanned PDF
    if not glob.glob(os.path.join(answers_dir, f"{exam_type}_*_answers.txt")):
        pdf_candidates = glob.glob(os.path.join("Exams", exam_type, "*.pdf"))
        if pdf_candidates:
            pdf_path = pdf_candidates[0]
            if verbose:
                print(f"[EvalMC] Extracting answers from scanned PDF: {pdf_path}")
            read_multichoice_answers_from_scanned_pdf(pdf_path, verbose=verbose)
        else:
            if verbose:
                print(f"[EvalMC] No answer PDF found in Exams/{exam_type}/")

    # Read solutions and answers
    solutions = read_solutions_from_txt()
    answers = read_answers_from_txt()

    # Load student database if available
    student_db = {}
    students = []
    if db_path and os.path.exists(db_path):
        try:
            students = load_database(db_path)
            for s in students:
                sid = str(getattr(s, "Student ID", "")).strip()
                if sid:
                    student_db[sid] = s
        except Exception:
            student_db = {}

    results = []
    for entry in answers:
        sheet_code = entry.get("sheet_code")
        student_id = entry.get("student_id")
        # Try to get student_name from database if possible
        student_name = ""
        if student_id and student_db.get(student_id):
            s = student_db[student_id]
            student_name = getattr(s, "Name", "")
        else:
            student_name = entry.get("student_name", "")
        student_answers = entry.get("answers", {})
        solution = solutions.get(sheet_code, {})
        total_questions = len(solution)
        correct_count = 0
        details = []
        for q_num in sorted(solution.keys()):
            correct_ans = solution[q_num]
            student_ans = student_answers.get(q_num, "")
            is_correct = (student_ans == correct_ans)
            if is_correct:
                correct_count += 1
            details.append({
                "question": q_num,
                "student_answer": student_ans,
                "correct_answer": correct_ans,
                "is_correct": is_correct
            })
        mark = correct_count * point_per_question

        # Add reward points from database if available
        reward_points = 0
        if student_id and student_db.get(student_id):
            s = student_db[student_id]
            reward_points = float(getattr(s, reward_field, 0) or 0)
        total_mark = mark + reward_points
        if total_mark > 10:
            total_mark = 10

        result_entry = {
            "student_id": student_id,
            "student_name": student_name,
            "sheet_code": sheet_code,
            "total_questions": total_questions,
            "correct_count": correct_count,
            "answers": student_answers,
            "solutions": solution,
            "details": details,
            "mark": total_mark,
            "reward_points": reward_points
        }
        results.append(result_entry)

        # Save evaluation to TXT file
        eval_filename = f"{exam_type}_{student_id}_{sheet_code}_evaluations.txt"
        eval_path = os.path.join(evaluations_dir, eval_filename)
        try:
            with open(eval_path, "w", encoding="utf-8") as f:
                f.write(f"Student ID: {student_id}\n")
                f.write(f"Student Name: {student_name}\n")
                f.write(f"Sheet Code: {sheet_code}\n")
                f.write(f"Total Questions: {total_questions}\n")
                f.write(f"Correct Answers: {correct_count}\n")
                f.write(f"Mark (with reward): {total_mark:.2f}\n")
                f.write(f"Reward Points: {reward_points}\n")
                f.write("\nDetails per question:\n")
                for d in details:
                    f.write(
                        f"  Q{d['question']}: Student Answer: {d['student_answer']} | "
                        f"Correct Answer: {d['correct_answer']} | "
                        f"{'Correct' if d['is_correct'] else 'Incorrect'}\n"
                    )
            if verbose:
                print(f"[EvalMC] Saved evaluation for {student_name} ({student_id}), Sheet: {sheet_code} to {eval_path}")
        except Exception as e:
            if verbose:
                print(f"[EvalMC] Failed to save evaluation for {student_name} ({student_id}): {e}")

        if verbose:
            print(f"[EvalMC] Student {student_name} ({student_id}), Sheet: {sheet_code}, Score: {correct_count}/{total_questions}, Mark: {total_mark:.2f} (Reward: {reward_points})")

    return results

def sync_multichoice_evaluations_to_canvas(
    exam_type=EXAM_TYPE,
    db_path=None,
    api_url=CANVAS_LMS_API_URL,
    api_key=CANVAS_LMS_API_KEY,
    course_id=CANVAS_LMS_COURSE_ID,
    verbose=False
):
    """
    Sync evaluation results from ./Exams/<EXAM_TYPE>/evaluations to Canvas assignment.
    For each TXT file in the evaluations folder, extract student id from filename,
    find Canvas ID from the database, and update the corresponding Canvas assignment score.
    The assignment id is determined by CANVAS_MIDTERM_ASSIGNMENT_ID or CANVAS_FINAL_ASSIGNMENT_ID.
    Additionally, send the evaluation details to each student via Canvas message and ask them to check for grading mistakes/errors.
    """
    # Determine assignment id
    exam_type_lower = str(exam_type).lower()
    if exam_type_lower == "midterm":
        assignment_id = CANVAS_MIDTERM_ASSIGNMENT_ID
    elif exam_type_lower == "final":
        assignment_id = CANVAS_FINAL_ASSIGNMENT_ID
    else:
        assignment_id = None
    if not assignment_id:
        if verbose:
            print(f"[SyncEval] No Canvas assignment id found for exam type '{exam_type}'.")
        else:
            print(f"No Canvas assignment id found for exam type '{exam_type}'.")
        return

    # Load student database
    student_db = {}
    student_name_db = {}
    if db_path and os.path.exists(db_path):
        students = load_database(db_path)
        for s in students:
            sid = str(getattr(s, "Student ID", "")).strip()
            canvas_id = getattr(s, "Canvas ID", None)
            name = getattr(s, "Name", "")
            if sid and canvas_id:
                student_db[sid] = canvas_id
            if sid and name:
                student_name_db[sid] = name

    # Find all evaluation TXT files
    eval_dir = os.path.join("Exams", exam_type, "evaluations")
    eval_files = glob.glob(os.path.join(eval_dir, f"{exam_type}_*_evaluations.txt"))
    if not eval_files:
        if verbose:
            print(f"[SyncEval] No evaluation files found in {eval_dir}.")
        else:
            print(f"No evaluation files found in {eval_dir}.")
        return

    # Connect to Canvas
    canvas = Canvas(api_url, api_key)
    course = canvas.get_course(course_id)
    assignment = course.get_assignment(assignment_id)
    assignment_title = getattr(assignment, "name", str(exam_type).capitalize())

    updated = 0
    skipped = 0
    messaged = 0
    for fname in eval_files:
        # Extract student id from filename
        m = re.match(rf"{exam_type}_([A-Za-z0-9]+)_[^_]+_evaluations\.txt", os.path.basename(fname))
        if not m:
            skipped += 1
            continue
        student_id = m.group(1)
        canvas_id = student_db.get(student_id)
        if not canvas_id:
            skipped += 1
            if verbose:
                print(f"[SyncEval] No Canvas ID found for student id {student_id}. Skipping.")
            continue
        # Extract mark and details from file
        mark = None
        details_lines = []
        with open(fname, "r", encoding="utf-8") as f:
            lines = f.readlines()
            for line in lines:
                if line.lower().startswith("mark"):
                    m_mark = re.search(r"([\d\.]+)", line)
                    if m_mark:
                        mark = float(m_mark.group(1))
                details_lines.append(line.rstrip())
        if mark is None:
            skipped += 1
            if verbose:
                print(f"[SyncEval] No mark found in {fname}. Skipping.")
            continue
        # Update Canvas assignment score
        try:
            submission = assignment.get_submission(canvas_id)
            submission.edit(submission={'posted_grade': mark})
            updated += 1
            if verbose:
                print(f"[SyncEval] Updated Canvas ID {canvas_id} (student id {student_id}) with mark {mark}.")
        except Exception as e:
            skipped += 1
            if verbose:
                print(f"[SyncEval] Failed to update Canvas ID {canvas_id} (student id {student_id}): {e}")
            continue
        # Send evaluation details to student via Canvas message
        try:
            student_name = student_name_db.get(student_id, "")
            subject = f"Kết quả chấm \"{assignment_title}\" của bạn"
            details_text = "\n".join(details_lines)
            message = (
                f"Chào {student_name or 'bạn'},\n\n"
                f"Đây là kết quả chấm \"{assignment_title}\" của bạn vừa được cập nhật trên Canvas.\n"
                f"---\n"
                f"{details_text}\n"
                f"---\n"
                f"Nếu bạn phát hiện có sai sót hoặc nhầm lẫn trong quá trình chấm, vui lòng phản hồi lại cho giảng viên càng sớm càng tốt.\n"
                f"Thông báo này được gửi tự động từ hệ thống."
            )
            canvas.create_conversation(
                recipients=[str(canvas_id)],
                subject=subject,
                body=message,
                force_new=True
            )
            messaged += 1
            if verbose:
                print(f"[SyncEval] Sent evaluation details to Canvas ID {canvas_id} (student id {student_id}).")
        except Exception as e:
            if verbose:
                print(f"[SyncEval] Failed to send message to Canvas ID {canvas_id} (student id {student_id}): {e}")

    if verbose:
        print(f"[SyncEval] Synced {updated} evaluations to Canvas assignment {assignment_id}. Skipped {skipped}. Sent messages to {messaged} students.")
    else:
        print(f"Synced {updated} evaluations to Canvas assignment {assignment_id}. Skipped {skipped}. Sent messages to {messaged} students.")

def calculate_and_print_final_grade_distribution(
    students,
    db_path=None,
    grade_field=None,
    verbose=False
):
    """
    Reimplemented: prefer letter-grade from database if available.
    Falls back to numeric grade distribution if no letter-grade field is found.
    Exports a human-readable report to ./final_grade_distribution.txt and returns a dict with distribution info.
    """

    # Load from db if requested
    if db_path:
        try:
            students = load_database(db_path, verbose=verbose)
        except Exception:
            if verbose:
                print(f"[GradeDist] Failed to load database from {db_path}. Using provided students list.")

    if not students:
        msg = "No students available to calculate distribution."
        if verbose:
            print(f"[GradeDist] {msg}")
        else:
            print(msg)
        return {}

    # 1) Try to find a letter-grade field in DB (most reliable)
    letter_field_candidates = [
        grade_field,
        "Letter Grade", "Final Grade", "Grade", "Grade Letter", "Final Letter",
        "Total Grade", "Total Final Grade", "Điểm chữ", "Letter", "Final Letter Grade"
    ]
    # preserve order, remove None/duplicates
    seen = set()
    letter_candidates = []
    for f in letter_field_candidates:
        if not f:
            continue
        if f in seen:
            continue
        seen.add(f)
        letter_candidates.append(f)

    chosen_letter_field = None
    for f in letter_candidates:
        found = False
        for s in students:
            if hasattr(s, f):
                val = getattr(s, f)
                if val is not None and str(val).strip() != "":
                    found = True
                    break
        if found:
            chosen_letter_field = f
            break

    result = {"total_students": len(students)}

    # helper to write report to file
    def _write_report(text):
        out_path = os.path.join(os.getcwd(), "final_grade_distribution.txt")
        try:
            with open(out_path, "w", encoding="utf-8") as outf:
                outf.write(text)
            if verbose:
                print(f"[GradeDist] Report written to: {out_path}")
        except Exception as e:
            if verbose:
                print(f"[GradeDist] Failed to write report to {out_path}: {e}")
            else:
                print(f"Failed to write report to {out_path}: {e}")

    if chosen_letter_field:
        # Build letter distribution (normalize common variants)
        normalize = lambda x: str(x).strip().upper().replace(" ", "")
        letters = []
        missing = 0
        for s in students:
            v = getattr(s, chosen_letter_field, None)
            if v is None or str(v).strip() == "":
                missing += 1
            else:
                letters.append(normalize(v))
        counter = Counter(letters)

        lines = []
        lines.append(f"Letter-grade distribution (field='{chosen_letter_field}')")
        lines.append("-" * 40)
        lines.append(f"Total students: {len(students)}")
        lines.append(f"With letter grade: {len(letters)}  |  Missing: {missing}")
        lines.append("")
        # Sort by typical order if recognizable, else by frequency
        common_order = ["A+", "A", "A-", "B+", "B", "B-", "C+", "C", "C-", "D+", "D", "D-", "F"]
        ordered = OrderedDict()
        # first add those in common order if present
        for k in common_order:
            if k in counter:
                ordered[k] = counter.pop(k)
        # then the rest by descending count
        for k, v in counter.most_common():
            ordered[k] = v
        for k, cnt in ordered.items():
            pct = cnt / len(letters) * 100 if letters else 0.0
            lines.append(f"  {k:<6}: {cnt:>4} ({pct:5.1f}%)")

        report_text = "\n".join(lines)
        print(report_text)
        _write_report(report_text)

        result.update({
            "letter_field": chosen_letter_field,
            "letter_distribution": dict(ordered),
            "missing_letter_count": missing
        })
        return result

    # 2) No letter field found -> fallback to numeric distribution (as before)
    candidate_fields = [
        grade_field,
        "Final",
        "Final Mark",
        "Final Final Score",
        "Total Final Score",
        "Total",
        "Tổng điểm (Canvas)",
        "Điểm tổng kết",
        "Midterm"
    ]
    # filter/unique
    seen = set()
    fields_to_try = []
    for f in candidate_fields:
        if not f:
            continue
        if f in seen:
            continue
        seen.add(f)
        fields_to_try.append(f)

    chosen_field = None
    for f in fields_to_try:
        found = False
        for s in students:
            if hasattr(s, f):
                val = getattr(s, f)
                if val is None:
                    continue
                try:
                    vv = str(val).strip().replace(",", ".")
                    if vv.endswith("%"):
                        vv = vv.rstrip("%")
                    float(vv)
                    found = True
                    break
                except Exception:
                    continue
        if found:
            chosen_field = f
            break

    if not chosen_field:
        if verbose:
            print("[GradeDist] No suitable grade field found. Provide grade_field or ensure students have numeric 'Final' or similar attribute.")
        else:
            print("No suitable grade field found.")
        return {}

    grades = []
    invalid = 0
    for s in students:
        val = getattr(s, chosen_field, None)
        if val is None or (isinstance(val, str) and not val.strip()):
            invalid += 1
            continue
        try:
            v = str(val).strip().replace(",", ".")
            if v.endswith("%"):
                v = v.rstrip("%")
                num = float(v)
                # percent -> convert to 0-10 scale if looks like 0-100
                if 0 <= num <= 100:
                    num = num / 10.0
            else:
                num = float(v)
                # if number appears in 0-100 but >10, treat as percent-like and scale to 0-10
                if num > 10 and num <= 100:
                    num = num / 10.0
            if math.isnan(num) or math.isinf(num):
                raise ValueError
            grades.append(num)
        except Exception:
            invalid += 1
            continue

    valid_count = len(grades)
    if valid_count == 0:
        if verbose:
            print(f"[GradeDist] No numeric grades found in field '{chosen_field}'.")
        else:
            print("No numeric grades found.")
        return {}

    mean = statistics.mean(grades)
    median = statistics.median(grades)
    stdev = statistics.pstdev(grades) if valid_count > 1 else 0.0
    minimum = min(grades)
    maximum = max(grades)

    # letter buckets
    buckets = OrderedDict([
        ("A+ [9.0-10.0]", lambda g: 9.0 <= g <= 10.0),
        ("A  [8.5-<9.0]", lambda g: 8.5 <= g < 9.0),
        ("B+ [8.0-<8.5]", lambda g: 8.0 <= g < 8.5),
        ("B  [7.0-<8.0]", lambda g: 7.0 <= g < 8.0),
        ("C+ [6.5-<7.0]", lambda g: 6.5 <= g < 7.0),
        ("C  [5.5-<6.5]", lambda g: 5.5 <= g < 6.5),
        ("D+ [5.0-<5.5]", lambda g: 5.0 <= g < 5.5),
        ("D  [4.0-<5.0]", lambda g: 4.0 <= g < 5.0),
        ("F  [<4.0]",     lambda g: g < 4.0),
    ])
    bucket_counts = {k: 0 for k in buckets}
    for g in grades:
        placed = False
        for k, fn in buckets.items():
            try:
                if fn(g):
                    bucket_counts[k] += 1
                    placed = True
                    break
            except Exception:
                continue
        if not placed:
            if g < 4.0:
                bucket_counts["F  [<4.0]"] += 1
            elif g >= 9.0:
                bucket_counts["A+ [9.0-10.0]"] += 1
            else:
                # nearest fallback
                if g < 5.0:
                    bucket_counts["D  [4.0-<5.0]"] += 1
                else:
                    bucket_counts["C  [5.5-<6.5]"] += 1

    # numeric bins
    bin_edges = [0.0, 4.0, 5.0, 6.5, 8.0, 10.0]
    bin_labels = [f"{bin_edges[i]} - {bin_edges[i+1]}" for i in range(len(bin_edges)-1)]
    bin_counts = [0] * (len(bin_edges) - 1)
    for g in grades:
        for i in range(len(bin_edges)-1):
            lo, hi = bin_edges[i], bin_edges[i+1]
            if (g >= lo and (g < hi or (i == len(bin_edges)-2 and g <= hi))):
                bin_counts[i] += 1
                break

    # build report lines
    lines = []
    header = f"Final grade distribution (field='{chosen_field}')"
    lines.append(header)
    lines.append("-" * len(header))
    lines.append(f"Total students: {len(students)}")
    lines.append(f"Valid numeric grades: {valid_count}  |  Invalid or missing: {invalid}")
    lines.append("")
    lines.append("Statistics:")
    lines.append(f"  Mean   : {mean:.2f}")
    lines.append(f"  Median : {median:.2f}")
    lines.append(f"  StdDev : {stdev:.2f}")
    lines.append(f"  Min    : {minimum:.2f}")
    lines.append(f"  Max    : {maximum:.2f}")
    lines.append("")
    lines.append("Letter grade buckets:")
    for k, cnt in bucket_counts.items():
        pct = cnt / valid_count * 100
        lines.append(f"  {k:<16}: {cnt:>4} ({pct:5.1f}%)")
    lines.append("")
    lines.append("Numeric bins:")
    max_bar = 40
    max_count = max(bin_counts) if bin_counts else 1
    for label, cnt in zip(bin_labels, bin_counts):
        pct = cnt / valid_count * 100
        bar_len = int((cnt / max_count) * max_bar) if max_count > 0 else 0
        bar = "#" * bar_len
        lines.append(f"  {label:<12}: {cnt:>4} ({pct:5.1f}%) |{bar}")

    # optionally show top students when verbose
    if verbose:
        try:
            sorted_students = sorted(
                [(getattr(s, chosen_field, None), getattr(s, "Name", getattr(s, "name", "")), s)
                    for s in students],
                key=lambda x: (float(str(x[0]).replace(",", ".")) if x[0] is not None and str(x[0]).strip() != "" else -999),
                reverse=True
            )
            lines.append("")
            lines.append("Top 5 students (by chosen field):")
            shown = 0
            for val, name, s in sorted_students:
                try:
                    v = float(str(val).replace(",", "."))
                except Exception:
                    continue
                lines.append(f"  {name[:30]:<30} {v:5.2f}")
                shown += 1
                if shown >= 5:
                    break
        except Exception:
            pass

    report_text = "\n".join(lines)
    print(report_text)
    _write_report(report_text)

    result.update({
        "field": chosen_field,
        "valid_count": valid_count,
        "invalid_count": invalid,
        "stats": {"mean": mean, "median": median, "stdev": stdev, "min": minimum, "max": maximum},
        "buckets": bucket_counts,
        "bins": {"labels": bin_labels, "counts": bin_counts}
    })
    return result

def export_roster_to_csv(students, file_path=None, verbose=False):
    """
    Export the classroom roster to a CSV file with one column: identifier.
    The identifier is set to the student's name.

    Args:
        students (list): List of Student objects.
        file_path (str, optional): Path to the output CSV file. Defaults to "classroom_roster.csv".
        verbose (bool): If True, print a message after exporting.

    Returns:
        None
    """
    if file_path is None:
        file_path = "classroom_roster.csv"

    # Ensure the directory exists
    dir_path = os.path.dirname(file_path)
    if dir_path:
        os.makedirs(dir_path, exist_ok=True)
    if DRY_RUN:
        print(f"[RosterExport] Dry run: would export roster to {file_path}")
        return

    with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        # Write header
        writer.writerow(["identifier"])
        # Write each student's identifier (name)
        for student in students:
            identifier = getattr(student, 'Name', '') or ''
            writer.writerow([identifier])

    # Print the path to the exported file if successful
    print(f"Roster exported to {file_path}")

    if verbose:
        print(f"Roster exported to {file_path}")
    append_run_report(
        "export-roster",
        outputs=file_path,
        verbose=verbose,
    )
_override_grades_cache = None
_override_grades_cache_path = None


def _normalize_vietnamese_name(value):
    if not value:
        return ""
    value = unicodedata.normalize("NFD", str(value))
    value = "".join(c for c in value if not unicodedata.combining(c))
    return value.lower().strip()


def _normalize_student_id(value):
    if value is None:
        return ""
    if isinstance(value, float):
        if pd.isna(value):
            return ""
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    raw = unicodedata.normalize("NFKC", str(value)).strip()
    if not raw:
        return ""
    raw = re.sub(r"\\s+", "", raw)
    if re.match(r"^\\d+\\.0+$", raw):
        return raw.split(".", 1)[0]
    if re.match(r"^\\d+$", raw):
        return raw
    return raw


def _looks_like_student_id(value):
    raw = _normalize_student_id(value)
    return bool(raw) and raw.isdigit() and len(raw) >= 5


def _infer_student_id_from_email(email):
    if not email:
        return ""
    local_part = str(email).split("@", 1)[0].strip()
    if _looks_like_student_id(local_part):
        return _normalize_student_id(local_part)
    digits = re.sub(r"\\D+", "", local_part)
    if _looks_like_student_id(digits):
        return digits
    return ""


def _get_student_sort_key(student, method=None):
    method = (method or STUDENT_SORT_METHOD or "first_last").strip().lower()
    name = str(getattr(student, "Name", "")).strip()
    parts = [p for p in name.split() if p]
    first = parts[-1] if parts else ""
    family = " ".join(parts[:-1]) if len(parts) > 1 else ""
    sid = _normalize_student_id(getattr(student, "Student ID", "")) or ""
    if method in ("id", "student_id"):
        return (sid, _normalize_vietnamese_name(name))
    if method in ("last_first", "family_first"):
        return (
            _normalize_vietnamese_name(family),
            _normalize_vietnamese_name(first),
            _normalize_vietnamese_name(name),
        )
    return (
        _normalize_vietnamese_name(first),
        _normalize_vietnamese_name(family),
        _normalize_vietnamese_name(name),
    )


def _looks_like_name(value):
    if value is None:
        return False
    raw = str(value).strip()
    if not raw:
        return False
    return any(ch.isalpha() for ch in raw)


def _is_grade_provided(value):
    if value is None:
        return False
    if isinstance(value, float) and pd.isna(value):
        return False
    if isinstance(value, str) and not value.strip():
        return False
    return True


def _coerce_score(value):
    if not _is_grade_provided(value):
        return None
    if isinstance(value, (int, float)) and not pd.isna(value):
        return float(value)
    if isinstance(value, str):
        raw = value.strip().replace(",", ".")
        try:
            return float(raw)
        except ValueError:
            return None
    return None


def load_override_grades(file_path="override_grades.xlsx", verbose=False):
    if not os.path.exists(file_path):
        if verbose:
            print(f"[OverrideGrades] File not found: {file_path}")
        return {"by_id": {}, "by_name": {}}

    # Header names are matched by aliases; order is not required.
    df = pd.read_excel(file_path)
    header_aliases = {
        "stt": ["stt", "no", "stt."],
        "student_id": ["m\u00e3 sinh vi\u00ean", "mssv", "m\u00e3 sv", "student id", "studentid", "id"],
        "full_name": ["h\u1ecd v\u00e0 t\u00ean", "h\u1ecd t\u00ean", "t\u00ean", "ho va ten", "ho ten", "name", "full name"],
        "cc": ["cc", "chuyen can", "chuy\u00ean c\u1ea7n", "attendance"],
        "gk": ["gk", "giua ky", "gi\u1eefa k\u1ef3", "midterm"],
        "ck": ["ck", "cuoi ky", "cu\u1ed1i k\u1ef3", "final", "final exam"],
        "reason": ["l\u00fd do", "ly do", "reason", "note", "notes"],
    }

    def normalize_header(value):
        raw = unicodedata.normalize("NFKC", str(value)).strip().lower()
        raw = unicodedata.normalize("NFD", raw)
        raw = "".join(c for c in raw if not unicodedata.combining(c))
        raw = re.sub(r"\\s+", " ", raw)
        return raw

    normalized_columns = {normalize_header(col): col for col in df.columns}

    def find_column(aliases):
        for alias in aliases:
            key = normalize_header(alias)
            if key in normalized_columns:
                return normalized_columns[key]
        return None

    col_id = find_column(header_aliases["student_id"])
    col_name = find_column(header_aliases["full_name"])
    col_cc = find_column(header_aliases["cc"])
    col_gk = find_column(header_aliases["gk"])
    col_ck = find_column(header_aliases["ck"])
    col_reason = find_column(header_aliases["reason"])

    if not col_id and not col_name:
        raise ValueError("[OverrideGrades] Missing required column: M\u00e3 Sinh Vi\u00ean or H\u1ecd v\u00e0 T\u00ean.")
    if not col_cc and not col_gk and not col_ck:
        raise ValueError("[OverrideGrades] Missing required score columns: CC, GK, or CK.")

    overrides = {"by_id": {}, "by_name": {}}
    for _, row in df.iterrows():
        raw_id = row.get(col_id) if col_id and _is_grade_provided(row.get(col_id)) else ""
        raw_name = row.get(col_name) if col_name and _is_grade_provided(row.get(col_name)) else ""
        student_id = _normalize_student_id(raw_id)
        full_name = str(raw_name).strip() if raw_name != "" else ""
        # Guard against user-swapped ID/name columns in the input file.
        if _looks_like_name(student_id) and _looks_like_student_id(full_name):
            student_id, full_name = _normalize_student_id(full_name), str(raw_id).strip()
            if verbose:
                print("[OverrideGrades] Detected swapped ID/name cells; swapping for this row.")
        if not student_id and not full_name:
            continue
        entry = {
            "CC": _coerce_score(row.get(col_cc)) if col_cc else None,
            "CK": _coerce_score(row.get(col_ck)) if col_ck else None,
            "GK": _coerce_score(row.get(col_gk)) if col_gk else None,
            "reason": str(row.get(col_reason)).strip() if col_reason and _is_grade_provided(row.get(col_reason)) else "",
        }
        if student_id:
            overrides["by_id"][student_id] = entry
        if full_name:
            overrides["by_name"][_normalize_vietnamese_name(full_name)] = entry

    if verbose:
        sample_ids = list(overrides["by_id"].keys())[:3]
        sample_names = list(overrides["by_name"].keys())[:3]
        print(
            "[OverrideGrades] Loaded entries - "
            f"by_id: {len(overrides['by_id'])}, by_name: {len(overrides['by_name'])}."
        )
        if sample_ids:
            print(f"[OverrideGrades] Sample IDs: {sample_ids}")
        if sample_names:
            print(f"[OverrideGrades] Sample names: {sample_names}")

    return overrides


def get_override_grades(file_path="override_grades.xlsx", verbose=False):
    global _override_grades_cache
    global _override_grades_cache_path
    if _override_grades_cache is None or _override_grades_cache_path != file_path:
        _override_grades_cache = load_override_grades(file_path=file_path, verbose=verbose)
        _override_grades_cache_path = file_path
    return _override_grades_cache


def apply_override_grades(student, CC, CK, GK, overrides):
    entry = None
    student_id = _normalize_student_id(getattr(student, "Student ID", ""))
    student_name = str(getattr(student, "Name", "")).strip()

    if overrides.get("by_id") and student_id in overrides["by_id"]:
        entry = overrides["by_id"][student_id]
    elif overrides.get("by_name") and student_name:
        entry = overrides["by_name"].get(_normalize_vietnamese_name(student_name))

    # Only override fields explicitly provided in override_grades.xlsx.
    if not entry:
        return CC, CK, GK, ""

    override_reason = entry.get("reason", "")
    if _is_grade_provided(entry.get("CC")):
        CC = entry["CC"]
    if _is_grade_provided(entry.get("CK")):
        CK = entry["CK"]
    if _is_grade_provided(entry.get("GK")):
        GK = entry["GK"]

    return CC, CK, GK, override_reason


def load_override_grades_to_database(override_file="override_grades.xlsx", db_path=None, verbose=False):
    """
    Load override grades from an Excel file and persist them into the student database.
    Stores override values on each matching student as:
      - Override CC
      - Override GK
      - Override CK
      - Override Reason
    """
    if not db_path:
        db_path = get_default_db_path()
    if not os.path.exists(db_path):
        if verbose:
            print(f"[OverrideGrades] Database not found: {db_path}")
        else:
            print(f"Database not found: {db_path}")
        return 0

    overrides = load_override_grades(file_path=override_file, verbose=verbose)
    if not overrides.get("by_id") and not overrides.get("by_name"):
        if verbose:
            print("[OverrideGrades] No override entries found.")
        else:
            print("No override entries found.")
        return 0

    students = load_database(db_path, verbose=verbose)
    if not students:
        if verbose:
            print("[OverrideGrades] No students loaded from database.")
        else:
            print("No students loaded from database.")
        return 0

    sid_map = {}
    name_map = {}
    for s in students:
        sid = str(getattr(s, "Student ID", "")).strip()
        name = str(getattr(s, "Name", "")).strip()
        if sid:
            sid_map[sid] = s
        if name:
            name_map[_normalize_vietnamese_name(name)] = s

    updated = 0

    def apply_entry(student, entry):
        nonlocal updated
        if _is_grade_provided(entry.get("CC")):
            setattr(student, "Override CC", entry.get("CC"))
        if _is_grade_provided(entry.get("GK")):
            setattr(student, "Override GK", entry.get("GK"))
        if _is_grade_provided(entry.get("CK")):
            setattr(student, "Override CK", entry.get("CK"))
        if entry.get("reason") is not None:
            setattr(student, "Override Reason", entry.get("reason", ""))
        updated += 1

    for sid, entry in overrides.get("by_id", {}).items():
        student = sid_map.get(str(sid).strip())
        if student:
            apply_entry(student, entry)

    for name_key, entry in overrides.get("by_name", {}).items():
        student = name_map.get(name_key)
        if student:
            apply_entry(student, entry)

    save_database(students, db_path, verbose=verbose, audit_source="override-grades")
    if verbose:
        print(f"[OverrideGrades] Applied overrides to {updated} student(s).")
    else:
        print(f"Applied overrides to {updated} student(s).")
    return updated


# Course calendar utilities

def _parse_calendar_date(value):
    return datetime.strptime(value.strip(), "%Y-%m-%d").date()


def _parse_calendar_time(value):
    return datetime.strptime(value.strip(), "%H:%M").time()


def _parse_calendar_time_range(value):
    parts = value.split("-")
    if len(parts) != 2:
        raise ValueError("Time range must be HH:MM-HH:MM")
    start = _parse_calendar_time(parts[0])
    end = _parse_calendar_time(parts[1])
    if end <= start:
        raise ValueError("Session end time must be after start time.")
    return start, end


def _parse_calendar_session_line(line):
    raw = line.strip()
    if not raw:
        return None
    if raw.lower().startswith("session"):
        _, _, raw = raw.partition(":")
        raw = raw.strip()
    if "|" in raw:
        parts = [p.strip() for p in raw.split("|") if p.strip()]
    else:
        parts = [p.strip() for p in raw.split(",") if p.strip()]
    if not parts:
        return None
    first = parts[0]
    tokens = first.split()
    if len(tokens) < 2:
        raise ValueError(f"Invalid session line: {line}")
    date_str = tokens[0]
    time_range = tokens[1]
    date_val = _parse_calendar_date(date_str)
    start_time, end_time = _parse_calendar_time_range(time_range)
    location = parts[1] if len(parts) > 1 else ""
    title = parts[2] if len(parts) > 2 else "Class session"
    return {
        "date": date_val,
        "start_time": start_time,
        "end_time": end_time,
        "location": location,
        "title": title,
    }


def _parse_calendar_bool(value):
    if value is None:
        return False
    return str(value).strip().lower() in ("1", "true", "yes", "y")


def get_vietnam_fixed_holidays(year):
    return [
        datetime(year, 1, 1).date(),
        datetime(year, 4, 30).date(),
        datetime(year, 5, 1).date(),
        datetime(year, 9, 2).date(),
    ]


def get_vietnam_holidays(years, extra_dates=None):
    holidays = set()
    for year in years:
        for day in get_vietnam_fixed_holidays(year):
            holidays.add(day)
    for day in extra_dates or []:
        holidays.add(day)
    return holidays


def fetch_vietnam_holidays_ai(years, method=None, verbose=False):
    if not method:
        return []
    if method not in ALL_AI_METHODS:
        if verbose:
            print(f"[Calendar] Unknown AI method for holidays: {method}.")
        return []
    years_list = sorted({int(y) for y in years if y})
    if not years_list:
        return []
    prompt = (
        "List Vietnam public holidays for the years "
        + ", ".join(str(y) for y in years_list)
        + ". Include lunar holidays (Tet, Hung Vuong) and fixed holidays. "
        "Return ONLY dates in YYYY-MM-DD format, one per line."
    )
    try:
        response = refine_text_with_ai(prompt, method=method, verbose=verbose)
    except Exception as exc:
        if verbose:
            print(f"[Calendar] Failed to fetch holidays via AI: {exc}")
        return []
    dates = re.findall(r"\b\d{4}-\d{2}-\d{2}\b", response or "")
    parsed = []
    for raw in dates:
        try:
            parsed.append(_parse_calendar_date(raw))
        except Exception:
            continue
    if verbose:
        print(f"[Calendar] AI returned {len(parsed)} holiday date(s).")
    return parsed


def _resolve_calendar_input_path(path):
    if not path:
        return None
    if os.path.exists(path):
        return path
    if os.path.isabs(path):
        return path
    package_root = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
    candidate = os.path.join(package_root, path)
    if os.path.exists(candidate):
        return candidate
    return path


def parse_course_calendar_txt(path, verbose=False):
    sessions = []
    holidays = []
    weeks = 15
    extra_week = None
    course_code = ""
    course_name = ""
    resolved_path = _resolve_calendar_input_path(path)
    with open(resolved_path, "r", encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            lowered = line.lower()
            if lowered.startswith("course_code") or lowered.startswith("course code"):
                _, _, value = line.partition(":")
                value = value.strip() or line.split("=", 1)[-1].strip()
                course_code = value
                continue
            if lowered.startswith("course_name") or lowered.startswith("course name"):
                _, _, value = line.partition(":")
                value = value.strip() or line.split("=", 1)[-1].strip()
                course_name = value
                continue
            if lowered.startswith("course"):
                _, _, value = line.partition(":")
                value = value.strip()
                if " - " in value:
                    code_part, name_part = value.split(" - ", 1)
                    course_code = course_code or code_part.strip()
                    course_name = course_name or name_part.strip()
                else:
                    course_name = course_name or value
                continue
            if lowered.startswith("weeks") or lowered.startswith("official_weeks"):
                _, _, value = line.partition(":")
                value = value.strip() or line.split("=", 1)[-1].strip()
                if value.isdigit():
                    weeks = int(value)
                continue
            if lowered.startswith("extra_week") or lowered.startswith("extra"):
                _, _, value = line.partition(":")
                value = value.strip() or line.split("=", 1)[-1].strip()
                extra_week = _parse_calendar_bool(value)
                continue
            if lowered.startswith("holiday") or lowered.startswith("extra_holiday") or lowered.startswith("unofficial_holiday"):
                _, _, value = line.partition(":")
                value = value.strip() or line.split("=", 1)[-1].strip()
                for part in value.split(","):
                    part = part.strip()
                    if part:
                        holidays.append(_parse_calendar_date(part))
                continue
            session = _parse_calendar_session_line(line)
            if session:
                sessions.append(session)
    if verbose:
        print(f"[Calendar] Loaded {len(sessions)} session(s) from {resolved_path}")
    return {
        "sessions": sessions,
        "weeks": weeks,
        "extra_week": extra_week,
        "holidays": holidays,
        "course_code": course_code,
        "course_name": course_name,
    }


def prompt_course_calendar_manual(course_code_default=None, course_name_default=None, verbose=False):
    course_code_prompt = "Course code"
    if course_code_default:
        course_code_prompt += f" [{course_code_default}]"
    course_code_prompt += ": "
    course_code = input(course_code_prompt).strip() or (course_code_default or "")
    course_name_prompt = "Course name"
    if course_name_default:
        course_name_prompt += f" [{course_name_default}]"
    course_name_prompt += ": "
    course_name = input(course_name_prompt).strip() or (course_name_default or "")
    weeks_raw = input("Number of official weeks [15]: ").strip()
    weeks = int(weeks_raw) if weeks_raw.isdigit() else 15
    extra_raw = input("Allow one make-up week if holidays occur? (y/n) [y]: ").strip().lower()
    extra_week = False if extra_raw in ("n", "no") else True
    count_raw = input("Number of sessions in the first week: ").strip()
    if not count_raw.isdigit():
        raise ValueError("Please provide a valid number of sessions.")
    count = int(count_raw)
    sessions = []
    for idx in range(1, count + 1):
        date_raw = input(f"Session {idx} date (YYYY-MM-DD): ").strip()
        start_raw = input(f"Session {idx} start time (HH:MM): ").strip()
        end_raw = input(f"Session {idx} end time (HH:MM): ").strip()
        location = input(f"Session {idx} location (optional): ").strip()
        title = input(f"Session {idx} title (optional) [Class session]: ").strip() or "Class session"
        start_time = _parse_calendar_time(start_raw)
        end_time = _parse_calendar_time(end_raw)
        if end_time <= start_time:
            raise ValueError("Session end time must be after start time.")
        sessions.append({
            "date": _parse_calendar_date(date_raw),
            "start_time": start_time,
            "end_time": end_time,
            "location": location,
            "title": title,
        })
    holidays = []
    default_method = DEFAULT_AI_METHOD if DEFAULT_AI_METHOD in ALL_AI_METHODS else ""
    default_label = default_method or "none"
    ai_raw = input(
        f"Fetch holidays via AI? (gemini/huggingface/local/none) [default: {default_label}]: "
    ).strip().lower()
    if not ai_raw:
        ai_raw = default_method
    if ai_raw and ai_raw != "none":
        ai_holidays = ai_raw
    add_holiday = input("Add extra holiday dates? (y/n) [n]: ").strip().lower()
    if add_holiday in ("y", "yes"):
        while True:
            holiday_raw = input("Holiday date (YYYY-MM-DD, blank to stop): ").strip()
            if not holiday_raw:
                break
            holidays.append(_parse_calendar_date(holiday_raw))
    if verbose:
        print(f"[Calendar] Collected {len(sessions)} session(s) from manual input")
    return {
        "sessions": sessions,
        "weeks": weeks,
        "extra_week": extra_week,
        "holidays": holidays,
        "course_code": course_code,
        "course_name": course_name,
    }


def build_course_calendar(first_week_sessions, weeks=15, extra_week=None, holidays=None, verbose=False):
    if not first_week_sessions:
        raise ValueError("No first-week sessions provided.")
    if weeks <= 0:
        raise ValueError("Number of weeks must be positive.")
    allow_extra_week = True if extra_week is None else bool(extra_week)
    total_weeks = weeks
    years = set()
    for session in first_week_sessions:
        years.add(session["date"].year)
        years.add((session["date"] + timedelta(days=7 * (weeks + 1))).year)
    holiday_set = get_vietnam_holidays(years, holidays)
    events = []
    skipped = 0
    for session in first_week_sessions:
        for week_idx in range(weeks):
            day = session["date"] + timedelta(days=7 * week_idx)
            start_dt = datetime.combine(day, session["start_time"])
            end_dt = datetime.combine(day, session["end_time"])
            title = session.get("title", "Class session")
            if day in holiday_set:
                if verbose:
                    print(f"[Calendar] Skipping holiday session on {day}")
                skipped += 1
                title = f"{title} [cancel]"
            events.append({
                "date": day,
                "start_dt": start_dt,
                "end_dt": end_dt,
                "location": session.get("location", ""),
                "title": title,
                "week": week_idx + 1,
            })
    if skipped and allow_extra_week:
        makeup_week = weeks
        total_weeks += 1
        for session in first_week_sessions:
            day = session["date"] + timedelta(days=7 * makeup_week)
            if day in holiday_set:
                if verbose:
                    print(f"[Calendar] Skipping holiday make-up session on {day}")
                continue
            start_dt = datetime.combine(day, session["start_time"])
            end_dt = datetime.combine(day, session["end_time"])
            events.append({
                "date": day,
                "start_dt": start_dt,
                "end_dt": end_dt,
                "location": session.get("location", ""),
                "title": session.get("title", "Class session"),
                "week": makeup_week + 1,
            })
    events.sort(key=lambda e: (e["start_dt"], e["title"]))
    if verbose:
        extra_note = " (make-up week added)" if skipped and allow_extra_week else ""
        print(f"[Calendar] Built {len(events)} session(s) across {total_weeks} week(s){extra_note}")
    return events


def _ics_escape(value):
    if value is None:
        return ""
    return str(value).replace("\\", "\\\\").replace(";", "\;").replace(",", "\,")

def _calendar_event_title(event, course_title=None):
    base_title = event.get("title") or "Class session"
    if course_title:
        if base_title.startswith(course_title):
            return base_title
        return f"{course_title} - {base_title}"
    return base_title


def export_course_calendar_ics(events, output_path, course_name=None, verbose=False):
    now = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//course//calendar//EN",
    ]
    for event in events:
        uid_source = f"{event['start_dt'].isoformat()}-{event.get('location','')}-{event.get('title','')}"
        uid = hashlib.md5(uid_source.encode("utf-8")).hexdigest() + "@course"
        summary = _calendar_event_title(event, course_name)
        lines.extend([
            "BEGIN:VEVENT",
            f"UID:{uid}",
            f"DTSTAMP:{now}",
            f"DTSTART:{event['start_dt'].strftime('%Y%m%dT%H%M%S')}",
            f"DTEND:{event['end_dt'].strftime('%Y%m%dT%H%M%S')}",
            f"SUMMARY:{_ics_escape(summary)}",
        ])
        location = event.get("location")
        if location:
            lines.append(f"LOCATION:{_ics_escape(location)}")
        lines.append("END:VEVENT")
    lines.append("END:VCALENDAR")
    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    if verbose:
        print(f"[Calendar] Wrote iCal to {output_path}")
    return output_path


def export_course_calendar_txt(events, output_path, course_name=None, verbose=False):
    lines = []
    if course_name:
        lines.append(f"Course: {course_name}")
    lines.append(f"Total sessions: {len(events)}")
    lines.append("")
    for event in events:
        date_str = event["date"].strftime("%Y-%m-%d")
        time_str = f"{event['start_dt'].strftime('%H:%M')}-{event['end_dt'].strftime('%H:%M')}"
        title = _calendar_event_title(event, course_name)
        location = event.get("location")
        loc_text = f" @ {location}" if location else ""
        lines.append(f"{date_str} (Week {event['week']}): {time_str} - {title}{loc_text}")
    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    if verbose:
        print(f"[Calendar] Wrote TXT to {output_path}")
    return output_path


def export_course_calendar_markdown(events, output_path, course_name=None, verbose=False):
    lines = []
    if course_name:
        lines.append(f"# {course_name} calendar")
    else:
        lines.append("# Course calendar")
    lines.append("")
    lines.append("| Date | Week | Time | Title | Location |")
    lines.append("| --- | --- | --- | --- | --- |")
    for event in events:
        date_str = event["date"].strftime("%Y-%m-%d")
        time_str = f"{event['start_dt'].strftime('%H:%M')}-{event['end_dt'].strftime('%H:%M')}"
        title = _calendar_event_title(event, course_name)
        location = event.get("location") or ""
        lines.append(f"| {date_str} | {event['week']} | {time_str} | {title} | {location} |")
    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    if verbose:
        print(f"[Calendar] Wrote Markdown to {output_path}")
    return output_path


def export_course_calendar_all(events, output_dir=None, base_name=None, course_name=None, verbose=False):
    output_dir = output_dir or os.getcwd()
    base_name = base_name or "course_calendar"
    txt_path = os.path.join(output_dir, f"{base_name}.txt")
    md_path = os.path.join(output_dir, f"{base_name}.md")
    ics_path = os.path.join(output_dir, f"{base_name}.ics")
    export_course_calendar_txt(events, txt_path, course_name=course_name, verbose=verbose)
    export_course_calendar_markdown(events, md_path, course_name=course_name, verbose=verbose)
    export_course_calendar_ics(events, ics_path, course_name=course_name, verbose=verbose)
    return {"txt": txt_path, "md": md_path, "ics": ics_path}


def build_and_export_course_calendar(input_path=None, weeks=None, extra_week=None, output_dir=None, base_name=None, course_code=None, course_name=None, extra_holidays=None, verbose=False):
    if input_path:
        payload = parse_course_calendar_txt(input_path, verbose=verbose)
    else:
        payload = prompt_course_calendar_manual(
            course_code_default=course_code or COURSE_CODE or get_cached_course_code(),
            course_name_default=course_name or COURSE_NAME,
            verbose=verbose,
        )
    sessions = payload.get("sessions", [])
    file_weeks = payload.get("weeks", 15)
    file_extra = payload.get("extra_week", False)
    holidays = payload.get("holidays", [])
    if extra_holidays:
        holidays.extend(extra_holidays)
    if not course_code:
        course_code = payload.get("course_code") or COURSE_CODE or get_cached_course_code() or ""
    if not course_name:
        course_name = payload.get("course_name") or COURSE_NAME or ""
    if weeks is None:
        weeks = file_weeks
    if extra_week is None:
        extra_week = file_extra
    course_code = str(course_code).strip()
    course_name = str(course_name).strip()
    if not course_code:
        raise ValueError("Course code is required. Provide --calendar-course-code, COURSE_CODE in config, or include course_code: in the input file.")
    if not course_name:
        raise ValueError("Course name is required. Provide --calendar-course-name, COURSE_NAME in config, or include course_name: in the input file.")
    course_title = f"{course_code} - {course_name}"
    if DEFAULT_AI_METHOD in ALL_AI_METHODS:
        years = {s["date"].year for s in sessions}
        years.add(datetime.now().year)
        holidays.extend(fetch_vietnam_holidays_ai(years, method=DEFAULT_AI_METHOD, verbose=verbose))
    events = build_course_calendar(
        first_week_sessions=sessions,
        weeks=weeks,
        extra_week=extra_week,
        holidays=holidays,
        verbose=verbose,
    )
    return export_course_calendar_all(events, output_dir=output_dir, base_name=base_name, course_name=course_title, verbose=verbose)
